"""
PP loader — reads a BOM xlsx (the "Source data for Material BOM" LTMC
template populated with user data) into a normalized internal structure.

Template row layout (verified against the SAP S/4HANA 2025 standard
template AND the customer's BOM_PHASE_1.xlsx)
-----------------------------------------------------------------------
Row 1  Title: "Source Data for Migration Object: Material BOM"
Row 2  Version banner
Row 3  (blank)
Row 4  SAP table name (S_BOM_HEADER, S_BOM_ITEM, S_STPU, S_GLOBAL_DEP, ...)
Row 5  SAP field codes (MATNR, WERKS, STLAN, ...) — AUTHORITATIVE
Row 6  ETE format spec (e.g. "ETE;80;0;C;80;0" → text/length=80/decimals=0)
Row 7  Field-group label ("Key", "Validity", "BOM and Alternative Text")
Row 8  Long human-readable description; trailing "*" on the first line
       indicates a mandatory field.
Row 9+ Data rows

The validator reads SAP codes from row 5; everything else (max length,
mandatory flag, friendly label) comes from the rulebook which is
populated programmatically from the LTMC template at process start.

Cell value normalization (consistent with mm_loader)
----------------------------------------------------
- None / empty string / pure whitespace → None
- Strings: stripped, NBSP → space
- Numbers kept as native float/int; ".0" trimming happens at format
  time (validator + generator), not here.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


PP_SHEETS_ORDER: tuple[str, ...] = (
    "BOM Header",
    "BOM Item",
    "BOM Subitem",
    "Global Dependency",
    "Local Dependency",
    "Local Dependency Description",
    "Documentation of Dependency",
    "Sources of Local Dependency",
    "BOM Item Document Assignment",
    "BOM Header Document Assignment",
)

_CODE_ROW = 5
_DATA_START_ROW = 9


@dataclass
class LoadedRow:
    excel_row: int
    values: dict[str, Any]

    def get(self, sap_code: str, default: Any = None) -> Any:
        return self.values.get(sap_code.upper(), default)


@dataclass
class LoadedSheet:
    name: str
    sap_fields: list[str] = field(default_factory=list)
    rows: list[LoadedRow] = field(default_factory=list)
    skipped_blank_rows: int = 0


@dataclass
class LoadedBom:
    filename: str
    sheets: dict[str, LoadedSheet] = field(default_factory=dict)


def _norm_cell(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.replace("\u00a0", " ").strip()
        return s if s else None
    return v


def _row_is_blank(values: dict[str, Any]) -> bool:
    return all(v is None for v in values.values())


def _read_sheet(ws, sheet_name: str) -> LoadedSheet:
    sheet = LoadedSheet(name=sheet_name)
    if ws.max_row < _CODE_ROW:
        return sheet

    # Read row 5 (SAP codes). We use random access here because we
    # need to know the rightmost populated column; iter_rows would
    # require a second pass.
    header_cells = [ws.cell(row=_CODE_ROW, column=c).value
                    for c in range(1, ws.max_column + 1)]
    last_col = 0
    for i, code in enumerate(header_cells):
        if code is not None and str(code).strip():
            last_col = i + 1
    if last_col == 0:
        return sheet

    col_to_sap: list[tuple[int, str]] = []
    sap_fields: list[str] = []
    for col_idx in range(last_col):
        code = header_cells[col_idx]
        if code is None or not str(code).strip():
            continue
        sap = str(code).strip().upper()
        col_to_sap.append((col_idx, sap))
        sap_fields.append(sap)
    sheet.sap_fields = sap_fields

    # Stream data rows starting at row 9. Streaming keeps memory
    # bounded for huge files (the customer's BOM_PHASE_1.xlsx has
    # 31582 BOM Item rows × 64 columns).
    data_rows: list[LoadedRow] = []
    for excel_row, row_values in enumerate(
        ws.iter_rows(min_row=_DATA_START_ROW, values_only=True), start=_DATA_START_ROW
    ):
        values: dict[str, Any] = {}
        for col_idx, sap in col_to_sap:
            cell = row_values[col_idx] if col_idx < len(row_values) else None
            values[sap] = _norm_cell(cell)
        data_rows.append(LoadedRow(excel_row=excel_row, values=values))

    # Drop trailing blank rows (Excel template leftovers).
    skipped = 0
    while data_rows and _row_is_blank(data_rows[-1].values):
        data_rows.pop()
        skipped += 1
    sheet.rows = data_rows
    sheet.skipped_blank_rows = skipped
    return sheet


def load_bom(xlsx_path: str | Path, filename: str | None = None) -> LoadedBom:
    """Load a BOM xlsx file from disk. Raises ValueError if the file has
    no recognizable BOM sheets at all."""
    xlsx_path = Path(xlsx_path)
    if filename is None:
        filename = xlsx_path.name

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    bom = LoadedBom(filename=filename)
    available = {ws_name.strip(): ws_name for ws_name in wb.sheetnames}

    for canonical_name in PP_SHEETS_ORDER:
        actual = None
        for canon, real in available.items():
            if canon.casefold() == canonical_name.casefold():
                actual = real
                break
        if actual is None:
            continue
        ws = wb[actual]
        bom.sheets[canonical_name] = _read_sheet(ws, canonical_name)

    if not bom.sheets:
        wb.close()
        raise ValueError(
            f"File '{filename}' has no recognizable BOM sheets. "
            f"Expected at least one of: BOM Header, BOM Item, BOM Subitem. "
            f"Found: {', '.join(wb.sheetnames)}. "
            f"If this is a Routing file, drop it in the Routing slot instead."
        )

    wb.close()
    return bom
