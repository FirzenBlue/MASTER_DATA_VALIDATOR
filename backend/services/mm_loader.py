"""
MM loader — reads each of the 3 detected MM xlsx files into a normalized
internal structure that the merger can consume.

Design principles:
  - Read row 2 (SAP codes) as the authoritative column header. Ignore row 1
    (friendly labels) — those are just for human consumption in Excel.
  - Each data row becomes a dict keyed by SAP code, e.g.:
       {"MATNR": "8903837589095", "MTART": "ZFRT", "MAKTX": "...", ...}
  - Values are normalized:
       * None / empty string → stored as None (consistent "missing" marker)
       * Strings: stripped, non-breaking spaces replaced
       * Numbers: kept as-is (float/int) — conversion to string happens
         at LTMC-generation time where length + format matter
  - No validation happens here. This is pure I/O.

The three files have distinct shapes so each gets its own loader function,
but they return a uniform `LoadedFile` dataclass for the merger.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


@dataclass
class LoadedRow:
    """One data row from an MM input file.

    Attributes:
        excel_row: 1-based Excel row number this row came from. First data
            row is excel_row=3 (rows 1,2 are headers). Used for error messages
            so SMEs can find the row in the original file.
        values: SAP-code → value. Keys are uppercased SAP codes (MATNR, WERKS).
            Values are the normalized cell values.
    """
    excel_row: int
    values: dict[str, Any]

    def get(self, sap_code: str, default: Any = None) -> Any:
        """Fetch a value by SAP code. Case-insensitive."""
        return self.values.get(sap_code.upper(), default)


@dataclass
class LoadedFile:
    """A fully-loaded MM input xlsx.

    Attributes:
        filename: the uploaded filename (for error messages)
        role: "main", "alt_uom", "longtext"
        sap_fields: row-2 SAP codes, in original column order. Preserves
            order so the LTMC generator can iterate columns deterministically.
        header_labels: row-1 friendly labels (e.g. "Material Number",
            "Industry"), one per sap_field. Used by the cleaned-xlsx
            export so the output is readable by humans, not just SAP
            codes. Empty strings for columns where row 1 is blank.
        rows: list of LoadedRow, in file order. An empty list means the file
            had only headers and no data.
        skipped_blank_rows: count of rows we skipped because every cell was
            blank (trailing empty rows at end of file, or mid-file gaps).
    """
    filename: str
    role: str
    sap_fields: list[str] = field(default_factory=list)
    header_labels: list[str] = field(default_factory=list)
    rows: list[LoadedRow] = field(default_factory=list)
    skipped_blank_rows: int = 0


def _clean_cell(value: Any) -> Any:
    """Normalize a raw cell value.

    - None / empty / whitespace-only → None
    - Strings: strip whitespace, replace non-breaking space (\xa0) with " "
    - Numbers: passed through unchanged
    - Dates: passed through unchanged (datetime → ISO format happens later)
    """
    if value is None:
        return None
    if isinstance(value, str):
        s = value.replace("\xa0", " ").strip()
        return s if s else None
    return value


def _read_xlsx_with_2row_header(xlsx_path: str | Path) -> tuple[list[str], list[LoadedRow], int, list[str]]:
    """Core reader: row 1 = friendly labels, row 2 = SAP codes, row 3+ = data.

    Returns (sap_fields_in_order, loaded_rows, skipped_blank_rows, header_labels).

    Performance note: openpyxl's pure-Python XML parser takes ~50 seconds
    on a 12 MB / 24,000-row file. python-calamine (Rust-backed) handles
    the same file in ~3 seconds. We try calamine first; if it isn't
    installed or fails, we fall back to openpyxl so deployments without
    the package still work.
    """
    # Fast path: python-calamine
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(str(xlsx_path))
        sheet = wb.get_sheet_by_index(0)
        data = sheet.to_python()
    except (ImportError, Exception) as cal_err:
        # Slow path: openpyxl. Same logic, just slower.
        if not isinstance(cal_err, ImportError):
            # Calamine installed but errored on this file — log and fall through.
            print(f"[mm_loader] calamine failed ({cal_err!r}), falling back to openpyxl")
        return _read_xlsx_with_openpyxl(xlsx_path)

    sap_fields: list[str] = []
    header_labels: list[str] = []
    rows: list[LoadedRow] = []
    skipped_blank = 0

    if len(data) < 2:
        return [], [], 0, []

    # Row 1: friendly labels for export rendering
    header_labels = [
        str(c).strip() if c is not None and str(c).strip() else ""
        for c in data[0]
    ]
    # Row 2: SAP codes
    sap_fields = [
        str(c).strip().upper() if c is not None and str(c).strip() else ""
        for c in data[1]
    ]

    # Data rows from index 2 onwards (Excel row 3+)
    for excel_idx, row in enumerate(data[2:], start=3):
        # Skip blank rows defensively
        if not any(
            c is not None and (not isinstance(c, str) or c.strip())
            for c in row
        ):
            skipped_blank += 1
            continue

        values: dict[str, Any] = {}
        for col_idx, raw in enumerate(row):
            if col_idx >= len(sap_fields):
                continue
            sap_code = sap_fields[col_idx]
            if not sap_code:
                continue
            cleaned = _clean_cell(raw)
            if cleaned is not None:
                values[sap_code] = cleaned

        rows.append(LoadedRow(excel_row=excel_idx, values=values))

    # Trailing empty cells in calamine output produce empty SAP codes;
    # the data rows above already skip those. For the returned field list,
    # we keep only the non-empty entries (matches openpyxl behaviour).
    sap_fields_clean: list[str] = []
    header_labels_clean: list[str] = []
    for f, lbl in zip(sap_fields, header_labels):
        if f:
            sap_fields_clean.append(f)
            header_labels_clean.append(lbl or f)

    return sap_fields_clean, rows, skipped_blank, header_labels_clean


def _read_xlsx_with_openpyxl(xlsx_path: str | Path) -> tuple[list[str], list[LoadedRow], int, list[str]]:
    """Slow openpyxl fallback — same shape as the calamine reader."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    sap_fields: list[str] = []
    header_labels: list[str] = []
    rows: list[LoadedRow] = []
    skipped_blank = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            header_labels = [
                str(c).strip() if c is not None and str(c).strip() else ""
                for c in row
            ]
            continue
        if i == 2:
            sap_fields = [
                str(c).strip().upper() if c is not None and str(c).strip() else ""
                for c in row
            ]
            continue
        if not any(c is not None and (not isinstance(c, str) or c.strip()) for c in row):
            skipped_blank += 1
            continue

        values: dict[str, Any] = {}
        for col_idx, raw in enumerate(row):
            if col_idx >= len(sap_fields):
                continue
            sap_code = sap_fields[col_idx]
            if not sap_code:
                continue
            cleaned = _clean_cell(raw)
            if cleaned is not None:
                values[sap_code] = cleaned

        rows.append(LoadedRow(excel_row=i, values=values))

    sap_fields_clean: list[str] = []
    header_labels_clean: list[str] = []
    for f, lbl in zip(sap_fields, header_labels):
        if f:
            sap_fields_clean.append(f)
            header_labels_clean.append(lbl or f)

    return sap_fields_clean, rows, skipped_blank, header_labels_clean


def load_main(xlsx_path: str | Path, filename: str) -> LoadedFile:
    """Load the MAIN MM file (~150 cols, 1 row per material)."""
    fields_, rows, skipped, labels = _read_xlsx_with_2row_header(xlsx_path)
    return LoadedFile(
        filename=filename, role="main",
        sap_fields=fields_, header_labels=labels, rows=rows,
        skipped_blank_rows=skipped,
    )


def load_alt_uom(xlsx_path: str | Path, filename: str) -> LoadedFile:
    """Load the Alternate Units file (4 cols, can have multiple rows per material)."""
    fields_, rows, skipped, labels = _read_xlsx_with_2row_header(xlsx_path)
    return LoadedFile(
        filename=filename, role="alt_uom",
        sap_fields=fields_, header_labels=labels, rows=rows,
        skipped_blank_rows=skipped,
    )


def load_longtext(xlsx_path: str | Path, filename: str) -> LoadedFile:
    """Load the Long Text file (~7 cols, multiple rows per material possible)."""
    fields_, rows, skipped, labels = _read_xlsx_with_2row_header(xlsx_path)
    return LoadedFile(
        filename=filename, role="longtext",
        sap_fields=fields_, header_labels=labels, rows=rows,
        skipped_blank_rows=skipped,
    )


# Dispatch table for the merger/upload endpoint — looks up the right
# loader by role without an if/elif chain.
LOADERS_BY_ROLE = {
    "main": load_main,
    "alt_uom": load_alt_uom,
    "longtext": load_longtext,
}
