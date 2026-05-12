"""
KDS Loader — parse a client-provided KDS workbook into catalogs the
validator can use.

This is a stopgap until KDS-in-DB (v32) ships; for now we read the
xlsx at startup and populate an in-memory catalog dict. The file lives
in the repo under `kds/Sales_and_Dist_KDS.xlsx` — replaceable without
a redeploy by admins who SCP a new file to that path.

Challenges handled:
  - Sheets have headers at different row positions (1, 2, 3, 5)
  - Some sheets start with a metadata row ("Sales Organization: IN02")
    that must be skipped before the real header is found
  - Trailing blank rows are common (Sales Groups has ~800 blanks at
    the end) — we stop reading when we see N consecutive blanks
  - Section breaks mid-sheet exist (Sales Groups has blank rows between
    the old-catalog section and the current-catalog section; client SMEs
    confirmed both sections are valid so we keep both)
  - Stray comment columns ("*Input from Business...") are ignored — we
    only read columns 1 (code) and 2 (description)
  - Non-breaking spaces (\xa0) and trailing whitespace on values
  - Excel numeric codes ("1" as int, not str "01") are coerced to str

Public API:
  load_catalogs(xlsx_path) -> dict[rule_id, dict[code, description]]
    Returns the exact shape that CATALOG_BY_RULE used to have, so
    drop-in replacement for the hardcoded catalogs.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any


# Which sheet in the KDS file maps to which rule_id.
# Column positions are 0-indexed; first two columns are always code + description
# in this client's KDS format (other columns are comments/metadata we ignore).
SHEET_RULE_MAP = {
    "Sales Groups":         "sales_group_not_in_kds",
    "Sales Office":         "sales_office_not_in_kds",
    "Customer Groups":      "customer_group_not_in_kds",
    "Distribution Channel": "distribution_channel_not_in_kds",
    "Division":             "division_not_in_kds",
    "Sales Organization":   "sales_org_not_in_kds",
    "Sales Districts":      "sales_district_not_in_kds",
    "Shipping Conditions":  "shipping_condition_not_in_kds",
    "Incoterms":            "inco_location_description",
}

# How many consecutive blank rows before we decide the sheet's data is over.
# 5 is enough to skip a legitimate section break (which is 1-2 blanks) while
# still catching the "rest of sheet is empty" pattern at the file end.
TRAILING_BLANK_THRESHOLD = 5


def _clean(value: Any) -> str:
    """Normalise a cell to a trimmed string. Empty → empty string.

    Handles:
      - None → ""
      - int/float codes: "1" not "1.0"
      - non-breaking space (\xa0) anywhere in the value
      - trailing/leading whitespace
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value)
    s = s.replace("\xa0", " ").strip()
    return s


def _find_header_row(rows: list[tuple]) -> int | None:
    """Locate the row that acts as the code/description header.

    Strategy: walk the first 10 rows looking for a row whose column-A value
    is a known header label ('Sales Group', 'Code', 'Sales District', ...).
    If no match, fall back to the first row whose column A is non-blank and
    column B is non-blank — that's almost always the header of a sheet whose
    first row isn't blank.

    Returns 0-indexed row number, or None if no header detected.
    """
    header_hints = {
        "sales group", "sales groups", "sales office", "customer group",
        "customer groups", "distribution channel", "division", "company code",
        "sales district", "sales districts", "shipping condition",
        "inco terms", "incoterms", "code",
    }
    for idx, row in enumerate(rows[:10]):
        if not row or row[0] is None:
            continue
        a = _clean(row[0]).lower()
        if a in header_hints:
            return idx
    # Fallback: first row where both col A and col B are non-blank.
    for idx, row in enumerate(rows[:10]):
        if not row or len(row) < 2:
            continue
        if _clean(row[0]) and _clean(row[1]):
            return idx
    return None


def _parse_sheet(rows: list[tuple]) -> dict[str, str]:
    """Extract {code: description} pairs from a single sheet's rows.

    Walks from the row after the header onward. Stops after
    TRAILING_BLANK_THRESHOLD consecutive blank rows — this collapses
    the 800+ empty rows at the end of Sales Groups without also
    cutting off legitimate section breaks (1-2 blanks between
    sub-sections of the catalog).
    """
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return {}

    result: dict[str, str] = {}
    consecutive_blanks = 0

    for row in rows[header_idx + 1:]:
        if not row:
            consecutive_blanks += 1
            if consecutive_blanks >= TRAILING_BLANK_THRESHOLD:
                break
            continue

        code = _clean(row[0]) if len(row) > 0 else ""
        desc = _clean(row[1]) if len(row) > 1 else ""

        if not code:
            # Blank code: could be a section break OR a comment row (e.g.
            # a sub-header like "Sales Organization: IN02" elsewhere in
            # the sheet). Either way, skip it — don't count it as the end
            # unless N of them in a row.
            consecutive_blanks += 1
            if consecutive_blanks >= TRAILING_BLANK_THRESHOLD:
                break
            continue

        # Heuristic guard: some sheets have sub-header rows after a blank
        # that look like "Code | Distribution Channel | Definition" — we
        # detect these by checking if the row is a plausible data row.
        # Data rows have: short code in col A, real text in col B.
        if code.lower() in {"code", "sales group", "sales district",
                             "inco terms", "incoterms"}:
            # A sub-header we accidentally walked into. Skip & reset blanks.
            consecutive_blanks = 0
            continue

        consecutive_blanks = 0

        # Last-write-wins if a code appears twice. Real KDS shouldn't have
        # duplicates, but if it does the later definition is probably the
        # more recent authoritative one.
        result[code] = desc

    return result


def load_catalogs(xlsx_path: str | Path) -> dict[str, dict[str, str]]:
    """Parse every sheet we care about into {rule_id: {code: desc}}.

    Returns a dict shaped exactly like the old hardcoded CATALOG_BY_RULE,
    so callers don't need to change.

    Failures are caught per-sheet: if one sheet is malformed, the others
    still load. A warning is printed so ops can see what went wrong.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        print(f"[kds_loader] could not open {xlsx_path}: {e}", flush=True)
        return {}

    catalogs: dict[str, dict[str, str]] = {}
    for sheet_name, rule_id in SHEET_RULE_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"[kds_loader] sheet '{sheet_name}' not found — skipping {rule_id}",
                  flush=True)
            continue
        try:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            catalog = _parse_sheet(rows)
            catalogs[rule_id] = catalog
            print(f"[kds_loader] {sheet_name}: {len(catalog):>5} entries → {rule_id}",
                  flush=True)
        except Exception as e:
            print(f"[kds_loader] error parsing '{sheet_name}': {e}", flush=True)

    return catalogs
