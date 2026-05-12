"""
MM KDS loader — reads the client's MM KDS workbook into catalog dicts
that mm_validator can use.

Conceptually similar to services/kds_loader.py (for SD/Customer) but the
MM KDS has different per-sheet shapes:
  - Some sheets use row 1 as the header, data from row 2 (Material Group,
    Material Type, External Material Group, Valuation Class, Storage Loc)
  - Some sheets have NO header row — row 1 is already data (Purchasing Group,
    HML Feedback on Plants)
  - Material Type has a spurious " Materials" sub-row between header and
    first real entry — we skip blanks + sub-headers
  - Some catalogs are SCOPED — a value is only valid within a context:
      Storage Location: (WERKS, LGORT) — same LGORT can exist across plants
      Valuation Class: (MTART, BKLAS) — class determines G/L account by type
    For these we return a NESTED dict: outer key is the scope, inner is code→desc.

Output contract (keys used by mm_validator.py):

  CATALOGS = {
    "plant":               {"PE01": "HML Ltd unit 1", ...},          # flat
    "purchasing_org":      {"IN02": ""},                              # flat
    "material_type":       {"ZFRT": "Finished Materials", ...},       # flat
    "material_group":      {"10001": "CARTON", ...},                  # flat
    "ext_material_group":  {"CN-AM": "AIR MASK", ...},                # flat
    "purchasing_group":    {"P01": "Domestic Purchase", ...},         # flat
    "storage_loc_by_plant":{"PE01": {"FEU1": "FG EXPORT", ...}, ...}, # nested
    "valclass_by_mtart":   {"ZFRT": "7920", ...},                     # flat (1-to-1)
  }

  Returned by load_mm_catalogs() as a plain dict. Not-yet-available
  catalogs (Division, Profit Center, DISPO, STRGR, FEVOR, SFCPF, MVGR1-5)
  are returned as empty dicts — validators check for emptiness before
  running the corresponding rule so they silently skip instead of raising.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def _clean(value: Any) -> str:
    """Normalise a cell: strip, NBSP-fix, float→int for whole numbers."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).replace("\xa0", " ").strip()
    return s


def _load_bwkey_by_werks(ws) -> dict[str, str]:
    """Plants sheet → {proposed_plant_code: valuation_area_code}.

    LTMC Valuation Data requires BWKEY per (material × plant). Typically
    BWKEY == the plant's valuation area which, for Healthium, is the
    numeric plant code (PE01→1010, PE02→1011, SC01→1020, etc.). The
    HML Plants sheet records this mapping in column F (Valuation area).

    Returns a dict keyed by the Proposed Plant Code (WERKS) with the
    Valuation Area as the value. Missing WERKS → no mapping, validator
    will flag the material with "BWKEY cannot be derived".
    """
    mapping: dict[str, str] = {}
    rows = list(ws.iter_rows(values_only=True))
    skip_first = (
        rows and rows[0] and _clean(rows[0][0]).lower() == "company code"
    )
    for i, row in enumerate(rows):
        if skip_first and i == 0:
            continue
        if not row or len(row) < 6:
            continue
        werks = _clean(row[1])       # col B: Proposed Plant Code
        bwkey = _clean(row[5])       # col F: Valuation area
        if werks and bwkey and werks not in mapping:
            mapping[werks] = bwkey
    return mapping


def _load_plants(ws) -> dict[str, str]:
    """HML FEEDBACK ON PLANTS → {proposed_plant_code: plant_name}.

    Row 1 is data (no header). Column B (index 1) = Proposed Plant Code,
    column E (index 4) = Plant name.
    """
    catalog: dict[str, str] = {}
    # Row 1 in this sheet is 'Company Code / Proposed Plant Code / Remarks / Plant / Name 1 ...'
    # Actually looking at it again - it IS a header because col A says 'Company Code'.
    # Tell apart by: if row 1 col A equals 'Company Code', skip it.
    skip_first = False
    rows = list(ws.iter_rows(values_only=True))
    if rows and rows[0] and _clean(rows[0][0]).lower() == "company code":
        skip_first = True
    for i, row in enumerate(rows):
        if skip_first and i == 0:
            continue
        if not row or len(row) < 5:
            continue
        code = _clean(row[1])      # col B: Proposed Plant Code (PE00, PE01 ...)
        name = _clean(row[4])      # col E: Plant name
        if code and code not in catalog:
            # Keep first occurrence — file has duplicates (same PE-code used
            # by multiple old plant numbers).
            catalog[code] = name
    return catalog


def _load_flat_catalog(ws, code_col: int, desc_col: int,
                       skip_header_rows: int = 1,
                       skip_blank_rows: bool = True) -> dict[str, str]:
    """Generic loader: {col[code_col]: col[desc_col]} across all data rows.

    Args:
        code_col: 0-indexed column with the code
        desc_col: 0-indexed column with the description
        skip_header_rows: how many leading rows to skip
        skip_blank_rows: skip rows where the code cell is empty
    """
    catalog: dict[str, str] = {}
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        if i < skip_header_rows:
            continue
        if not row or len(row) <= code_col:
            continue
        code = _clean(row[code_col])
        if skip_blank_rows and not code:
            continue
        # Guard against sub-header rows: if the cell looks like a header
        # word (e.g. " Materials" in Material Type sheet), skip it.
        if code.lower() in {"materials", "material type", "material group",
                             "code", "plant", "storage location"}:
            continue
        desc = _clean(row[desc_col]) if len(row) > desc_col else ""
        if code not in catalog:
            catalog[code] = desc
    return catalog


def _load_storage_loc(ws) -> dict[str, dict[str, str]]:
    """Storage Location → {plant: {lgort: description}}.

    Row 1 is header. Col A = Plant, col B = Storage Location, col C = desc.
    Same LGORT code can appear under different plants, each with its own
    description — hence the nested dict.
    """
    nested: dict[str, dict[str, str]] = {}
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        if i == 0:   # skip header
            continue
        if not row or len(row) < 3:
            continue
        plant = _clean(row[0])
        lgort = _clean(row[1])
        desc = _clean(row[2])
        if not plant or not lgort:
            continue
        nested.setdefault(plant, {})[lgort] = desc
    return nested


def _load_valclass_mapping(ws) -> dict[str, str]:
    """Valuation Class Mapping → {mtart: valuation_class}.

    Row 1 is header. Col A = Material Type, col C = Valuation Class.
    This is a 1-to-1 mapping per material type (each MTART has ONE
    expected valuation class). Used by mm_validator Checklist rule 27
    to verify BKLAS matches what's expected for the MTART.
    """
    mapping: dict[str, str] = {}
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        if i == 0:
            continue
        if not row or len(row) < 3:
            continue
        mtart = _clean(row[0])
        vclass = _clean(row[2])
        if mtart and vclass:
            mapping[mtart] = vclass
    return mapping


def _load_mtart_ranges(ws) -> dict[str, dict]:
    """Read the Material Type sheet's range config into a per-MTART dict.

    For each MTART row, captures:
      - range_type: "Internal" or "External" (case preserved as-is)
      - range_from: raw string of the From No.
      - range_to:   raw string of the End No.
      - range_is_numeric: True if both bounds parse as integers (meaning
        the MATNR must be numeric and within the integer range). False
        for alphanumeric ranges like ZHLB's "A" to "ZZZZZZZZ".
      - range_min_len / range_max_len: length bounds derived from the
        from/to strings. Used to validate MATNRs that use the alphanumeric
        "A to ZZZZZZZZ" shape.

    External ranges still have bounds — SAP validates them on save. This
    is what catches 'TBC-2' on a ZFRT material: ZFRT range is numeric
    13-digit, 'TBC-2' is 5 chars + alphabetic, so it fails.

    Returns a dict keyed by MTART. MTARTs missing from the KDS are
    absent — the validator treats that as "unknown, skip" rather than
    "invalid". Protects against KDS updates that add new MTARTs mid-use.
    """
    ranges: dict[str, dict] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # header
        if not row or row[0] is None:
            continue
        mtart = _clean(row[0])
        if not mtart or mtart.isspace():
            continue
        # Skip the spurious " Materials" sub-header row that has no data
        if len(row) < 6 or all(row[k] is None for k in (3, 4, 5)):
            continue
        range_type = _clean(row[3]) if len(row) > 3 else ""
        range_from = _clean(row[4]) if len(row) > 4 else ""
        range_to = _clean(row[5]) if len(row) > 5 else ""
        if not (range_type and range_from and range_to):
            continue
        # Is the range numeric? Both bounds must parse as integers.
        try:
            from_int = int(range_from)
            to_int = int(range_to)
            range_is_numeric = True
        except (ValueError, TypeError):
            from_int = None
            to_int = None
            range_is_numeric = False
        # Length bounds (used for alphanumeric external ranges)
        range_min_len = min(len(range_from), len(range_to)) if range_from and range_to else 0
        range_max_len = max(len(range_from), len(range_to)) if range_from and range_to else 0
        ranges[mtart] = {
            "range_type": range_type,       # "Internal" | "External"
            "range_from": range_from,
            "range_to": range_to,
            "range_is_numeric": range_is_numeric,
            "range_from_int": from_int,
            "range_to_int": to_int,
            "range_min_len": range_min_len,
            "range_max_len": range_max_len,
        }
    return ranges


def load_mm_catalogs(xlsx_path: str | Path) -> dict[str, Any]:
    """Main entry point. Returns a dict of all MM catalogs.

    See module docstring for the key shape. Missing/unreadable sheets
    result in empty catalogs — validators handle that gracefully by
    skipping rules that depend on them.
    """
    xlsx_path = Path(xlsx_path)  # normalize — call sites pass Path in prod but tests pass str
    catalogs: dict[str, Any] = {
        "plant": {},
        "bwkey_by_werks": {},
        "purchasing_org": {},
        "material_type": {},
        "mtart_ranges": {},
        "material_group": {},
        "ext_material_group": {},
        "purchasing_group": {},
        "storage_loc_by_plant": {},
        "valclass_by_mtart": {},
        # Pending catalogs — populated when the SAP team provides the data.
        # Keeping them here (as empty dicts) so validator rule wiring is
        # complete; rules silently skip when the catalog is empty.
        "division": {},
        "profit_center": {},
        "mrp_controller": {},
        "strategy_group": {},
        "production_supervisor": {},
        "scheduling_profile": {},
        # v66: ISO Unit-of-Measure catalog — loaded from the separate file
        # `ISO_Unit_Of_Measure_tentative_file.xlsx` (NOT from MM_KDS).
        # Populated below after MM_KDS sheets are processed. Used to
        # validate MEINS in the main file and MEINH in the alt-uom file.
        # Keyed by the "Internal UoM" code (col A in the ISO file): %, CMS,
        # kB, A, MMI, etc. Value is the descriptor (matches the 3-char or
        # 6-char external code when available).
        "iso_uom": {},
    }

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        print(f"[mm_kds] could not open {xlsx_path}: {e}", flush=True)
        return catalogs

    # Each sheet → handler. Missing sheets are logged but don't crash.
    handlers: list[tuple[str, str, Any]] = [
        ("HML FEEDBACK ON PLANTS", "plant", _load_plants),
        # Same Plants sheet, second catalog: plant → valuation area mapping
        # used by the BWKEY-derivation rule in the validator.
        ("HML FEEDBACK ON PLANTS", "bwkey_by_werks", _load_bwkey_by_werks),
        ("Storage Location",        "storage_loc_by_plant", _load_storage_loc),
        ("Material Type",           "material_type",
            lambda ws: _load_flat_catalog(ws, code_col=0, desc_col=1, skip_header_rows=1)),
        # Same sheet, second catalog: per-MTART range bounds for MATNR validation
        ("Material Type",           "mtart_ranges", _load_mtart_ranges),
        ("Material Group",          "material_group",
            lambda ws: _load_flat_catalog(ws, code_col=0, desc_col=1, skip_header_rows=1)),
        ("External Material Group", "ext_material_group",
            lambda ws: _load_flat_catalog(ws, code_col=0, desc_col=1, skip_header_rows=1)),
        ("Purchasing Group",        "purchasing_group",
            # Sheet has NO real header row — row 1 column A says
            # "Proposed Purchasing groups" (wraps header) so skip 1.
            lambda ws: _load_flat_catalog(ws, code_col=0, desc_col=1, skip_header_rows=1)),
        ("Valuation Class Mapping", "valclass_by_mtart", _load_valclass_mapping),
        # Purchasing Org is a tiny 3-col sheet where col A = Purch Org.
        # We only need the distinct set of Purch Org codes.
        ("Purchasing Org",          "purchasing_org",
            lambda ws: _load_flat_catalog(ws, code_col=0, desc_col=1, skip_header_rows=1)),
    ]

    for sheet_name, catalog_key, handler in handlers:
        if sheet_name not in wb.sheetnames:
            print(f"[mm_kds] sheet '{sheet_name}' not found — skipping {catalog_key}",
                  flush=True)
            continue
        try:
            ws = wb[sheet_name]
            cat = handler(ws)
            catalogs[catalog_key] = cat
            if isinstance(cat, dict):
                count = sum(len(v) for v in cat.values()) if \
                    (cat and isinstance(next(iter(cat.values()), None), dict)) else len(cat)
                print(f"[mm_kds] {sheet_name}: {count:>4} entries → {catalog_key}",
                      flush=True)
        except Exception as e:
            print(f"[mm_kds] error parsing '{sheet_name}': {e}", flush=True)

    # v66: load ISO Unit-of-Measure catalog from the dedicated file
    # bundled alongside MM_KDS.xlsx in backend/kds/. The file is a single
    # "Data" sheet with row 1 = headers. Col A holds the canonical
    # "Internal UoM" code; col B holds the 3-char external unit; col C
    # the 6-char external unit. We index by Internal UoM (canonical) and
    # also accept the 3-char/6-char codes (since main MEINS may use any
    # of them depending on SAP customization). Three lookup sets keep the
    # ISO check forgiving but precise.
    iso_path = xlsx_path.parent / "ISO_Unit_Of_Measure_tentative_file.xlsx"
    if iso_path.exists():
        try:
            iso_wb = openpyxl.load_workbook(str(iso_path), read_only=True, data_only=True)
            if "Data" in iso_wb.sheetnames:
                iso_ws = iso_wb["Data"]
                iso_codes: dict[str, str] = {}
                for row in iso_ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    # col 0 = Internal UoM, col 1 = 3-char external, col 2 = 6-char
                    internal = str(row[0]).strip() if row[0] else ""
                    three = str(row[1]).strip() if (len(row) > 1 and row[1] and str(row[1]).strip() != "X") else ""
                    six = str(row[2]).strip() if (len(row) > 2 and row[2] and str(row[2]).strip() != "X") else ""
                    if internal:
                        iso_codes[internal] = internal
                    if three:
                        iso_codes[three] = internal
                    if six:
                        iso_codes[six] = internal
                catalogs["iso_uom"] = iso_codes
                print(f"[mm_kds] ISO UoM: {len(iso_codes):>4} entries → iso_uom",
                      flush=True)
        except Exception as e:
            print(f"[mm_kds] could not load ISO UoM file at {iso_path}: {e}", flush=True)
    else:
        print(f"[mm_kds] ISO UoM file not found at {iso_path} — iso_uom catalog empty",
              flush=True)

    # v70: Load Division catalog from the dedicated Divison_KDS.xlsx file
    # (preserve SME's filename spelling). This file has one sheet per
    # sales organisation — INO2 (Sutures), IN03 (QNPM), IN04 (CareNow
    # Medical), IN05 (CareNow Lifesciences), UK02 (Healthium Medtech) —
    # with division codes that vary by sales org. We flatten ALL division
    # values across all sheets into one suggestion catalog so the SME
    # sees every valid code in the Group & Replace dropdown regardless
    # of which sales org the material belongs to. SME quote (2026-05-11):
    # "for division column you can use this file 'division KDS' for user
    # u can suggest every value present in those sheets in this file
    # 'Division KDS' so user can input anything he want."
    #
    # Each sheet has:
    #   Row 1: ['Sales Organization', sales_org_code, sales_org_name]
    #   Row 2: blank
    #   Row 3: ['Code', 'Division', 'Definition'] (header)
    #   Row 4+: data rows
    # IN04 has no data (only metadata) — handled gracefully.
    # Code 04 appears in both INO2 ('Arthroscopy') and UK02 ('Arthoscopy')
    # — last-write-wins for the description; both interpretations remain
    # valid since the validator only checks code-membership, not desc.
    #
    # v70 supersedes the v69 SD-KDS Division loader (which read from the
    # SD KDS Division sheet — a single flat list). The v69 SD-KDS load
    # block is removed below.
    div_path = xlsx_path.parent / "Divison_KDS.xlsx"
    if div_path.exists():
        try:
            div_wb = openpyxl.load_workbook(str(div_path), read_only=True, data_only=True)
            division: dict[str, str] = {}
            # v70.1: ALSO build a full options list (one entry per sheet per
            # unique code+desc pair) so the dropdown shows every entry from
            # every sheet, not just the dedup'd 17. SME quote (2026-05-11):
            # "for division show every values those are present in the excel
            # sheet 'INO2','IN03','IN04','IN05','UK02' so user can use any
            # value". Same code can legitimately mean different things per
            # sales org — e.g. '04' is 'Arthroscopy' in INO2 (Sutures) but
            # 'Arthoscopy' [sic] in UK02; '08 Needles' appears in both INO2
            # and IN03. We surface both so the SME picks the right one.
            division_options: list[dict[str, str]] = []
            seen_per_sheet: set[tuple[str, str, str]] = set()
            per_sheet_counts: dict[str, int] = {}
            for sheet_name in div_wb.sheetnames:
                ws = div_wb[sheet_name]
                # Read row 1 to get the sales-org name ("Sutures Sales Org" etc)
                sales_org_name = ""
                try:
                    r1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if r1 and len(r1) >= 3 and r1[2]:
                        sales_org_name = str(r1[2]).strip()
                except Exception:
                    pass
                count_this_sheet = 0
                # Start at row 4 — rows 1-3 are metadata + header.
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if not row or row[0] is None:
                        continue
                    code = str(row[0]).strip()
                    if not code or code.lower() in ("code", "division"):
                        continue
                    # Pad numeric codes to 2 digits (Excel stores '01' as int 1)
                    if code.isdigit() and len(code) == 1:
                        code = code.zfill(2)
                    desc = (str(row[1]).strip() if len(row) > 1 and row[1] else "")

                    # Dict catalog (validator existence check) — last-write-wins
                    division[code] = desc

                    # Options list (dropdown) — dedup by (sheet, code, desc)
                    # so within-sheet duplicate rows don't show twice, but
                    # cross-sheet variants stay visible.
                    key = (sheet_name, code, desc)
                    if key in seen_per_sheet:
                        continue
                    seen_per_sheet.add(key)
                    label_suffix = f" ({sheet_name}"
                    if sales_org_name:
                        label_suffix += f" — {sales_org_name}"
                    label_suffix += ")"
                    label = f"{code} — {desc}{label_suffix}" if desc else f"{code}{label_suffix}"
                    division_options.append({"value": code, "label": label})
                    count_this_sheet += 1
                if count_this_sheet:
                    per_sheet_counts[sheet_name] = count_this_sheet
            catalogs["division"] = division
            catalogs["division_options"] = division_options
            sheets_summary = ", ".join(f"{s}={n}" for s, n in per_sheet_counts.items())
            print(f"[mm_kds] Division (from Divison_KDS): {len(division)} unique codes / "
                  f"{len(division_options)} dropdown options [{sheets_summary}]", flush=True)
        except Exception as e:
            print(f"[mm_kds] could not load Divison_KDS at {div_path}: {e}", flush=True)
    else:
        # Fallback to SD KDS if Divison_KDS.xlsx not bundled (v69 behaviour).
        # Kept for graceful degradation during deploy transitions.
        sd_kds_path = xlsx_path.parent / "Sales_and_Dist_KDS.xlsx"
        if sd_kds_path.exists():
            try:
                sd_wb = openpyxl.load_workbook(str(sd_kds_path), read_only=True, data_only=True)
                if "Division" in sd_wb.sheetnames:
                    div_ws = sd_wb["Division"]
                    division = {}
                    blank_streak = 0
                    for row in div_ws.iter_rows(min_row=6, values_only=True):
                        if not row or row[0] is None or str(row[0]).strip() == "":
                            blank_streak += 1
                            if blank_streak >= 2:
                                break
                            continue
                        blank_streak = 0
                        code = str(row[0]).strip()
                        if code.lower() in ("code", "division"):
                            continue
                        desc = (str(row[1]).strip() if len(row) > 1 and row[1] else "")
                        division[code] = desc
                    catalogs["division"] = division
                    print(f"[mm_kds] Division (fallback SD KDS): {len(division):>4} entries → division",
                          flush=True)
            except Exception as e:
                print(f"[mm_kds] fallback Division load failed: {e}", flush=True)

    return catalogs
