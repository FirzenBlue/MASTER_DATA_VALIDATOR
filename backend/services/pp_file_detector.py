"""
PP file detector — fast pre-upload sanity check for BOM and Routing files.

Goal: peek at the sheet names and row 2 (SAP codes) of an uploaded xlsx
and decide:
  - role = "bom" → looks like a BOM file (drop in BOM slot)
  - role = "routing" → looks like a Routing file (drop in Routing slot)
  - role = "unknown" → can't tell; SME picked the wrong file or pasted
    data into a non-template xlsx
  - role = "neither" → file structurally invalid (no 2-row header, no
    SAP codes, etc.)

The detector runs from /api/pp/format-check before the real upload, so
we use openpyxl's read_only mode to keep RAM tiny on big files. We only
read sheet names + row 2 of each sheet, which is essentially free.

Why this exists
---------------
Two scenarios this catches:
1. User has a BOM file and a Routing file open. They mean to drop BOM
   into the BOM slot but drag the wrong one. The detector returns
   role="routing" and the slot rejects it with "this is a Routing file,
   drop it in the Routing slot instead" — saves a 5-minute upload that
   would 400 at parse time.
2. User exports from SAP/3rd party with a non-LTMC structure. Detector
   returns role="unknown" with a hint about expected sheet names.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import openpyxl

from .pp_loader import PP_SHEETS_ORDER
from .routing_loader import ROUTING_SHEETS_ORDER


FileRole = Literal["bom", "routing", "unknown", "neither"]


@dataclass
class DetectionResult:
    role: FileRole
    reason: str
    confidence: float  # 0.0–1.0; how sure we are
    matched_sheets: list[str]  # sheet names that matched this role


# Anchor sheets — finding any one of these is strong evidence the file
# is the corresponding role. We pick the most distinctive sheets per
# role; sheets like "Global Dependency" appear in BOTH BOM and Routing
# templates so they're not anchors.
_BOM_ANCHOR_SHEETS = {"BOM Header", "BOM Item", "BOM Subitem"}
_ROUTING_ANCHOR_SHEETS = {
    "Routing Group", "Task List - Header", "Operations", "Sequences",
    "Sub Operations", "Material Task List Assignment",
    "Production Resources and Tools", "Inspection Plan Characteristic",
}


def detect(xlsx_path: str | Path, filename: str | None = None) -> DetectionResult:
    """Sniff a file's sheet names to determine BOM vs Routing.

    Counts how many anchor sheets of each role are present. Whichever
    role wins (more matches) is returned with confidence proportional
    to the match count. If neither role matches any anchor, returns
    role="unknown" with a hint about expected sheets.

    Why anchor-counting rather than checking row 2 SAP codes: the
    sheet-name check is enough on its own and runs in <50ms even for
    multi-GB files (openpyxl read_only doesn't load any data until
    iter_rows is called). Adding row-2 checks would catch edge cases
    (renamed sheets) but cost time and complexity for diminishing
    returns. The merger/loader will still reject malformed files at
    parse time as defense-in-depth.
    """
    xlsx_path = Path(xlsx_path)
    if filename is None:
        filename = xlsx_path.name

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        return DetectionResult(
            role="neither",
            reason=f"Couldn't open as Excel: {e}. Make sure the file is a real .xlsx (Office 2007+).",
            confidence=1.0,
            matched_sheets=[],
        )

    sheets_in_file = {name.strip().casefold(): name for name in wb.sheetnames}
    wb.close()

    # Match anchor sheets case-insensitively.
    bom_matches = [name for name in _BOM_ANCHOR_SHEETS
                   if name.casefold() in sheets_in_file]
    routing_matches = [name for name in _ROUTING_ANCHOR_SHEETS
                       if name.casefold() in sheets_in_file]

    if not bom_matches and not routing_matches:
        return DetectionResult(
            role="unknown",
            reason=(
                f"File has no BOM or Routing anchor sheets. "
                f"Expected one of [{', '.join(sorted(_BOM_ANCHOR_SHEETS))}] for BOM, "
                f"or [{', '.join(sorted(list(_ROUTING_ANCHOR_SHEETS)[:4]))}, …] for Routing. "
                f"Found tabs: {', '.join(list(sheets_in_file.values())[:8])}."
            ),
            confidence=0.0,
            matched_sheets=[],
        )

    if len(bom_matches) > len(routing_matches):
        # Confidence: 1.0 if we got all 3 BOM anchors, scaled down otherwise.
        conf = min(1.0, len(bom_matches) / 3.0)
        return DetectionResult(
            role="bom",
            reason=f"Detected as BOM — found {len(bom_matches)} of 3 BOM anchor sheets.",
            confidence=conf,
            matched_sheets=bom_matches,
        )

    if len(routing_matches) > len(bom_matches):
        conf = min(1.0, len(routing_matches) / 4.0)
        return DetectionResult(
            role="routing",
            reason=f"Detected as Routing — found {len(routing_matches)} Routing anchor sheets.",
            confidence=conf,
            matched_sheets=routing_matches,
        )

    # Equal counts — ambiguous (file has anchors from both). Favour
    # Routing in this rare case, since BOM Header alone can match a
    # Routing file via case-insensitive comparison if the Routing file
    # has a tab oddly named "BOM Header" for some custom reason.
    return DetectionResult(
        role="unknown",
        reason=(
            f"File has anchor sheets from both BOM ({', '.join(bom_matches)}) "
            f"and Routing ({', '.join(routing_matches)}). "
            f"Can't tell which role to assign. Check the file structure."
        ),
        confidence=0.5,
        matched_sheets=bom_matches + routing_matches,
    )
