"""
Routing loader — reads a Routing xlsx (the "Source data for Routing"
LTMC template populated with user data) into a normalized internal
structure.

Same row layout as BOM (verified against the SAP S/4HANA 2025 Routing
template AND the customer's New_Routing_Sheet_peenya.xlsx):
  Row 4 = SAP table name (S_GROUP, S_PLKO, S_OPERATION, …)
  Row 5 = SAP field codes (PLNNR, PLNAL, ...) — AUTHORITATIVE
  Row 6 = ETE format spec
  Row 7 = field-group label
  Row 8 = description (trailing "*" = mandatory)
  Row 9+ = data rows

Routing has 14 data sheets (BOM has 10). Operations is the largest by
far — the customer's file has 72701 Operations rows.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


ROUTING_SHEETS_ORDER: tuple[str, ...] = (
    "Routing Group",
    "Task List - Header",
    "Material Task List Assignment",
    "Sequences",
    "Operations",
    "Component Assignment",
    "Production Resources and Tools",
    "Sub Operations",
    "Inspection Plan Characteristic",
    "Global Dependency",
    "Local Dependency",
    "Local Dependency Description",
    "Documentation of Dependency",
    "Sources of Local Dependency",
)

_CODE_ROW = 5
_DATA_START_ROW = 9


@dataclass
class LoadedRow:
    excel_row: int
    values: dict[str, Any]

    def get(self, sap_code: str, default: Any = None) -> Any:
        return self.values.get(sap_code.upper(), default)


@dataclass
class LoadedSheet:
    name: str
    sap_fields: list[str] = field(default_factory=list)
    rows: list[LoadedRow] = field(default_factory=list)
    skipped_blank_rows: int = 0


@dataclass
class LoadedRouting:
    filename: str
    sheets: dict[str, LoadedSheet] = field(default_factory=dict)


def _norm_cell(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.replace("\u00a0", " ").strip()
        return s if s else None
    return v


def _row_is_blank(values: dict[str, Any]) -> bool:
    return all(v is None for v in values.values())


def _read_sheet(ws, sheet_name: str) -> LoadedSheet:
    sheet = LoadedSheet(name=sheet_name)
    if ws.max_row < _CODE_ROW:
        return sheet

    header_cells = [ws.cell(row=_CODE_ROW, column=c).value
                    for c in range(1, ws.max_column + 1)]
    last_col = 0
    for i, code in enumerate(header_cells):
        if code is not None and str(code).strip():
            last_col = i + 1
    if last_col == 0:
        return sheet

    col_to_sap: list[tuple[int, str]] = []
    sap_fields: list[str] = []
    for col_idx in range(last_col):
        code = header_cells[col_idx]
        if code is None or not str(code).strip():
            continue
        sap = str(code).strip().upper()
        col_to_sap.append((col_idx, sap))
        sap_fields.append(sap)
    sheet.sap_fields = sap_fields

    data_rows: list[LoadedRow] = []
    for excel_row, row_values in enumerate(
        ws.iter_rows(min_row=_DATA_START_ROW, values_only=True), start=_DATA_START_ROW
    ):
        values: dict[str, Any] = {}
        for col_idx, sap in col_to_sap:
            cell = row_values[col_idx] if col_idx < len(row_values) else None
            values[sap] = _norm_cell(cell)
        data_rows.append(LoadedRow(excel_row=excel_row, values=values))

    skipped = 0
    while data_rows and _row_is_blank(data_rows[-1].values):
        data_rows.pop()
        skipped += 1
    sheet.rows = data_rows
    sheet.skipped_blank_rows = skipped
    return sheet


def load_routing(xlsx_path: str | Path, filename: str | None = None) -> LoadedRouting:
    """Load a Routing xlsx file from disk. Raises ValueError if no
    recognizable Routing sheets are found."""
    xlsx_path = Path(xlsx_path)
    if filename is None:
        filename = xlsx_path.name

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    routing = LoadedRouting(filename=filename)
    available = {ws_name.strip(): ws_name for ws_name in wb.sheetnames}

    for canonical_name in ROUTING_SHEETS_ORDER:
        actual = None
        for canon, real in available.items():
            if canon.casefold() == canonical_name.casefold():
                actual = real
                break
        if actual is None:
            continue
        ws = wb[actual]
        routing.sheets[canonical_name] = _read_sheet(ws, canonical_name)

    if not routing.sheets:
        wb.close()
        raise ValueError(
            f"File '{filename}' has no recognizable Routing sheets. "
            f"Expected at least one of: Routing Group, Task List - Header, Operations. "
            f"Found: {', '.join(wb.sheetnames)}. "
            f"If this is a BOM file, drop it in the BOM slot instead."
        )

    wb.close()
    return routing
