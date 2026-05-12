"""
MM file detector — figures out which of the three expected MM input
files a given xlsx is, by sniffing row 2 (the SAP field code header).

The user uploads N xlsx files at once. We don't force them to name
files a specific way or upload in a specific order — we look at each
file's column headers and decide its role.

Expected roles (as of 23-Apr-2026):
  - MAIN    : wide file, 1 row per material. Has MATNR + MBRSH + ~150 cols.
  - ALT_UOM : 1 row per (material × alt unit). Has MATNR + MEINH + UMREZ + UMREN.
  - LONGTEXT: 1 row per (material × text type). Has MATNR + BASE_TEXT.

If a file doesn't match any signature, we return UNKNOWN. The UI then
prompts the SME to identify it manually (or ignore / remove).

We deliberately DO NOT match on filenames — users copy/rename/duplicate
files all the time and we need this to be robust to that.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import openpyxl

FileRole = Literal["main", "alt_uom", "longtext", "unknown"]


@dataclass
class DetectionResult:
    """What we detected about one uploaded file.

    Attributes:
        filename: the filename as uploaded (for display + debug)
        role: detected role, or "unknown" if no signature matched
        sap_fields: the set of row-2 SAP field codes we found (for debug
            display when detection is ambiguous or wrong — the SME can
            see *why* we think it's that role).
        column_count: total columns in row 2 (some diagnostics use this)
        data_rows: number of rows with at least one non-empty cell after
            the 2-row header (useful preview info for the UI)
        reason: short human-readable explanation of why we picked this role
        hint: actionable next-step text shown to the user when role is
            "unknown" or when the file is in the wrong slot. Empty string
            for clean matches — the UI only shows this when there's a
            problem. Examples:
              "This looks like the Long Text file. Try the Long Text slot."
              "Row 2 should contain SAP codes. Did row 1 and row 2 get
               swapped? Open the file and check."
              "Download the Main Template (button at top) to see the
               expected layout."
    """
    filename: str
    role: FileRole
    sap_fields: set[str]
    column_count: int
    data_rows: int
    reason: str
    hint: str = ""


# Signatures in priority order. First match wins. Order matters: MAIN is
# checked before ALT_UOM and LONGTEXT because MAIN contains MATNR + MEINS
# (the base UoM) and we don't want that to accidentally trigger the ALT_UOM
# rule when a file is actually the MAIN file.
#
# Each signature is: (role, required_fields, reason_template, min_columns)
#   - required_fields: set of SAP codes that MUST all appear in row 2
#   - min_columns: minimum column count (rules out tiny "sheet with just
#     MATNR" files that would otherwise match every signature)
#
# We DELIBERATELY use row 2 (SAP codes) not row 1 (friendly labels). Row 1
# labels vary by how the SME renamed them; row 2 is the SAP-standard code
# and is stable across clients.
_SIGNATURES: list[tuple[FileRole, set[str], str, int]] = [
    (
        "main",
        # Signature proof: MAIN has both MBRSH (industry, only in main file)
        # and MATNR. Alt-UoM doesn't have MBRSH, Long-Text doesn't have MBRSH.
        # MBRSH alone is distinctive enough — no other MM file has it — so
        # we don't need a column-count floor as an anti-false-positive guard.
        # This matters because test files / partial exports may have fewer
        # columns but are legitimately the "main" role.
        {"MATNR", "MBRSH"},
        "has MATNR and MBRSH (industry) — wide main-material file",
        3,
    ),
    (
        "alt_uom",
        # Alt-UoM is narrow (4 cols), has UMREZ and UMREN which don't appear
        # anywhere else.
        {"MATNR", "MEINH", "UMREZ", "UMREN"},
        "has MATNR, MEINH, UMREZ, UMREN — alternate units of measure file",
        3,
    ),
    (
        "longtext",
        # Long-Text has BASE_TEXT or PO_TEXT or SALES_TEXT — none appear elsewhere.
        # Any one of these distinctive fields is enough along with MATNR.
        {"MATNR", "BASE_TEXT"},
        "has MATNR and BASE_TEXT — long-text file",
        3,
    ),
]


def _read_row2_codes(xlsx_path: str | Path) -> tuple[set[str], int, int, set[str]]:
    """Read just enough of the xlsx to get the row-2 SAP codes.

    Returns (set_of_sap_codes, total_column_count, data_row_count, row1_codes).

    row1_codes is the same set-of-uppercase-codes shape as the return
    value but from ROW 1 instead of row 2 — so we can detect the common
    "row 1 and row 2 got swapped" mistake. When row 2 looks wrong but
    row 1 has the SAP codes, we tell the user exactly that.

    Uses read_only=True so we don't pay for a full load of a big file just
    to see the headers.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    # Convention: MM files have data on the first sheet. We ignore subsequent
    # sheets — if the user split MM data across sheets of a workbook they're
    # using the tool wrong and we'll error loudly later.
    ws = wb[wb.sheetnames[0]]

    codes: set[str] = set()
    row1_codes: set[str] = set()
    n_cols = 0
    data_row_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            # Row 1 is friendly labels in a correct file, but we capture
            # it into a set so we can detect swap-with-row-2 errors.
            for c in row:
                if c is None:
                    continue
                s = str(c).strip().upper()
                if s:
                    row1_codes.add(s)
            continue
        if i == 2:
            # Row 2 = SAP codes. Normalize: strip whitespace, uppercase,
            # drop empty cells so our set is clean.
            for c in row:
                if c is None:
                    continue
                s = str(c).strip().upper()
                if s:
                    codes.add(s)
            n_cols = sum(1 for c in row if c is not None and str(c).strip())
            continue
        # Row 3+: data. Count non-empty rows for preview.
        if any(c is not None and str(c).strip() for c in row):
            data_row_count += 1

    return codes, n_cols, data_row_count, row1_codes


def detect(xlsx_path: str | Path, filename: str | None = None) -> DetectionResult:
    """Identify an uploaded file's MM role.

    Args:
        xlsx_path: filesystem path to the xlsx
        filename: what to report in the result (defaults to the path's filename)

    Returns a DetectionResult. role == "unknown" if no signature matched —
    the caller should prompt the user rather than guessing.
    """
    fname = filename or Path(xlsx_path).name
    try:
        codes, n_cols, data_rows, row1_codes = _read_row2_codes(xlsx_path)
    except Exception as e:
        # Bad/corrupt file: don't crash the upload, just mark unknown.
        # Hint covers the common causes: actual-CSV-renamed, zip corruption,
        # locked-by-Excel file, password protection.
        return DetectionResult(
            filename=fname, role="unknown",
            sap_fields=set(), column_count=0, data_rows=0,
            reason=f"Could not read the file: {e}",
            hint=("The file is not a valid .xlsx. Common causes: "
                  "(1) it's actually a .csv or .xls renamed with an .xlsx "
                  "extension — re-save it as xlsx from Excel; "
                  "(2) the file is still open in Excel — close it first; "
                  "(3) the file is password-protected — remove protection. "
                  "Download a template (button at top of this modal) to see "
                  "the expected format."),
        )

    # Empty file — no row 2 at all
    if n_cols == 0 and not codes:
        return DetectionResult(
            filename=fname, role="unknown",
            sap_fields=set(), column_count=0, data_rows=data_rows,
            reason="Row 2 is empty — no SAP field codes found",
            hint=("The file should have friendly labels in row 1 and SAP "
                  "field codes (MATNR, MBRSH, MTART etc.) in row 2. Your "
                  "row 2 is empty. Download a template to see the correct "
                  "layout."),
        )

    # Happy path — signature match
    for role, required, reason, min_cols in _SIGNATURES:
        if required.issubset(codes) and n_cols >= min_cols:
            return DetectionResult(
                filename=fname, role=role,
                sap_fields=codes, column_count=n_cols,
                data_rows=data_rows, reason=reason,
                hint="",
            )

    # No match. Build a targeted hint based on what we DID see.
    has_matnr = "MATNR" in codes
    row1_looks_like_sap_codes = (
        "MATNR" in row1_codes
        and ("MBRSH" in row1_codes or "UMREZ" in row1_codes or "BASE_TEXT" in row1_codes)
    )

    # Case 1: row 1 has the SAP codes — user swapped rows 1 and 2.
    # This is the single most specific, high-confidence recovery case
    # we can point at. Check it before the generic fallbacks.
    if row1_looks_like_sap_codes:
        reason = "Row 1 contains SAP codes; row 2 should but doesn't"
        hint = ("Looks like row 1 and row 2 got swapped. In MM templates: "
                "row 1 = friendly labels ('Material Number', 'Industry'), "
                "row 2 = SAP codes ('MATNR', 'MBRSH'), row 3+ = data. "
                "Insert an empty row at the top so your SAP codes end up on row 2.")
        return DetectionResult(
            filename=fname, role="unknown",
            sap_fields=codes, column_count=n_cols,
            data_rows=data_rows, reason=reason, hint=hint,
        )

    # Case 2: no MATNR anywhere — probably not an MM file at all.
    if not has_matnr and "MATNR" not in row1_codes:
        reason = "No MATNR column found — not a recognised MM file"
        hint = ("MM files must have a MATNR column. This file doesn't. "
                "Common mistakes: (1) wrong file type — did you mean to upload "
                "a Sales Order export? Master Data Validator is for MATERIAL "
                "MASTER, not sales. (2) the column was renamed — MATNR must "
                "appear literally as 'MATNR' in row 2. Download the template "
                "(button at top) to see what an MM file should look like.")
        return DetectionResult(
            filename=fname, role="unknown",
            sap_fields=codes, column_count=n_cols,
            data_rows=data_rows, reason=reason, hint=hint,
        )

    # Case 3: has MATNR but ambiguous — missing distinguishing codes.
    # Tell the user what we expected vs found.
    missing = []
    if "MBRSH" not in codes: missing.append("MBRSH (for Main)")
    if "UMREZ" not in codes: missing.append("UMREZ (for Alt UoM)")
    if "BASE_TEXT" not in codes: missing.append("BASE_TEXT (for Long Text)")
    reason = (
        f"Row 2 has MATNR but can't tell which MM file type this is. "
        f"Missing: {', '.join(missing)}"
    )
    hint = ("MATNR alone isn't enough to identify the file. A Main file "
            "also has MBRSH; an Alt UoM file also has UMREZ/UMREN; a Long "
            "Text file also has BASE_TEXT. At least one of these must be "
            "in row 2. If you exported only partial columns, re-export with "
            "all standard columns — or start from the blank template "
            "(button at top) and paste your data into it.")
    return DetectionResult(
        filename=fname, role="unknown",
        sap_fields=codes, column_count=n_cols,
        data_rows=data_rows, reason=reason, hint=hint,
    )


def detect_all(paths_and_names: list[tuple[str | Path, str]]) -> list[DetectionResult]:
    """Run detect() on a list of (path, filename) tuples.

    Order preserved. Useful for the upload endpoint which gets a list of
    SpooledTemporaryFile + filenames from FastAPI.
    """
    return [detect(p, n) for p, n in paths_and_names]


def check_detection_set(results: list[DetectionResult]) -> dict:
    """Given N DetectionResults, tell the caller what's missing / duplicate.

    Returns:
        {
          "roles_seen": {"main": ["file1.xlsx"], "alt_uom": [...]},
          "missing": ["main", "alt_uom"],     # roles NOT present
          "duplicates": ["main"],             # roles present more than once
          "unknown_files": ["weird.xlsx"],    # files that didn't match
          "ready_to_load": bool,              # main present, no dups, no unknowns
        }

    MAIN is the only strictly-required role. Alt-UoM and Long-Text are
    optional (a client might not have any alternate units or long text
    to migrate). So "ready_to_load" = main present AND no duplicates AND
    no unknowns. If the client explicitly confirms "no alt UoM file
    available" we still proceed with just main.
    """
    roles_seen: dict[str, list[str]] = {}
    unknown: list[str] = []
    for r in results:
        if r.role == "unknown":
            unknown.append(r.filename)
        else:
            roles_seen.setdefault(r.role, []).append(r.filename)

    all_roles = {"main", "alt_uom", "longtext"}
    missing = sorted(all_roles - roles_seen.keys())
    duplicates = sorted([role for role, files in roles_seen.items() if len(files) > 1])

    ready = (
        "main" in roles_seen
        and not duplicates
        and not unknown
    )

    return {
        "roles_seen": roles_seen,
        "missing": missing,
        "duplicates": duplicates,
        "unknown_files": unknown,
        "ready_to_load": ready,
    }
