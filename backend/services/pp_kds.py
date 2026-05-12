"""
PP KDS (reference catalogs) loader for BOM and Routing validation.

Catalogs we maintain
--------------------
- units_of_measure   — loaded from backend/kds/ISO_Unit_Of_Measure_tentative_file.xlsx
                       (230 ISO codes provided by the customer)
- bom_usages         — SAP-standard {1, 2, 3, 4, 5, 6} (Production,
                       Engineering, Universal, Plant Maintenance,
                       Sales/Distribution, Costing)
- bom_status         — SAP-standard {1, 2, 3, 4} (active, inactive,
                       blocked, planned)
- item_categories    — SAP-standard {L, N, R, T, D, M, K, I, …}
- control_keys       — SAP-standard {PP01, PP02, PP03, PP99} + customer Z-codes
- capacity_categories — SAP-standard {001, 002, 003, 004}
- routing_usages     — SAP-standard {1, 2, 3, 4, 5, 6}
- routing_status     — SAP-standard {1, 2, 3, 4}
- sequence_categories — SAP-standard {0, 1, 2, 3} (standard, parallel,
                       alternative, finalized)
- prt_categories     — SAP-standard {M, E, T, F} (Material, Equipment,
                       Tool, Document)
- plants             — empty by default; populated from a customer-
                       provided Plants KDS file when available

When no customer-specific catalog is provided, we run with SAP standards
only. The validator emits "warning" (not "error") for fields whose
catalog is empty so the user isn't spammed with false positives during
initial setup.

Customer note: the BOM_PHASE_1.xlsx data uses BASE_UNIT="BOX", which is
NOT in the ISO catalog (which has BX1/BX2/BX3 but not bare BOX). The
validator will flag this as a unit-not-in-catalog warning; the customer
either:
  a) Fixes the data to use a valid ISO code (BX, BX1, etc.), or
  b) Adds BOX to the customer's UoM catalog as a custom code.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


_KDS_DIR = Path(__file__).parent.parent / "kds"
_ISO_UOM_FILE = _KDS_DIR / "ISO_Unit_Of_Measure_tentative_file.xlsx"


@dataclass
class PpCatalogs:
    """All reference catalogs used by PP/Routing validators."""
    plants: set[str] = field(default_factory=set)
    units_of_measure: set[str] = field(default_factory=set)
    bom_usages: set[str] = field(default_factory=set)
    bom_status: set[str] = field(default_factory=set)
    item_categories: set[str] = field(default_factory=set)
    control_keys: set[str] = field(default_factory=set)
    capacity_categories: set[str] = field(default_factory=set)
    routing_usages: set[str] = field(default_factory=set)
    routing_status: set[str] = field(default_factory=set)
    sequence_categories: set[str] = field(default_factory=set)
    prt_categories: set[str] = field(default_factory=set)
    work_centres: dict[tuple[str, str], dict[str, Any]] = field(default_factory=dict)


# ─── SAP-standard defaults ─────────────────────────────────────────────
_STANDARD_BOM_USAGES = {"1", "2", "3", "4", "5", "6", "7", "8", "9"}
_STANDARD_BOM_STATUS = {"1", "2", "3", "4"}
_STANDARD_ITEM_CATEGORIES = {"L", "N", "R", "T", "D", "I", "M", "K"}
_STANDARD_CONTROL_KEYS = {
    "PP01", "PP02", "PP03", "PP99",
    "ZPP1", "ZPP2", "ZPP3", "ZPRT",
}
_STANDARD_CAPACITY_CATEGORIES = {"001", "002", "003", "004"}
_STANDARD_ROUTING_USAGES = {"1", "2", "3", "4", "5", "6"}
_STANDARD_ROUTING_STATUS = {"1", "2", "3", "4"}
_STANDARD_SEQUENCE_CATEGORIES = {"0", "1", "2", "3"}
_STANDARD_PRT_CATEGORIES = {"M", "E", "T", "F"}


# ─── ISO UoM loader ─────────────────────────────────────────────────────

def _load_iso_units(path: Path) -> set[str]:
    """Load the customer's ISO UoM catalog from the bundled xlsx.

    Format (column 1 of the Data sheet, row 1 = header):
      Internal UoM | 3-char external | 6-char.external | ...

    We use column 1 (Internal UoM) — that's what the customer's BOM
    files use in the COMP_UNIT and BASE_UNIT columns.
    """
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    units: set[str] = set()
    # The file has one sheet named "Data" with the header in row 1.
    sheet_name = "Data" if "Data" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            s = str(row[0]).strip()
            if s:
                units.add(s.upper())
    wb.close()
    return units


# ─── Public API ────────────────────────────────────────────────────────

_CATALOGS_CACHE: PpCatalogs | None = None


def load_pp_catalogs(plants_path: str | Path | None = None) -> PpCatalogs:
    """Build PpCatalogs with SAP-standard defaults plus the bundled ISO
    UoM file. If a Plants KDS file is provided, plants are loaded from
    it; otherwise plants is empty (validator skips plant-catalog checks
    until plants are populated).

    Cached across calls in the same process — the catalogs are immutable
    once loaded so caching is safe.
    """
    global _CATALOGS_CACHE
    if _CATALOGS_CACHE is not None and plants_path is None:
        return _CATALOGS_CACHE

    cats = PpCatalogs(
        units_of_measure=_load_iso_units(_ISO_UOM_FILE),
        bom_usages=set(_STANDARD_BOM_USAGES),
        bom_status=set(_STANDARD_BOM_STATUS),
        item_categories=set(_STANDARD_ITEM_CATEGORIES),
        control_keys=set(_STANDARD_CONTROL_KEYS),
        capacity_categories=set(_STANDARD_CAPACITY_CATEGORIES),
        routing_usages=set(_STANDARD_ROUTING_USAGES),
        routing_status=set(_STANDARD_ROUTING_STATUS),
        sequence_categories=set(_STANDARD_SEQUENCE_CATEGORIES),
        prt_categories=set(_STANDARD_PRT_CATEGORIES),
    )

    # Load Plants from a customer KDS file if provided. Format (any
    # sheet name; first column = WERKS code):
    if plants_path is not None:
        path = Path(plants_path)
        if path.exists():
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    s = str(row[0]).strip().upper()
                    if s:
                        cats.plants.add(s)
            wb.close()

    if plants_path is None:
        _CATALOGS_CACHE = cats

    return cats
