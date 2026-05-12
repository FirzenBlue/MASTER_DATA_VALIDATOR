"""Review xlsx export — round-trippable.

Goals (v53 spec from SME):
  1. Same column structure as the SOURCE upload — row 1 friendly
     labels, row 2 SAP codes, row 3+ data. NO banner row, no extra
     headers. The SME wants to fix it in Excel and re-upload to the
     validator without restructuring.
  2. Errors → red fill (one color, not a per-rule rainbow).
  3. Errors that have a KDS-derived suggested fix → green fill
     instead of red. The hover comment shows the suggested value.
  4. Summary sheet ("Review Notes") lists each error column by:
       - Excel column letter (e.g. "AL", "AS", "AM")
       - SAP code
       - Friendly label
       - Error count
       - Sample rule
     so the SME can navigate to the right column quickly and see
     the headline issue.

Why this differs from the v51 colored review:
  v51 used a 24-color palette per rule, a banner row, and a 3-row
  header (banner + labels + codes). This made re-uploading impossible
  because the file no longer matched the upload contract (which
  expects 2-row header). v53 strips the extras: same upload shape,
  red/green coloring only, hover comments preserved, summary moved
  out to a dedicated sheet.

Two output sheets:
  - "Review Notes": summary + legend + per-column error breakdown
  - "Materials"   : the round-trippable data sheet — open this in
                    Excel, fix the colored cells, save, and re-upload
                    via the same MM upload modal.
"""
from __future__ import annotations

from typing import Any
import io
import string
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment


# Colors
_FILL_ERROR        = "FFCCCC"  # generic light red — error, no fix or team
_FILL_FIXABLE      = "C8E6C9"  # light green — error with KDS-based fix
_FILL_INVALID_PINK = "FFC7CE"  # critical: cell has an INVALID value
                                # (per Color Guide row 13)
_FILL_HEADER_LABEL = "F1F5F9"  # light slate — friendly label row
_FILL_HEADER_CODE  = "E2E8F0"  # slightly darker slate — SAP code row

_COLOR_HEADER_TEXT = "0F172A"  # near-black for both header rows
_COLOR_CODE_TEXT   = "475569"  # mid-slate for SAP code row

# Team tints from the customer's Color Guide. When a mandatory field is
# blank we fill it with the team's tint so the SME knows which team owns
# the gap. Header row 1 also takes the saturated team color (matches the
# customer's master template visual).
try:
    from .mm_mandatory_by_team import (
        TEAM_HEADER_COLOR, TEAM_TINT_COLOR, INVALID_VALUE_FILL,
        MANDATORY_FIELDS_BY_TEAM, field_to_team,
    )
    _FIELD_TO_TEAM = field_to_team()
except ImportError:
    # Defensive: if the team module is missing for any reason, fall back
    # to plain red/green. The xlsx still works, just without team colors.
    TEAM_HEADER_COLOR = {}
    TEAM_TINT_COLOR = {}
    INVALID_VALUE_FILL = "FFC7CE"
    MANDATORY_FIELDS_BY_TEAM = {}
    _FIELD_TO_TEAM = {}


def _excel_col_letter(col_idx: int) -> str:
    """1 → 'A', 26 → 'Z', 27 → 'AA', 28 → 'AB', ..., 154 → 'EX'.

    Standard Excel column-letter conversion. We use this to give
    the SME a navigable reference in the summary ("errors in column
    AL" — which they can then jump to via Excel's name box).
    """
    s = ""
    while col_idx > 0:
        col_idx, r = divmod(col_idx - 1, 26)
        s = string.ascii_uppercase[r] + s
    return s


def build_mm_review_xlsx(merged, errors: list, main_loaded,
                          base_filename: str = "MM_Review",
                          alt_loaded=None, lt_loaded=None) -> tuple[str, bytes]:
    """Build the round-trippable review xlsx. Returns (filename, bytes).

    The Materials sheet matches the source upload's structure exactly:
      Row 1: friendly labels (e.g. "Material Number", "Material Description")
      Row 2: SAP codes (e.g. "MATNR", "MAKTX")
      Row 3+: one row per material × plant (mirroring source)

    Cells with errors are filled:
      red   (FFCCCC)  — error, no fix known
      green (C8E6C9)  — error with a suggested KDS-derived fix

    Hover comments on every error cell show:
      [SEVERITY] rule_name: message
      Suggested: <value> (when present)

    Cross-file sheets (v57): when `alt_loaded` / `lt_loaded` are provided,
    two additional sheets render the alt-uom and long-text source rows
    with cross-file error coloring (orphan MATNRs, sales-area mismatches).
    These sheets have the same 2-row-header shape so they're round-
    trippable too.

    The Review Notes sheet has:
      - Total error / warning / fixable counts
      - Per-column breakdown table: Excel letter, SAP code, friendly
        label, error count, sample rule
      - Color legend
    """
    wb = openpyxl.Workbook()
    # Materials must be the FIRST sheet for round-trip — the MM loader
    # picks `wb.sheetnames[0]` regardless of name. If Review Notes were
    # first, re-uploading this file would point the loader at the notes
    # sheet (which has different shape) and the upload would fail with
    # "this doesn't look like an MM main file".
    ws = wb.active
    ws.title = "Materials"
    notes = wb.create_sheet("Review Notes")

    # ── Resolve friendly labels per SAP code ─────────────────────────────
    # main_loaded.sap_fields is the column order; main_loaded.header_labels
    # is a parallel flat list of human-readable strings (e.g. "Material
    # Number" for MATNR, "Material Description" for MAKTX). They line up
    # 1:1 by index. (Earlier code mistakenly treated header_labels as a
    # list-of-rows and indexed [0] which gave a single character; this
    # uses [i] directly which gives the right label.)
    sap_fields = list(main_loaded.sap_fields)
    label_map: dict[str, str] = {}
    for i, code in enumerate(sap_fields):
        try:
            lbl = main_loaded.header_labels[i]
            label_map[code] = str(lbl) if lbl else code
        except (IndexError, TypeError, AttributeError):
            label_map[code] = code

    # ── Materials sheet: 2-row header + data ─────────────────────────────
    # Row 1: friendly labels (matches source row 1)
    # Row 2: SAP codes       (matches source row 2)
    # Row 3+: data           (matches source row 3+)
    # NOTE: This shape is REQUIRED by the upload format-check — putting a
    # banner row at top breaks re-upload. Don't add one.
    #
    # Row 1 cells colored by team when the field is owned by a team
    # (matches the customer's color-coded template visual: green for
    # Production fields, purple for Sales, blue for Planning, orange
    # for Finance). Other fields keep the neutral slate background.
    for i, code in enumerate(sap_fields, 1):
        c = ws.cell(row=1, column=i, value=label_map.get(code, code))
        team = _FIELD_TO_TEAM.get(code)
        if team and team in TEAM_HEADER_COLOR:
            c.fill = PatternFill("solid", fgColor=TEAM_HEADER_COLOR[team])
            c.font = Font(color="FFFFFF", bold=True, size=10)
        else:
            c.fill = PatternFill("solid", fgColor=_FILL_HEADER_LABEL)
            c.font = Font(color=_COLOR_HEADER_TEXT, bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        c = ws.cell(row=2, column=i, value=code)
        c.fill = PatternFill("solid", fgColor=_FILL_HEADER_CODE)
        c.font = Font(color=_COLOR_CODE_TEXT, italic=True, size=9, name="Consolas")
        c.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A3"

    # ── Index errors by (material_index, sap_field) ──────────────────────
    # MM validator emits errors with `row_idx` — the 0-based index into
    # merged.materials. (NOT the excel_row.) That maps directly to the
    # row position we'll write in the output below.
    #
    # IMPORTANT: only include errors whose `sheet` is "Materials". Cross-
    # file errors (sheet="AlternateUnits" / "LongText", from v57's
    # mm_cross_file_validator) reference rows in DIFFERENT files and
    # would land on the wrong main-sheet rows if mixed in here. They get
    # rendered on dedicated sheets below.
    code_to_col = {code: i + 1 for i, code in enumerate(sap_fields)}
    errors_by_cell: dict[tuple[int, str], list] = {}
    for e in errors:
        if e.sap_field is None:
            continue
        if (getattr(e, "sheet", None) or "Materials") != "Materials":
            continue
        key = (e.row_idx, e.sap_field)
        errors_by_cell.setdefault(key, []).append(e)

    # ── Data rows (start at row 3, matching source) ──────────────────────
    materials = list(merged.materials)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="F1F5F9"),
        bottom=Side(style="thin", color="F1F5F9"),
    )

    # Per-column error tracking, used for the summary sheet
    col_error_counts: dict[int, int] = {}    # col_idx → count
    col_fixable_counts: dict[int, int] = {}  # col_idx → count
    col_sample_rule: dict[int, str] = {}     # col_idx → first rule_id seen

    for ridx, mat in enumerate(materials):
        excel_row = ridx + 3   # row 1 labels, row 2 codes, data starts row 3
        values = mat.main.values if hasattr(mat.main, "values") else {}

        for code, col_idx in code_to_col.items():
            v = values.get(code)
            cell_value = v if (v is not None and v != "") else None
            c = ws.cell(row=excel_row, column=col_idx, value=cell_value)
            c.border = thin_border

            cell_errs = errors_by_cell.get((ridx, code), [])
            if not cell_errs:
                continue

            # Pick the most severe error (errors > warnings) — if multiple
            # errors land on the same cell, the worst one drives the color.
            worst = max(
                cell_errs,
                key=lambda e: 2 if (e.severity or "").lower() == "error" else 1
            )
            has_suggestion = bool(worst.suggested_value)

            # Color resolution priority:
            #   1. Green if any error on this cell has a KDS-derived
            #      suggested fix — SME just accepts the suggestion.
            #   2. Pink (INVALID_VALUE_FILL) if the cell has a value but
            #      it doesn't match a catalog or format rule. The Color
            #      Guide row 13 specifies this for "this cell contains
            #      an INVALID value".
            #   3. Team tint if the field is in MANDATORY_FIELDS_BY_TEAM
            #      and the value is blank. Color matches which team owns
            #      the field (Production green, Sales purple, Planning
            #      blue, Finance orange).
            #   4. Generic red as the catch-all.
            cell_value = (v if (v is not None and v != "") else None)
            team = _FIELD_TO_TEAM.get(code)

            if has_suggestion:
                fill_color = _FILL_FIXABLE
            elif cell_value is not None:
                # Cell has SOMETHING in it but it's invalid (e.g. wrong
                # length, wrong format, not in catalog). Pink.
                fill_color = _FILL_INVALID_PINK
            elif team and team in TEAM_TINT_COLOR:
                # Cell is blank AND the field is mandatory for a known
                # team. Use the team's tint.
                fill_color = TEAM_TINT_COLOR[team]
            else:
                # Blank, no team — fall back to generic red so it's still
                # noticeable.
                fill_color = _FILL_ERROR

            c.fill = PatternFill("solid", fgColor=fill_color)

            # Hover comment: rule name + message + suggested value.
            # When the cell is colored by team, prepend "[<TEAM>]" so the
            # SME sees who owns the missing field.
            comment_lines = []
            if team and cell_value is None and not has_suggestion:
                comment_lines.append(f"[{team} team] This field is mandatory.")
            for e in cell_errs:
                sev_tag = (e.severity or "info").upper()
                line = f"[{sev_tag}] {e.rule_name or e.rule_id}: {e.message}"
                if e.suggested_value:
                    line += f"\nSuggested: {e.suggested_value}"
                comment_lines.append(line)
            cmt = Comment("\n\n".join(comment_lines), "Validator")
            cmt.width = 380
            cmt.height = 100 + 24 * (len(comment_lines) - 1)
            c.comment = cmt

            # Track for summary
            col_error_counts[col_idx] = col_error_counts.get(col_idx, 0) + 1
            if has_suggestion:
                col_fixable_counts[col_idx] = col_fixable_counts.get(col_idx, 0) + 1
            if col_idx not in col_sample_rule:
                col_sample_rule[col_idx] = worst.rule_name or worst.rule_id

    # Best-effort column width (sample first 50 rows)
    for col_idx in range(1, len(sap_fields) + 1):
        col_letter = _excel_col_letter(col_idx)
        sap_code = sap_fields[col_idx - 1]
        widths = [len(label_map.get(sap_code, sap_code)), len(sap_code)]
        for r_idx in range(min(50, len(materials))):
            mat = materials[r_idx]
            v = mat.main.values.get(sap_code) if hasattr(mat.main, "values") else None
            if v is not None:
                widths.append(min(36, len(str(v))))
        ws.column_dimensions[col_letter].width = max(8, min(36, max(widths) + 2))

    # ── Cross-file sheets (v57) ──────────────────────────────────────────
    # When the alt-uom and/or long-text companion files were uploaded,
    # render their source rows on dedicated sheets with the same 2-row-
    # header round-trip shape, and color any cells that triggered a
    # cross-file error (orphan MATNRs, sales-area mismatches). The SME
    # sees errors highlighted in the same review file regardless of
    # which file the data came from.
    cross_file_summary: list[tuple[str, int, int]] = []  # (sheet_name, rows, errors)
    if alt_loaded is not None and alt_loaded.rows:
        n_errs = _render_cross_file_sheet(
            wb, sheet_name="Alternate Units",
            loaded_file=alt_loaded,
            errors=errors,
            error_sheet_label="AlternateUnits",
        )
        cross_file_summary.append(("Alternate Units", len(alt_loaded.rows), n_errs))

    if lt_loaded is not None and lt_loaded.rows:
        n_errs = _render_cross_file_sheet(
            wb, sheet_name="Long Text",
            loaded_file=lt_loaded,
            errors=errors,
            error_sheet_label="LongText",
        )
        cross_file_summary.append(("Long Text", len(lt_loaded.rows), n_errs))

    # ── Review Notes sheet ───────────────────────────────────────────────
    notes.cell(row=1, column=1, value="MM Validation Review").font = Font(
        bold=True, size=18, color="0F172A"
    )

    msg = (
        "This file has the same column structure as your upload — fix the "
        "colored cells in the Materials sheet, save, and re-upload via the "
        "MM upload screen. Hover any colored cell to see the issue."
    )
    notes.cell(row=2, column=1, value=msg).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    notes.row_dimensions[2].height = 50
    notes.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    # Legend — covers all the colors the SME might see in the Materials sheet.
    # Order: green (fixable) → pink (invalid value) → 4 team tints (mandatory
    # field blank) → red (catch-all). Matches the order they appear in the
    # color-resolution priority list inside the cell-fill loop.
    notes.cell(row=4, column=1, value="Legend").font = Font(bold=True, size=12)
    legend_rows = [
        (5, _FILL_FIXABLE,
         "Green — error in this cell, AND the validator has a suggested "
         "value from the KDS or a default rule. Hover the cell to see the "
         "suggestion; you can accept it or override."),
        (6, _FILL_INVALID_PINK,
         "Pink — this cell has a value but it's INVALID (wrong format, "
         "wrong length, or not in the KDS catalog). Replace with a valid value."),
    ]
    # Per-team mandatory tints
    next_row = 7
    for team, tint in TEAM_TINT_COLOR.items():
        legend_rows.append((
            next_row,
            tint,
            f"{team} (tint) — mandatory field blank. Hover the cell to see "
            f"which field; the {team} team owns this value."
        ))
        next_row += 1
    legend_rows.append((
        next_row, _FILL_ERROR,
        "Red — generic error, no team or fix association. Less common; "
        "hover for details."
    ))
    summary_start_row = next_row + 2  # leave a blank row before Summary
    for row, color, text in legend_rows:
        c1 = notes.cell(row=row, column=1, value="")
        c1.fill = PatternFill("solid", fgColor=color)
        c1.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        notes.cell(row=row, column=2, value=text).alignment = Alignment(
            wrap_text=True, vertical="center"
        )
        notes.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        notes.row_dimensions[row].height = 32

    # Summary stats — row index now depends on legend size (4 teams + 3 generic = 7 rows).
    n_err = sum(1 for e in errors if (e.severity or "").lower() == "error")
    n_warn = sum(1 for e in errors if (e.severity or "").lower() == "warning")
    n_fixable = sum(1 for e in errors if e.suggested_value)
    notes.cell(row=summary_start_row, column=1, value="Summary").font = Font(bold=True, size=12)
    notes.cell(row=summary_start_row + 1, column=1, value="Materials:").font = Font(bold=True)
    notes.cell(row=summary_start_row + 1, column=2, value=len(materials))
    notes.cell(row=summary_start_row + 2, column=1, value="Total errors:").font = Font(bold=True)
    notes.cell(row=summary_start_row + 2, column=2, value=n_err).font = Font(color="C62828", bold=True)
    notes.cell(row=summary_start_row + 3, column=1, value="Of those, fixable from KDS:").font = Font(bold=True)
    notes.cell(row=11, column=2, value=n_fixable).font = Font(color="2E7D32", bold=True)
    notes.cell(row=summary_start_row + 4, column=1, value="Warnings:").font = Font(bold=True)
    notes.cell(row=summary_start_row + 4, column=2, value=n_warn).font = Font(color="EF6C00", bold=True)

    # Per-column breakdown table — starts 2 rows below the last summary stat
    breakdown_title_row = summary_start_row + 6
    notes.cell(row=breakdown_title_row, column=1, value="Errors by column").font = Font(
        bold=True, size=12
    )
    header_row = breakdown_title_row + 1
    # Per-SME spec: don't include a "Team" column in the breakdown.
    # Team coloring still happens on the Materials sheet's header row 1
    # (saturated team color) and on data cells (team tints) — the SME
    # sees who owns a field at a glance there. The summary table is just
    # the error counts per column, no team annotation.
    headers = [
        ("Excel column", 14),
        ("SAP code", 16),
        ("Field name", 36),
        ("Errors", 10),
        ("Fixable (green)", 16),
        ("Sample rule", 50),
    ]
    for i, (text, width) in enumerate(headers, 1):
        c = notes.cell(row=header_row, column=i, value=text)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="334155")
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Sort columns by error count descending — biggest issues at top
    sorted_cols = sorted(
        col_error_counts.items(), key=lambda x: -x[1]
    )
    body_row = header_row + 1
    thin_b = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for col_idx, count in sorted_cols:
        sap_code = sap_fields[col_idx - 1]
        excel_letter = _excel_col_letter(col_idx)
        friendly = label_map.get(sap_code, sap_code)
        fixable = col_fixable_counts.get(col_idx, 0)
        sample = col_sample_rule.get(col_idx, "")

        cells = [
            notes.cell(row=body_row, column=1, value=excel_letter),
            notes.cell(row=body_row, column=2, value=sap_code),
            notes.cell(row=body_row, column=3, value=friendly),
            notes.cell(row=body_row, column=4, value=count),
            notes.cell(row=body_row, column=5, value=fixable),
            notes.cell(row=body_row, column=6, value=sample),
        ]
        for c in cells:
            c.border = thin_b
        # Excel column letter cell — monospace, slate background, centered
        cells[0].font = Font(name="Consolas", bold=True, color="0F172A")
        cells[0].fill = PatternFill("solid", fgColor="F1F5F9")
        cells[0].alignment = Alignment(horizontal="center", vertical="center")
        # SAP code — monospace
        cells[1].font = Font(name="Consolas", color="475569")
        # Errors count — emphasize
        cells[3].font = Font(bold=True, color="C62828")
        cells[3].alignment = Alignment(horizontal="right")
        # Fixable count — green if any, otherwise muted
        if fixable > 0:
            cells[4].font = Font(bold=True, color="2E7D32")
        else:
            cells[4].font = Font(color="94A3B8")
        cells[4].alignment = Alignment(horizontal="right")

        body_row += 1

    # Set column widths for the summary sheet
    for i, (_, width) in enumerate(headers, 1):
        notes.column_dimensions[_excel_col_letter(i)].width = width

    notes.freeze_panes = f"A{header_row + 1}"

    # ── Serialize ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    out_name = f"{base_filename}_Review.xlsx"
    return out_name, buf.getvalue()


def _render_cross_file_sheet(wb, *, sheet_name: str, loaded_file,
                              errors: list, error_sheet_label: str) -> int:
    """Render an alt-uom or long-text source file as a colored review sheet.

    Args:
        wb: openpyxl workbook to add the sheet to.
        sheet_name: human-readable name for the new sheet ("Alternate Units"
                    or "Long Text"). Doubles as round-trip target — if the
                    SME re-uploads this xlsx with this sheet exposed, the
                    loader can read it back.
        loaded_file: LoadedFile from mm_loader (rows, sap_fields,
                     header_labels — same shape as the main loader's output).
        errors: full validator error list. Filtered here to errors whose
                `sheet` matches `error_sheet_label`.
        error_sheet_label: "AlternateUnits" or "LongText" — the value the
                           cross-file validator stamps on Error.sheet for
                           rows in this file.

    Returns: number of cells colored (i.e. error rows in this file).

    Sheet layout (matches the source upload's 2-row header):
      Row 1: friendly labels from loaded_file.header_labels
      Row 2: SAP codes from loaded_file.sap_fields
      Row 3+: data, with cells colored where cross-file errors fired

    Cell coloring: red for errors, light pink for warnings (sales-area-
    unverified). Comments hover-show the rule + suggested fix where one
    exists. Same color scheme as the main Materials sheet for consistency.
    """
    ws = wb.create_sheet(sheet_name)
    sap_fields = list(loaded_file.sap_fields)
    labels = list(loaded_file.header_labels) if loaded_file.header_labels else sap_fields

    # Row 1: friendly labels. Slate background — these sheets aren't
    # team-owned (the team color scheme applies to MAIN-file fields only).
    for i, code in enumerate(sap_fields, 1):
        lbl = labels[i - 1] if i - 1 < len(labels) else code
        c = ws.cell(row=1, column=i, value=lbl)
        c.fill = PatternFill("solid", fgColor=_FILL_HEADER_LABEL)
        c.font = Font(color=_COLOR_HEADER_TEXT, bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")

        c = ws.cell(row=2, column=i, value=code)
        c.fill = PatternFill("solid", fgColor=_FILL_HEADER_CODE)
        c.font = Font(color=_COLOR_CODE_TEXT, italic=True, size=9, name="Consolas")
        c.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A3"

    # Index errors for this sheet by (row_idx, sap_field). row_idx is the
    # 0-based index into loaded_file.rows — that's how the cross-file
    # validator stores it.
    code_to_col = {code: i + 1 for i, code in enumerate(sap_fields)}
    sheet_errors_by_cell: dict[tuple[int, str], list] = {}
    for e in errors:
        if (getattr(e, "sheet", None) or "") != error_sheet_label:
            continue
        if e.sap_field is None:
            continue
        # Only color if the field is present in this sheet's columns.
        if e.sap_field not in code_to_col:
            continue
        sheet_errors_by_cell.setdefault((e.row_idx, e.sap_field), []).append(e)

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="F1F5F9"),
        bottom=Side(style="thin", color="F1F5F9"),
    )

    n_colored = 0
    for ridx, row in enumerate(loaded_file.rows):
        excel_row = ridx + 3   # row 1 labels, row 2 codes
        for code, col_idx in code_to_col.items():
            v = row.values.get(code) if hasattr(row, "values") else None
            cell_value = v if (v is not None and v != "") else None
            c = ws.cell(row=excel_row, column=col_idx, value=cell_value)
            c.border = thin_border

            cell_errs = sheet_errors_by_cell.get((ridx, code), [])
            if not cell_errs:
                continue
            n_colored += 1

            # Most-severe wins for fill choice
            worst = max(cell_errs,
                        key=lambda e: 2 if (e.severity or "").lower() == "error" else 1)
            sev = (worst.severity or "error").lower()

            if worst.suggested_value:
                c.fill = PatternFill("solid", fgColor=_FILL_FIXABLE)
            elif sev == "warning":
                # Lighter tint for warnings (sales-area unverified) so SMEs
                # can spot real errors first.
                c.fill = PatternFill("solid", fgColor=_FILL_INVALID_PINK)
            else:
                c.fill = PatternFill("solid", fgColor=_FILL_ERROR)

            # Hover comment
            comment_lines = []
            for e in cell_errs:
                tag = (e.severity or "info").upper()
                line = f"[{tag}] {e.rule_name or e.rule_id}: {e.message}"
                if e.suggested_value:
                    line += f"\nSuggested: {e.suggested_value}"
                comment_lines.append(line)
            cmt = Comment("\n\n".join(comment_lines), "Validator")
            cmt.width = 380
            cmt.height = 100 + 24 * (len(comment_lines) - 1)
            c.comment = cmt

    # Column widths (best-effort sampling)
    for col_idx in range(1, len(sap_fields) + 1):
        col_letter = _excel_col_letter(col_idx)
        code = sap_fields[col_idx - 1]
        widths = [len(labels[col_idx - 1] if col_idx - 1 < len(labels) else code), len(code)]
        for r_idx in range(min(30, len(loaded_file.rows))):
            v = loaded_file.rows[r_idx].values.get(code) if hasattr(loaded_file.rows[r_idx], "values") else None
            if v is not None:
                widths.append(min(40, len(str(v))))
        ws.column_dimensions[col_letter].width = max(10, min(40, max(widths) + 2))

    return n_colored

