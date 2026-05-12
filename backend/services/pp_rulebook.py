"""
PP rulebook — declarative validation rules for BOM and Routing fields.

Source of truth: the SAP S/4HANA LTMC standard templates that ship with
the codebase under backend/pp_templates/:
  - Source_data_for_Material_BOM.xlsx (BOM, .xlsx — openpyxl)
  - Source_data_for_Routing.xml       (Routing, SpreadsheetML — lxml)

We extract:
  - sap_field codes from row 5
  - max_length from the ETE format spec on row 6 (e.g. "ETE;80;0;C;80;0"
    → length=80, decimals=0)
  - mandatory flag from a trailing "*" on the first line of the row-8
    description ("Material Number*\\n\\nType: Text\\nLength: 80")
  - friendly_label from the first line of row 8 (the part before "\\n")

Catalog references (catalog="plants", catalog="units_of_measure", etc.)
are added programmatically AFTER template extraction — these are
business decisions, not in the template:
  - WERKS / WERKS_MAT / WERKS_WORK_CNTR / WERKS_ROOT → plants
  - BASE_UNIT / COMP_UNIT / EMPTIES_UOM / PLNME / MEINH /
    VGE01..VGE06 / MGEINH / EWEINH / EHOFFB / EHOFFE → units_of_measure
  - STLAN → bom_usages (1=Production, 2=Engineering, …; SAP standard)
  - STLST → bom_status (1=active, 2=inactive, 3=blocked, 4=planned)
  - POSTP → item_categories (L, N, R, T, D, M, K, …)
  - STEUS → control_keys (PP01, PP02, PP03, …)
  - VERWE → routing_usages (1=Production, …)
  - SPRAS → languages (EN, DE, etc.) — TODO: catalog
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


# Where the templates live in the codebase
_TEMPLATES_DIR = Path(__file__).parent.parent / "pp_templates"
_BOM_TEMPLATE = _TEMPLATES_DIR / "Source_data_for_Material_BOM.xlsx"
_ROUTING_TEMPLATE = _TEMPLATES_DIR / "Source_data_for_Routing.xml"

# Where the SAP codes / format / description live in the template
_CODE_ROW = 5
_ETE_ROW = 6
_DESC_ROW = 8


# Sheets we extract rules for. Anything else in the template (Introduction,
# Field List) is ignored. Matches PP_SHEETS_ORDER in pp_loader and
# ROUTING_SHEETS_ORDER in routing_loader.
_BOM_SHEETS = (
    "BOM Header", "BOM Item", "BOM Subitem",
    "Global Dependency", "Local Dependency", "Local Dependency Description",
    "Documentation of Dependency", "Sources of Local Dependency",
    "BOM Item Document Assignment", "BOM Header Document Assignment",
)
_ROUTING_SHEETS = (
    "Routing Group", "Task List - Header", "Material Task List Assignment",
    "Sequences", "Operations", "Component Assignment",
    "Production Resources and Tools", "Sub Operations",
    "Inspection Plan Characteristic",
    "Global Dependency", "Local Dependency", "Local Dependency Description",
    "Documentation of Dependency", "Sources of Local Dependency",
)


# ─── FieldRule ──────────────────────────────────────────────────────────
@dataclass
class FieldRule:
    sheet: str
    sap_field: str
    friendly_label: str
    is_mandatory: bool = False
    max_length: int | None = None
    decimal_places: int = 0
    field_kind: str = "text"  # text | number | date | code | unit
    catalog: str | None = None  # PpCatalogs attribute name
    description: str = ""
    rule_id: str = ""

    def __post_init__(self):
        if not self.rule_id:
            slug = self.sheet.lower().replace(" ", "_").replace("-", "_")
            self.rule_id = f"pp_{slug}_{self.sap_field.lower()}"


@dataclass
class Rulebook:
    rule_index: dict[str, dict[str, FieldRule]] = field(default_factory=dict)
    sap_fields_by_sheet: dict[str, list[str]] = field(default_factory=dict)

    def get(self, sheet: str, sap_field: str) -> FieldRule | None:
        return self.rule_index.get(sheet, {}).get(sap_field)


# ─── ETE parser ─────────────────────────────────────────────────────────
# ETE format: "ETE;80;0;C;80;0" or "ENU;13;3;P;13;3"
# Position 1 = max length (chars for text, total digits for number)
# Position 2 = decimal places (0 for text)
# Position 3 = data type code (C=character, D=date, N=number, P=packed, g=long-text)
_ETE_RE = re.compile(r"^([A-Z]+);(\d+);(\d+);([A-Za-z])")

def _parse_ete(ete: Any) -> tuple[int | None, int, str]:
    """Return (max_length, decimal_places, type_code) or (None, 0, "?")."""
    if not ete:
        return None, 0, "?"
    s = str(ete).strip()
    m = _ETE_RE.match(s)
    if not m:
        return None, 0, "?"
    type_prefix, length_s, dec_s, type_code = m.groups()
    try:
        length = int(length_s)
    except ValueError:
        length = None
    try:
        dec = int(dec_s)
    except ValueError:
        dec = 0
    return length, dec, type_code


def _extract_label_and_mandatory(desc: Any) -> tuple[str, bool]:
    """Pull the short label (first line) and mandatory flag from the
    row-8 description cell. Mandatory if the first line ends with "*"."""
    if desc is None:
        return "", False
    first_line = str(desc).split("\n", 1)[0].strip()
    mandatory = first_line.endswith("*")
    label = first_line.rstrip("*").strip()
    return label, mandatory


def _field_kind_from_type(type_code: str, ete_dec: int) -> str:
    """Map an ETE type letter to the validator/UI's field kind."""
    if type_code == "D":
        return "date"
    if type_code in ("N", "P"):
        return "number"
    if type_code == "g":
        return "longtext"
    return "text"


# ─── Catalog assignment overlay ─────────────────────────────────────────
# Field-to-catalog mappings. These are business decisions, not in the
# template. Reviewed against the SAP S/4 LTMC for Material BOM and
# Routing standard guides.
_CATALOG_BY_FIELD: dict[str, str] = {
    # Plants — wherever a plant column appears
    "WERKS":            "plants",
    "WERKS_MAT":        "plants",
    "WERKS_ROOT":       "plants",
    "WERKS_WORK_CNTR":  "plants",
    "FHWRK":            "plants",
    "QPMK_ZAEHL":       "plants",
    "QMTB_WERKS":       "plants",
    "AUSWMGWRK1":       "plants",

    # Units of measure (ISO format)
    "BASE_UNIT":        "units_of_measure",
    "COMP_UNIT":        "units_of_measure",
    "EMPTIES_UOM":      "units_of_measure",
    "PLNME":            "units_of_measure",
    "MEINH":            "units_of_measure",
    "VGE01":            "units_of_measure",
    "VGE02":            "units_of_measure",
    "VGE03":            "units_of_measure",
    "VGE04":            "units_of_measure",
    "VGE05":            "units_of_measure",
    "VGE06":            "units_of_measure",
    "ZEILM":            "units_of_measure",
    "ZEILP":            "units_of_measure",
    "ZEIWN":            "units_of_measure",
    "ZEIWM":            "units_of_measure",
    "ZEITN":            "units_of_measure",
    "ZEITM":            "units_of_measure",
    "ZEIMB":            "units_of_measure",
    "ZEIMU":            "units_of_measure",
    "MGEINH":           "units_of_measure",
    "EWEINH":           "units_of_measure",
    "EHOFFB":           "units_of_measure",
    "EHOFFE":           "units_of_measure",
    "PROBEMGEH":        "units_of_measure",
    "MASSEINHSW":       "units_of_measure",

    # BOM usage / status / item category
    "STLAN":            "bom_usages",
    "STLAN_HDR":        "bom_usages",
    "BOM_TYPE":         "bom_usages",
    "STLST":            "bom_status",
    "POSTP":            "item_categories",

    # Routing-specific
    "VERWE":            "routing_usages",
    "STATU":            "routing_status",
    "STEUS":            "control_keys",
    "CAPID":            "capacity_categories",
    "FLGAT":            "sequence_categories",
    "PSNFH":            "prt_categories",
}


# ─── Template parsers ──────────────────────────────────────────────────

def _parse_bom_template(path: Path) -> dict[str, list[FieldRule]]:
    """Read the BOM xlsx template, return {sheet: [FieldRule, ...]}."""
    wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)
    rules: dict[str, list[FieldRule]] = {}

    for sheet_name in _BOM_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        codes = [ws.cell(row=_CODE_ROW, column=c).value
                 for c in range(1, ws.max_column + 1)]
        etes = [ws.cell(row=_ETE_ROW, column=c).value
                for c in range(1, ws.max_column + 1)]
        descs = [ws.cell(row=_DESC_ROW, column=c).value
                 for c in range(1, ws.max_column + 1)]

        sheet_rules: list[FieldRule] = []
        for i, code in enumerate(codes):
            if code is None or not str(code).strip():
                continue
            sap = str(code).strip().upper()
            ete = etes[i] if i < len(etes) else None
            desc = descs[i] if i < len(descs) else None
            length, dec, type_code = _parse_ete(ete)
            label, mand = _extract_label_and_mandatory(desc)
            kind = _field_kind_from_type(type_code, dec)
            # Long-text fields have length 0 in the ETE — translate to
            # None (unlimited) so the validator doesn't complain.
            mx = length if (length and length > 0) else None
            sheet_rules.append(FieldRule(
                sheet=sheet_name,
                sap_field=sap,
                friendly_label=label or sap,
                is_mandatory=mand,
                max_length=mx,
                decimal_places=dec,
                field_kind=kind,
                catalog=_CATALOG_BY_FIELD.get(sap),
                description=str(desc).strip() if desc else "",
            ))
        rules[sheet_name] = sheet_rules

    wb.close()
    return rules


def _parse_routing_template(path: Path) -> dict[str, list[FieldRule]]:
    """Read the Routing SpreadsheetML XML template, return rules dict.

    The template has malformed inline tags (<LS> etc.) inside text
    nodes, so we use lxml's recover mode."""
    from lxml import etree

    parser = etree.XMLParser(recover=True)
    tree = etree.parse(str(path), parser)
    root = tree.getroot()

    ns = "{urn:schemas-microsoft-com:office:spreadsheet}"

    def sheet_rows(ws):
        """Build {row_idx: [cell_text, ...]} from a Worksheet element,
        respecting ss:Index attributes for skipped rows/cells."""
        table = ws.find(f"{ns}Table")
        if table is None:
            return {}
        rows_dict: dict[int, list] = {}
        cur_row = 0
        for row in table.findall(f"{ns}Row"):
            idx = row.attrib.get(f"{ns}Index")
            cur_row = int(idx) if idx else cur_row + 1
            cells: list = []
            cur_col = 0
            for cell in row.findall(f"{ns}Cell"):
                cidx = cell.attrib.get(f"{ns}Index")
                cur_col = int(cidx) if cidx else cur_col + 1
                while len(cells) < cur_col - 1:
                    cells.append(None)
                data = cell.find(f"{ns}Data")
                cells.append(data.text if data is not None else None)
            rows_dict[cur_row] = cells
        return rows_dict

    rules: dict[str, list[FieldRule]] = {}
    for ws in root.findall(f"{ns}Worksheet"):
        sheet_name = ws.attrib.get(f"{ns}Name", "")
        if sheet_name not in _ROUTING_SHEETS:
            continue
        rows = sheet_rows(ws)
        codes = rows.get(_CODE_ROW, [])
        etes = rows.get(_ETE_ROW, [])
        descs = rows.get(_DESC_ROW, [])

        sheet_rules: list[FieldRule] = []
        for i, code in enumerate(codes):
            if not code:
                continue
            sap = str(code).strip().upper()
            ete = etes[i] if i < len(etes) else None
            desc = descs[i] if i < len(descs) else None
            length, dec, type_code = _parse_ete(ete)
            label, mand = _extract_label_and_mandatory(desc)
            kind = _field_kind_from_type(type_code, dec)
            mx = length if (length and length > 0) else None
            sheet_rules.append(FieldRule(
                sheet=sheet_name,
                sap_field=sap,
                friendly_label=label or sap,
                is_mandatory=mand,
                max_length=mx,
                decimal_places=dec,
                field_kind=kind,
                catalog=_CATALOG_BY_FIELD.get(sap),
                description=str(desc).strip() if desc else "",
            ))
        rules[sheet_name] = sheet_rules

    return rules


# ─── Singleton accessor ────────────────────────────────────────────────
_RULEBOOK_CACHE: Rulebook | None = None


def get_rulebook() -> Rulebook:
    """Return the singleton rulebook. Built once per process from the
    LTMC templates bundled in backend/pp_templates/."""
    global _RULEBOOK_CACHE
    if _RULEBOOK_CACHE is not None:
        return _RULEBOOK_CACHE

    book = Rulebook()

    # Parse BOM template
    if _BOM_TEMPLATE.exists():
        bom_rules = _parse_bom_template(_BOM_TEMPLATE)
        for sheet_name, sheet_rules in bom_rules.items():
            book.rule_index[sheet_name] = {r.sap_field: r for r in sheet_rules}
            book.sap_fields_by_sheet[sheet_name] = [r.sap_field for r in sheet_rules]

    # Parse Routing template
    if _ROUTING_TEMPLATE.exists():
        rt_rules = _parse_routing_template(_ROUTING_TEMPLATE)
        for sheet_name, sheet_rules in rt_rules.items():
            # BOM and Routing both have sheets named "Global Dependency",
            # "Local Dependency", etc. The rulebook keys are flat — to
            # keep them separate we prefix Routing's shared-name sheets
            # with "Routing · ". The validator does the same prefixing
            # when looking up rules for a Routing sheet.
            #
            # NOTE: sheets unique to Routing (Routing Group, Operations,
            # …) keep their natural names since there's no collision.
            if sheet_name in book.rule_index:
                key = f"Routing · {sheet_name}"
            else:
                key = sheet_name
            book.rule_index[key] = {r.sap_field: r for r in sheet_rules}
            book.sap_fields_by_sheet[key] = [r.sap_field for r in sheet_rules]

    _RULEBOOK_CACHE = book
    return book


# Aliases retained for compatibility with code that imported these names
# during the rebuild — both empty now since rules come from templates.
ALL_BOM_RULES: tuple = ()
ALL_ROUTING_RULES: tuple = ()
