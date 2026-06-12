"""
Master Data Validator — Backend API
FastAPI with auth, module-gated repository, validation engine.
"""
from __future__ import annotations

import io
import json
import os
import sys
import time
import datetime as dt_module
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, HTTPException, Request, Response, UploadFile, Depends, Form, Cookie
from fastapi.responses import FileResponse, HTMLResponse, Response as FastResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from services.xml_engine import parse as parse_xml, write as write_xml, Workbook
from services.validator import validate, Error
from services.decision_engine import group_errors, summarize, Decision
from services.applier import apply_decision, apply_single_edit, delete_single_row, undo_entry, AuditEntry
from services import auth as auth_svc
from services import repository as repo
from services import audit_log
from services import db as db_svc
from services import kds_reference as kds
# MM module services + routes (module-scoped; Customer module doesn't import these)
from services.mm_loader import LOADERS_BY_ROLE as MM_LOADERS
from services.mm_merger import merge as mm_merge
from services.mm_validator import validate_mm
from mm_routes import router as mm_router, _get_mm_catalogs
from pp_routes import router as pp_router
# from grievance_routes import router as grievance_router
from chatbot_routes import router as chatbot_router

app = FastAPI(title="Master Data Validator", version="0.3")

# MM module routes live in a separate file to keep main.py navigable.
# See backend/mm_routes.py for the upload / template / format-check routes.
app.include_router(mm_router)

# PP module routes — BOM and Routing upload, format-check, manifest+chunk
# export. See backend/pp_routes.py.
app.include_router(pp_router)

#app.include_router(grievance_router)
app.include_router(chatbot_router)


@app.on_event("startup")
def on_startup():
    """Initialize database schema and seed users on first run."""
    try:
        ok, info = db_svc.ping()
        if not ok:
            print(f"[FATAL] Cannot connect to Postgres: {info}")
            print(f"[HINT] Set DATABASE_URL env var or ensure Postgres is running on localhost:5432")
            print(f"[HINT] Default: postgresql://postgres:Lumbini@localhost:5432/masterdata")
            raise RuntimeError("Postgres unreachable — see logs")
        print(f"[db] Connected: {info[:60]}")
        db_svc.init_schema()
        # Migrate any legacy JSON data (one-time)
        db_svc.migrate_from_json_if_present(repo.STORAGE_ROOT)
        db_svc.seed_default_users()
        print("[db] Schema + seed ready")
    except Exception as e:
        print(f"[startup error] {e}")
        raise

ROOT = Path(__file__).parent.parent
FRONTEND_DIR = ROOT / "frontend"
app.mount("/static", StaticFiles(directory=FRONTEND_DIR / "static"), name="static")


# Prevent browser caching of app.js / index.html / main.css during active development.
# Stale cached JS has been a recurring pain point — force a fresh fetch every time.
@app.middleware("http")
async def no_cache_for_html_and_js(request, call_next):
    response = await call_next(request)
    path = request.url.path
    if (path == "/"
            or path.endswith(".html")
            or path.endswith(".js")
            or path.endswith(".css")):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


# ── In-memory per-user working sessions ─────────────────────────────────────
# Key: session_token → working state (loaded workbook + errors + decisions)
# This is separate from auth sessions (which hold user identity).
#
# HARDENING for production multi-user:
# - Every state access updates last_accessed.
# - _sweep_stale_sessions() called opportunistically to free stale workbooks.
# - MAX_LOADED_WORKBOOKS caps total memory — oldest evicted first.
# - Explicit DELETE /api/session/close lets users voluntarily free memory.
WORKING: dict[str, dict] = {}

# Eviction policy constants — tuned for 10 users / 16 GB VM / 28 MB files.
# Override via environment variables at boot time.
import os as _os_env
SESSION_TTL_SECONDS = int(_os_env.environ.get("MDV_SESSION_TTL", "1800"))  # 30 min default
MAX_LOADED_WORKBOOKS = int(_os_env.environ.get("MDV_MAX_WORKBOOKS", "15"))  # hard cap
SWEEP_EVERY_SECONDS = int(_os_env.environ.get("MDV_SWEEP_INTERVAL", "60"))

import time as _time
_LAST_SWEEP = [0.0]  # mutable holder for last-sweep timestamp


def _sweep_stale_sessions():
    """Free workbooks untouched for SESSION_TTL_SECONDS.
    Called opportunistically on every state access (throttled to once a minute).
    Also enforces MAX_LOADED_WORKBOOKS via LRU eviction."""
    now = _time.time()
    if now - _LAST_SWEEP[0] < SWEEP_EVERY_SECONDS:
        return
    _LAST_SWEEP[0] = now

    # Pass 1 — TTL-based eviction (only evict ones with a loaded workbook,
    # keeping empty session shells lets users reconnect without re-login)
    to_drop = []
    for token, state in WORKING.items():
        if state.get("workbook") is None:
            continue  # nothing to free
        last = state.get("last_accessed", 0.0)
        if now - last > SESSION_TTL_SECONDS:
            to_drop.append(token)
    for token in to_drop:
        _free_workbook_in_session(WORKING.get(token, {}))

    # Pass 2 — if still over the cap, LRU-evict until under
    loaded = [(t, s.get("last_accessed", 0.0)) for t, s in WORKING.items() if s.get("workbook")]
    if len(loaded) > MAX_LOADED_WORKBOOKS:
        loaded.sort(key=lambda x: x[1])  # oldest first
        for token, _ in loaded[: len(loaded) - MAX_LOADED_WORKBOOKS]:
            _free_workbook_in_session(WORKING.get(token, {}))


def _free_workbook_in_session(state: dict) -> None:
    """Clear the heavy pieces of a working session; keep the shell so the
    user's auth remains intact and they can re-open another file without
    re-logging in. Persists any unflushed edits before releasing the workbook
    so the TTL sweeper never drops unsaved work."""
    if not state:
        return
    if state.get("dirty") and state.get("workbook") and state.get("file_id"):
        try:
            xml_bytes = write_xml(state["workbook"])
            repo.save_working_copy(state["file_id"], xml_bytes)
        except Exception as e:
            print(f"[sweep] flush failed for {state.get('file_id')}: {e}")
    state["workbook"] = None
    state["errors"] = []
    state["decisions"] = []
    state["audit_log"] = []
    state["resolved_decisions"] = set()
    state["accepted_errors"] = set()
    state["file_id"] = None
    state["filename"] = ""
    state["module"] = None
    state["dirty"] = False
    # MM sessions keep the merged material list alongside the raw loaded
    # files so re-validation can run against edits without re-reading disk.
    # state["mm_bundle"] is None for non-MM sessions.
    state["mm_bundle"] = None
    # PP sessions: holds the loaded BOM + optional Routing, plus their
    # merged (MATNR-grouped / PLNNR-grouped) views. None for non-PP.
    state["pp_bundle"] = None


def _get_session_state(user_token: str, create: bool = False) -> dict:
    _sweep_stale_sessions()
    state = WORKING.get(user_token)
    if state is None and create:
        state = {
            "workbook": None,
            "file_id": None,
            "filename": "",
            "module": None,
            "errors": [],
            "decisions": [],
            "audit_log": [],
            "resolved_decisions": set(),
            "accepted_errors": set(),
            "chat_history":[], # initializing chat history 
            "mm_bundle": None,
            "pp_bundle": None,
            "last_accessed": _time.time(),
        }
        WORKING[user_token] = state
    if state is not None:
        state["last_accessed"] = _time.time()
    return state


def _revalidate_and_rebuild(state: dict):
    """After any mutation, rebuild errors + decisions from current workbook.
    Respects the accepted_errors set — errors explicitly accepted-as-is
    by the user are filtered out, so they don't show as pending any more."""
    if state.get("mm_bundle"):
        # MM re-validation uses the loaded main/alt/longtext + current merged state
        bundle = state["mm_bundle"]
        catalogs = _get_mm_catalogs()
        # v57: pass cross-file context so orphan/sales-area errors stay
        # consistent across edits. (Edits to main material values can
        # also fix or break cross-file consistency — VKORG change can
        # break long-text sales area, etc.)
        alt_loaded = bundle.get("alt_loaded")
        lt_loaded = bundle.get("lt_loaded")
        errors = validate_mm(
            bundle["merged"].materials,
            catalogs,
            bundle["main_loaded"].sap_fields,
            merged_result=bundle["merged"],
            alt_uom_rows=(alt_loaded.rows if alt_loaded else None),
            longtext_rows=(lt_loaded.rows if lt_loaded else None),
            friendly_labels=getattr(bundle["main_loaded"], "header_labels", None),
        )
    else:
        errors = validate(state["workbook"])
    accepted = state.get("accepted_errors") or set()
    if accepted:
        errors = [
            e for e in errors
            if (e.rule_id, e.sheet, e.col_idx, e.row_idx) not in accepted
        ]
    decisions = group_errors(errors)
    state["errors"] = errors
    state["decisions"] = decisions
    return errors


def _session_loaded(state: dict | None) -> bool:
    """True if the session has any module's data loaded.

    Used by every per-session endpoint to gate access without caring
    which module is active. Saves duplicating the `not state or not
    state["workbook"]` check and adding MM/PP alongside it everywhere.
    """
    if not state:
        return False
    return (
        state.get("workbook") is not None
        or state.get("mm_bundle") is not None
        or state.get("pp_bundle") is not None
    )


def _ensure_session_loaded(session_token: str | None, user: dict) -> dict | None:
    """Resolve session state. Currently just a thin wrapper around
    _get_session_state — kept as a separate function so that future
    work (lazy rebuild from disk after a backend restart) can be added
    here in one place rather than duplicated across every per-session
    endpoint.

    Used by the PP routes for parity with the MM/SD pattern that
    expects this helper to exist."""
    return _get_session_state(session_token)


def get_token_from_cookie(session: str | None = Cookie(None)) -> str | None:
    return session


def current_user(session: str | None = Cookie(None)) -> dict:
    user = auth_svc.get_user(session)
    if not user:
        raise HTTPException(401, "Not authenticated")
    return user


def current_user_optional(session: str | None = Cookie(None)) -> dict | None:
    return auth_svc.get_user(session)


# ──────────────────────────────────────────────────────────────────────────
# Auth routes
# ──────────────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/api/auth/login")
def login(req: LoginRequest, response: Response):
    user = auth_svc.login(req.username, req.password)
    if not user:
        raise HTTPException(401, "Invalid credentials")
    response.set_cookie("session", user["token"], httponly=True, samesite="lax", max_age=8*3600)
    return {
        "username": user["username"],
        "display_name": user["display_name"],
        "role": user["role"],
        "module": user.get("module"),
    }


@app.post("/api/auth/logout")
def logout(response: Response, session: str | None = Cookie(None)):
    if session:
        auth_svc.logout(session)
    WORKING.pop(session, None) if session else None
    response.delete_cookie("session")
    return {"ok": True}


@app.get("/api/auth/me")
def me(user: dict = Depends(current_user)):
    return {
        "username": user["username"],
        "display_name": user["display_name"],
        "role": user["role"],
        "module": user.get("module"),
    }


# ──────────────────────────────────────────────────────────────────────────
# Admin — User Management
# ──────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/users")
def list_users(user: dict = Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    return {"users": auth_svc.list_users()}


class CreateUserRequest(BaseModel):
    username: str
    password: str
    display_name: str
    role: str
    module: str | None = None


@app.post("/api/admin/users")
def create_user(req: CreateUserRequest, user: dict = Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    try:
        created = auth_svc.create_user(req.username, req.password, req.display_name, req.role, req.module)
        return {"ok": True, "username": created["username"]}
    except ValueError as e:
        raise HTTPException(400, str(e))


class ChangePasswordRequest(BaseModel):
    username: str
    new_password: str


@app.post("/api/admin/users/password")
def change_pw(req: ChangePasswordRequest, user: dict = Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    try:
        auth_svc.change_password(req.username, req.new_password)
        return {"ok": True}
    except ValueError as e:
        raise HTTPException(400, str(e))


class ChangeRoleRequest(BaseModel):
    username: str
    new_role: str
    new_module: str | None = None


@app.post("/api/admin/users/role")
def change_role(req: ChangeRoleRequest, user: dict = Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    try:
        auth_svc.change_role(req.username, req.new_role, req.new_module)
        audit_log.log(user=user, action="user_role_changed",
                      details={"target": req.username, "new_role": req.new_role, "new_module": req.new_module})
        return {"ok": True}
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.delete("/api/admin/users/{username}")
def delete_user(username: str, user: dict = Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    try:
        auth_svc.delete_user(username)
        audit_log.log(user=user, action="user_deleted",
                      details={"deleted_username": username})
        return {"ok": True}
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.get("/api/admin/audit")
def admin_audit(limit: int = 500, user: dict = Depends(current_user)):
    """Full audit trail — admin or IT."""
    if user["role"] not in ("admin", "it"):
        raise HTTPException(403, "Admin or IT only")
    return {"entries": audit_log.list_all(limit=limit)}


@app.get("/api/admin/audit/file/{file_id}")
def admin_audit_for_file(file_id: str, user: dict = Depends(current_user)):
    if user["role"] not in ("admin", "it"):
        raise HTTPException(403, "Admin or IT only")
    return {"entries": audit_log.list_for_file(file_id)}


# ──────────────────────────────────────────────────────────────────────────
# Repository
# ──────────────────────────────────────────────────────────────────────────

@app.get("/api/repo/files")
def repo_list(module: str | None = None, status: str | None = None, user: dict = Depends(current_user)):
    files = repo.list_files(user, module=module, status=status)
    return {"files": files, "total": len(files)}


@app.get("/api/repo/modules")
def get_modules(user: dict = Depends(current_user)):
    """Return list of modules visible to this user + their status."""
    modules = []
    for mod in auth_svc.ALL_MODULES:
        if not auth_svc.can_access_module(user, mod):
            continue
        files = repo.list_files(user, module=mod)
        modules.append({
            "code": mod,
            "name": mod,
            "accessible": True,
            "file_count": len(files),
            "in_progress": sum(1 for f in files if f["status"] == "in_progress"),
            "validated": sum(1 for f in files if f["status"] == "validated"),
            "ltmc_uploaded": sum(1 for f in files if f["status"] == "ltmc_uploaded"),
        })
    # Also list modules they can't access (greyed-out in UI)
    for mod in auth_svc.ALL_MODULES:
        if not auth_svc.can_access_module(user, mod):
            modules.append({
                "code": mod,
                "name": mod,
                "accessible": False,
                "file_count": 0,
                "in_progress": 0,
                "validated": 0,
                "ltmc_uploaded": 0,
            })
    # De-duplicate (keep accessible entries first)
    seen = set()
    out = []
    for m in modules:
        if m["code"] not in seen:
            out.append(m)
            seen.add(m["code"])
    return {"modules": out}


@app.post("/api/repo/upload")
async def repo_upload(module: str = Form(...), file: UploadFile = File(...),
                       user: dict = Depends(current_user)):
    """Upload a file to the repository for a specific module."""
    if not auth_svc.can_access_module(user, module):
        raise HTTPException(403, f"You don't have access to module {module}")
    if module == "SD":
        pass  # SD uses single-file upload (one XML per customer dataset)
    elif module == "MM":
        # MM requires 3 files and has its own upload endpoint at /api/mm/upload.
        # Bounce the user there rather than accepting a partial upload here.
        raise HTTPException(
            400,
            "MM uploads must use /api/mm/upload (requires 3 files: main + alt_uom + longtext)",
        )
    elif module not in auth_svc.ALL_MODULES:
        raise HTTPException(400, "Invalid module")

    contents = await file.read()
    if not file.filename.lower().endswith((".xml", ".xlsx")):
        raise HTTPException(400, "Only .xml or .xlsx files are supported")

    entry = repo.save_file(module, file.filename, contents, user)

    # Auto-parse + validate for quick stats
    try:
        wb = parse_xml(contents)
        errors = validate(wb)
        decisions = group_errors(errors)
        repo.update_stats(entry["file_id"], wb.customer_count, len(errors), len(decisions))
        entry = repo.get_file(entry["file_id"])
    except Exception as e:
        entry["parse_error"] = str(e)

    audit_log.log(
        user=user, action="file_upload", file_id=entry["file_id"],
        filename=entry["filename"], module=module,
        affected_count=entry.get("row_count") or 0,
        details={"size_bytes": entry["size_bytes"]},
    )

    return entry


@app.post("/api/repo/files/{file_id}/validated")
def mark_validated(file_id: str, user: dict = Depends(current_user)):
    entry = repo.get_file(file_id)
    if not entry:
        raise HTTPException(404, "Not found")
    if not auth_svc.can_validate(user, entry["module"]):
        raise HTTPException(403, "You cannot validate this file")
    try:
        updated = repo.mark_validated(file_id, user)
        audit_log.log(user=user, action="mark_validated", file_id=file_id,
                      filename=entry["filename"], module=entry["module"])
        return updated
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.post("/api/repo/files/{file_id}/revoke")
def revoke_validation(file_id: str, user: dict = Depends(current_user)):
    try:
        entry = repo.get_file(file_id)
        updated = repo.revoke_validation(file_id, user)
        audit_log.log(user=user, action="revoke_validation", file_id=file_id,
                      filename=entry["filename"] if entry else None,
                      module=entry["module"] if entry else None)
        return updated
    except ValueError as e:
        raise HTTPException(400, str(e))
    except PermissionError as e:
        raise HTTPException(403, str(e))


@app.post("/api/repo/files/{file_id}/ltmc_uploaded")
def mark_ltmc(file_id: str, user: dict = Depends(current_user)):
    if not auth_svc.can_mark_ltmc_uploaded(user):
        raise HTTPException(403, "Only IT or admin can mark LTMC uploaded")
    try:
        entry = repo.get_file(file_id)
        updated = repo.mark_ltmc_uploaded(file_id, user)
        audit_log.log(user=user, action="mark_ltmc_uploaded", file_id=file_id,
                      filename=entry["filename"] if entry else None,
                      module=entry["module"] if entry else None)
        return updated
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.delete("/api/repo/files/{file_id}")
def delete_file(file_id: str, user: dict = Depends(current_user)):
    try:
        entry = repo.get_file(file_id)
        repo.delete_file(file_id, user)
        if entry:
            audit_log.log(user=user, action="file_deleted", file_id=file_id,
                          filename=entry["filename"], module=entry["module"])
        return {"ok": True}
    except PermissionError as e:
        raise HTTPException(403, str(e))


@app.get("/api/repo/files/{file_id}/download")
def download_file(file_id: str, user: dict = Depends(current_user)):
    entry = repo.get_file(file_id)
    if not entry:
        raise HTTPException(404, "Not found")
    if not auth_svc.can_access_module(user, entry["module"]):
        raise HTTPException(403, "Access denied")
    path = repo.get_file_path(file_id)
    if not path or not path.exists():
        raise HTTPException(404, "File missing on disk")
    return FileResponse(path=path, media_type="application/xml", filename=entry["filename"])


# ──────────────────────────────────────────────────────────────────────────
# Working session (open file for editing)
# ──────────────────────────────────────────────────────────────────────────

@app.post("/api/session/open/{file_id}")
def open_file(file_id: str, session: str | None = Cookie(None), user: dict = Depends(current_user)):
    entry = repo.get_file(file_id)
    if not entry:
        raise HTTPException(404, "Not found")
    if not auth_svc.can_access_module(user, entry["module"]):
        raise HTTPException(403, "Access denied")

    # Dispatch by module — SD reads the single XML via xml_engine; MM reads
    # the 3-file bundle via mm_loader + merges. Downstream endpoints all work
    # off state["errors"] so they don't need to know which path was taken.
    state = _get_session_state(session, create=True)

    if entry["module"] == "MM":
        paths = repo.get_mm_bundle_paths(file_id)
        if not paths or not paths["main"].exists():
            raise HTTPException(404, "MM bundle missing on disk")
        # Read + merge.
        # The Main file is always present and required. Alt UoM and Long
        # Text are optional — when the user uploaded just the Main file,
        # the bundle directory contains 0-byte placeholders for those
        # slots. We detect that and pass None to the merger rather than
        # calling the loader on an empty file (which would throw a
        # parse error). The merger's signature already accepts None for
        # alt_uom_file and longtext_file. See mm_routes.mm_upload for
        # the symmetric upload-side logic.
        # v62: LTMC source data form (XML) upload mode. When the main
        # file on disk is an XML rather than xlsx, route through the
        # LTMC form loader. The loader extracts main + alt-UoM + long-
        # text from the form's own sheets — no separate xlsx files
        # are needed in those slots (though if they were uploaded
        # explicitly, they win).
        from services.mm_ltmc_form_loader import (
            looks_like_ltmc_form, load_ltmc_form,
        )
        main_path = paths["main"]
        # Read just the first 8 KB for the detector (the file may be
        # large; full read happens only if we need to parse).
        head = main_path.read_bytes()[:8192] if main_path.exists() else b""
        if looks_like_ltmc_form(entry["filename"], head):
            full_bytes = main_path.read_bytes()
            main_loaded, alt_from_form, lt_from_form = load_ltmc_form(
                full_bytes, filename=entry["filename"],
            )
            # Explicit alt/lt slots win over form-embedded sheets
            alt_path = paths["alt_uom"]
            if alt_path.exists() and alt_path.stat().st_size > 0:
                alt_loaded = MM_LOADERS["alt_uom"](alt_path, "alt_uom.xlsx")
            else:
                alt_loaded = alt_from_form
            lt_path = paths["longtext"]
            if lt_path.exists() and lt_path.stat().st_size > 0:
                lt_loaded = MM_LOADERS["longtext"](lt_path, "longtext.xlsx")
            else:
                lt_loaded = lt_from_form
        else:
            main_loaded = MM_LOADERS["main"](paths["main"], entry["filename"])
            alt_path = paths["alt_uom"]
            alt_loaded = (MM_LOADERS["alt_uom"](alt_path, "alt_uom.xlsx")
                          if alt_path.exists() and alt_path.stat().st_size > 0
                          else None)
            lt_path = paths["longtext"]
            lt_loaded = (MM_LOADERS["longtext"](lt_path, "longtext.xlsx")
                         if lt_path.exists() and lt_path.stat().st_size > 0
                         else None)
        merged = mm_merge(main_loaded, alt_loaded, lt_loaded)
        # MM catalogs come from the bundled MM_KDS.xlsx (loaded once at
        # process start, cached). Per-customer KDS override was tried in
        # v51 and removed in v54 — single source of truth for catalogs.
        catalogs = _get_mm_catalogs()
        # v57: pass MergeResult + alt/lt source rows so cross-file checks
        # run (orphan MATNRs in alt/lt files, sales-area mismatches in
        # long-text). Backward compat: when alt_loaded/lt_loaded are
        # None this path collapses to the old behavior.
        errors = validate_mm(
            merged.materials, catalogs, main_loaded.sap_fields,
            merged_result=merged,
            alt_uom_rows=(alt_loaded.rows if alt_loaded else None),
            longtext_rows=(lt_loaded.rows if lt_loaded else None),
            friendly_labels=getattr(main_loaded, "header_labels", None),
        )
        decisions = group_errors(errors)

        # Stash everything the in-session flow needs.
        state["workbook"] = None            # MM has no "workbook" — bundle replaces it
        state["mm_bundle"] = {
            "main_loaded": main_loaded,
            "alt_loaded": alt_loaded,
            "lt_loaded": lt_loaded,
            "merged": merged,
        }
        state["file_id"] = file_id
        state["filename"] = entry["filename"]
        state["module"] = "MM"
        state["errors"] = errors
        state["decisions"] = decisions
        state["audit_log"] = []
        state["resolved_decisions"] = set()
        state["accepted_errors"] = set()
        state["dirty"] = False

        repo.update_stats(file_id, len(merged.materials), len(errors), len(decisions))

        return {
            "file_id": file_id,
            "filename": entry["filename"],
            "module": "MM",
            "object_name": "Material Master",
            "customer_count": len(merged.materials),  # reused key name; UI shows "materials"
            "sheet_count": 1,
            "sheets": [{"name": "Materials", "rows": len(merged.materials),
                        "columns": len(main_loaded.sap_fields)}],
            "error_count": len(errors),
            "decision_count": len(decisions),
            "mm_summary": merged.summary,   # extra hint for MM dashboard
        }

    # ── PP path ──
    if entry["module"] == "PP":
        # PP bundles store the BOM file as the "main" slot (re-uses
        # the MM bundle directory layout) and the optional Routing
        # file as the "alt_uom" slot. The "longtext" slot is unused.
        # See pp_routes.pp_upload_bom for the symmetric upload-side
        # logic.
        from services.pp_loader import load_bom
        from services.routing_loader import load_routing
        from services.pp_kds import load_pp_catalogs as _load_pp_cats
        from services.pp_rulebook import get_rulebook as _pp_rulebook
        from services.pp_validator import validate_bom, validate_routing
        from services.pp_merger import merge_bom
        from services.routing_merger import merge_routing

        paths = repo.get_mm_bundle_paths(file_id)
        if not paths or not paths["main"].exists():
            raise HTTPException(404, "PP bundle missing on disk")

        bom_path = paths["main"]
        routing_path = paths["alt_uom"]

        bom = load_bom(bom_path, entry["filename"])
        merged_bom = merge_bom(bom)

        routing = None
        merged_routing = None
        if routing_path.exists() and routing_path.stat().st_size > 0:
            routing = load_routing(routing_path, "routing.xlsx")
            merged_routing = merge_routing(routing)

        catalogs = _load_pp_cats(None)
        rulebook = _pp_rulebook()

        errors = validate_bom(bom, catalogs, rulebook)
        if routing is not None:
            errors.extend(validate_routing(routing, catalogs, rulebook))

        # Decision grouping for PP — TODO: extend group_errors to be
        # PP-aware. For now we just leave decisions empty so the
        # Decisions tab is visible but reports "0 decisions" — the
        # Error Grid still surfaces every error individually.
        decisions: list = []

        state["workbook"] = None
        state["mm_bundle"] = None
        state["pp_bundle"] = {
            "bom": bom,
            "routing": routing,
            "merged_bom": merged_bom,
            "merged_routing": merged_routing,
            "catalogs": catalogs,
            "rulebook": rulebook,
        }
        state["file_id"] = file_id
        state["filename"] = entry["filename"]
        state["module"] = "PP"
        state["errors"] = errors
        state["decisions"] = decisions
        state["audit_log"] = []
        state["resolved_decisions"] = set()
        state["accepted_errors"] = set()
        state["dirty"] = False

        material_count = len(merged_bom.materials)
        repo.update_stats(file_id, material_count, len(errors), len(decisions))

        sheets_listing = []
        for n, sh in bom.sheets.items():
            sheets_listing.append({"name": n, "rows": len(sh.rows),
                                   "columns": len(sh.sap_fields)})
        if routing is not None:
            for n, sh in routing.sheets.items():
                # Prefix routing sheet names so they don't collide with
                # BOM sheet names (Global Dependency, Local Dependency,
                # etc. exist in both file types).
                sheets_listing.append({"name": f"Routing · {n}",
                                       "rows": len(sh.rows),
                                       "columns": len(sh.sap_fields)})

        return {
            "file_id": file_id,
            "filename": entry["filename"],
            "module": "PP",
            "object_name": "BOM" + (" + Routing" if routing else ""),
            "customer_count": material_count,
            "sheet_count": len(sheets_listing),
            "sheets": sheets_listing,
            "error_count": len(errors),
            "decision_count": len(decisions),
            "pp_summary": {
                "bom": merged_bom.summary,
                "routing": (merged_routing.summary if merged_routing else None),
            },
        }

    # ── SD path (unchanged) ──
    path = repo.get_file_path(file_id)
    if not path or not path.exists():
        raise HTTPException(404, "File missing on disk")
    contents = path.read_bytes()

    wb = parse_xml(contents)
    errors = validate(wb)
    decisions = group_errors(errors)

    state["workbook"] = wb
    state["mm_bundle"] = None
    state["file_id"] = file_id
    state["filename"] = entry["filename"]
    state["module"] = entry["module"]
    state["errors"] = errors
    state["decisions"] = decisions
    state["audit_log"] = []
    state["resolved_decisions"] = set()
    state["accepted_errors"] = set()
    state["dirty"] = False

    repo.update_stats(file_id, wb.customer_count, len(errors), len(decisions))

    return {
        "file_id": file_id,
        "filename": entry["filename"],
        "module": entry["module"],
        "object_name": wb.object_name,
        "customer_count": wb.customer_count,
        "sheet_count": len(wb.sheets),
        "sheets": [
            {"name": n, "rows": len(s.data_rows), "columns": len(s.specs)}
            for n, s in wb.sheets.items() if len(s.data_rows) > 0
        ],
        "error_count": len(errors),
        "decision_count": len(decisions),
    }


@app.get("/api/session/dashboard")
def dashboard(session: str | None = Cookie(None), user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        return {"loaded": False}

    active_decisions = [d for d in state["decisions"] if d.decision_id not in state["resolved_decisions"]]
    summary = summarize(active_decisions)

    resolved_errors = sum(
        d.affected_count for d in state["decisions"] if d.decision_id in state["resolved_decisions"]
    )
    total_errors_ever = sum(d.affected_count for d in state["decisions"])

    # Build module-specific header fields. SD workbook has wb.object_name /
    # wb.customer_count / wb.sheets; MM has merged.materials; PP has the
    # bom + optional routing in pp_bundle.
    mm_warnings: list[str] = []
    mm_stats: dict | None = None
    pp_stats: dict | None = None
    if state.get("pp_bundle"):
        # PP path — BOM + optional Routing.
        bundle = state["pp_bundle"]
        merged_bom = bundle["merged_bom"]
        merged_routing = bundle.get("merged_routing")
        record_count = merged_bom.summary["material_count"]
        object_name = "BOM" + (" + Routing" if merged_routing else "")

        # One sheet entry per BOM and (if present) Routing tab. Errors
        # are attributed to whichever sheet they came from.
        sheets_info = []
        for n, sh in bundle["bom"].sheets.items():
            if len(sh.rows) == 0:
                continue
            sheets_info.append({
                "name": n,
                "rows": len(sh.rows),
                "errors": sum(1 for e in state["errors"] if e.sheet == n),
            })
        if bundle.get("routing"):
            for n, sh in bundle["routing"].sheets.items():
                if len(sh.rows) == 0:
                    continue
                sheets_info.append({
                    "name": f"Routing · {n}",
                    "rows": len(sh.rows),
                    "errors": sum(1 for e in state["errors"] if e.sheet == n),
                })
        sheet_count = len(sheets_info)

        pp_stats = {
            "material_count": merged_bom.summary["material_count"],
            "bom_total_rows": merged_bom.summary["total_rows_all_sheets"],
            "routing_count": (merged_routing.summary["routing_count"]
                              if merged_routing else 0),
            "routing_total_rows": (merged_routing.summary["total_rows_all_sheets"]
                                   if merged_routing else 0),
            "has_routing": bool(merged_routing),
        }
    elif state.get("mm_bundle"):
        merged = state["mm_bundle"]["merged"]
        object_name = "Material Master"
        record_count = len(merged.materials)
        sheets_info = [{
            "name": "Materials",
            "rows": record_count,
            "errors": len(state["errors"]),
        }]
        sheet_count = 1

        # Multi-plant + orphan stats for the UI header. Phase-2 Peenya
        # test exposed that these are real migration issues SMEs need to
        # see before they start fixing individual records.
        mm_stats = {
            "plant_row_count": merged.summary.get("plant_row_count", record_count),
            "multi_plant_count": merged.summary.get("multi_plant_count", 0),
            "duplicate_count": len(merged.duplicate_matnrs_in_main),
            "orphan_alt_uom_count": len(merged.orphans_in_alt_uom),
            "orphan_longtext_count": len(merged.orphans_in_longtext),
            "distinct_plants": merged.summary.get("distinct_plants", []),
            "distinct_material_types": merged.summary.get("distinct_material_types", []),
        }
        if merged.file_pair_warning:
            mm_warnings.append(merged.file_pair_warning)
        if mm_stats["multi_plant_count"] > 0:
            mp = mm_stats["multi_plant_count"]
            plants = mm_stats["distinct_plants"]
            mm_warnings.append(
                f"{mp} of {record_count} materials are assigned to multiple plants "
                f"({'/'.join(plants)}). LTMC export will need to fan out each "
                f"material into one Plant Data row per assignment."
            )
        if mm_stats["duplicate_count"] > 0:
            mm_warnings.append(
                f"{mm_stats['duplicate_count']} true duplicates in the main file "
                f"(same MATNR + same WERKS appearing more than once). First "
                f"occurrence kept, rest discarded."
            )
    else:
        wb = state["workbook"]
        object_name = wb.object_name
        record_count = wb.customer_count
        sheets_info = [
            {
                "name": n,
                "rows": len(s.data_rows),
                "errors": sum(1 for e in state["errors"] if e.sheet == n),
            }
            for n, s in wb.sheets.items() if len(s.data_rows) > 0
        ]
        sheet_count = sum(1 for s in wb.sheets.values() if len(s.data_rows) > 0)

    return {
        "loaded": True,
        "file_id": state["file_id"],
        "filename": state["filename"],
        "module": state["module"],
        "object_name": object_name,
        "customer_count": record_count,   # key kept for UI backward compat (it labels "materials" for MM, "BOMs" for PP)
        "sheet_count": sheet_count,
        "total_errors_found": total_errors_ever,
        "resolved_errors": resolved_errors,
        "pending_errors": summary["total_errors"],
        "pending_decisions": summary["total_decisions"],
        "pattern_decisions": summary["pattern_decisions"],
        "individual_decisions": summary["individual_decisions"],
        "completion_pct": round(resolved_errors / total_errors_ever * 100, 1) if total_errors_ever else 100.0,
        "sheets": sheets_info,
        "audit_count": len(state["audit_log"]),
        # MM-specific fields (None for SD/PP)
        "mm_stats": mm_stats,
        "mm_warnings": mm_warnings,
        # PP-specific fields (None for SD/MM)
        "pp_stats": pp_stats,
    }


@app.get("/api/session/decisions")
def get_decisions(status: str = "pending", session: str | None = Cookie(None), user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    if status == "pending":
        decs = [d for d in state["decisions"] if d.decision_id not in state["resolved_decisions"]]
    elif status == "resolved":
        decs = [d for d in state["decisions"] if d.decision_id in state["resolved_decisions"]]
    else:
        decs = state["decisions"]

    # v62: include current LTMC overrides so the UI can show what's
    # already set. Empty dict for non-MM sessions and for MM sessions
    # where no overrides have been entered yet.
    ltmc_overrides = dict(state.get("ltmc_overrides") or {})

    return {
        "decisions": [d.as_dict() for d in decs],
        "total": len(decs),
        "ltmc_overrides": ltmc_overrides,
    }


@app.get("/api/session/errors_by_rule/{decision_id}")
def errors_by_rule(decision_id: str, session: str | None = Cookie(None), user: dict = Depends(current_user)):
    """Preview rows affected by a decision.
    Capped at PREVIEW_MAX rows so the modal stays responsive on 5k+ duplicate sets.
    Returns total_count so the UI can show "Showing N of M"."""
    PREVIEW_MAX = 500

    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")
    decision = next((d for d in state["decisions"] if d.decision_id == decision_id), None)
    if not decision:
        raise HTTPException(404, "Decision not found")

    # MM branch: the "sheet" model doesn't apply — we preview per material
    # using the merged bundle. Simpler than SD; no sales-area cross-lookup,
    # no duplicate clustering yet. Shows the material's MATNR, MAKTX,
    # plant + the value for the column under scrutiny.
    if state.get("mm_bundle"):
        bundle = state["mm_bundle"]
        materials = bundle["merged"].materials
        main_fields = bundle["main_loaded"].sap_fields

        # v66: Cross-file decisions (sheet=LongText / AlternateUnits) have
        # row_idx values that point into the SOURCE file's rows array
        # (alt_loaded.rows or lt_loaded.rows), NOT into `materials`. The
        # main-file preview logic below would silently dereference the
        # wrong material if we let cross-file row indices fall through.
        # So we short-circuit here: build the preview from the source
        # rows, surfacing MATNR + the offending field's value so the
        # cross-file guidance modal can show the SME exactly which rows
        # in their alt-UoM or long-text xlsx need editing.
        if decision.sheet in ("LongText", "AlternateUnits"):
            source_rows = []
            if decision.sheet == "AlternateUnits":
                alt_loaded = bundle.get("alt_loaded")
                if alt_loaded is not None:
                    source_rows = list(alt_loaded.rows)
            else:  # LongText
                lt_loaded = bundle.get("lt_loaded")
                if lt_loaded is not None:
                    source_rows = list(lt_loaded.rows)

            all_row_ids = list(decision.error_row_indexes)
            total_count = len(all_row_ids)
            unique_row_ids = list(dict.fromkeys(all_row_ids))[:PREVIEW_MAX]

            rows = []
            for r_idx in unique_row_ids:
                if not (0 <= r_idx < len(source_rows)):
                    continue
                src = source_rows[r_idx]
                # Pull MATNR + the field under scrutiny + value as flat keys
                # the modal renders directly. Keeping the shape lean since
                # cross-file guidance doesn't need full record data.
                matnr = src.values.get("MATNR")
                matnr_str = ("" if matnr is None
                             else (str(int(matnr)) if isinstance(matnr, float) and matnr.is_integer()
                                   else str(matnr)))
                val = src.values.get(decision.sap_field) if decision.sap_field else None
                rows.append({
                    "row_idx": r_idx,
                    "xml_row": getattr(src, "excel_row", r_idx + 3),
                    "matnr": matnr_str,
                    "sap_field": decision.sap_field or "",
                    "column_label": decision.column_label or decision.sap_field or "",
                    "value": "" if val is None else str(val),
                })

            return {
                "decision_id": decision_id,
                "rule_id": decision.rule_id,
                "rule_name": decision.rule_name,
                "column_label": decision.column_label,
                "sheet": decision.sheet,
                "key_columns": [
                    {"col_idx": 1, "label": "MATNR"},
                    {"col_idx": 2, "label": decision.sap_field or "Field"},
                ],
                "rows": rows,
                "total_count": total_count,
                "truncated": total_count > len(rows),
                "is_cross_file": True,  # signals the frontend to use guidance modal
            }

        key_cols = [
            {"col_idx": 1,  "label": "MATNR"},
            {"col_idx": 2,  "label": "MTART"},
            {"col_idx": 3,  "label": "WERKS"},
            {"col_idx": 4,  "label": "MAKTX"},
        ]
        # Always add the field under scrutiny so SMEs see the bad value
        if decision.sap_field and decision.sap_field not in [k["label"] for k in key_cols]:
            # col_idx here is synthetic index used by UI for ordering only
            key_cols.append({"col_idx": 99, "label": decision.sap_field})

        all_row_ids = list(decision.error_row_indexes)
        total_count = len(all_row_ids)
        # Dedupe row_idx for plant-scoped rules (each appears once per plant
        # in error_row_indexes). We re-expand below per plant.
        unique_row_ids = list(dict.fromkeys(all_row_ids))

        # Plant-scoped fields — for these, render one row per plant_row of
        # the material so the SME sees each plant's value separately. The
        # field under scrutiny is read from the plant row, not main.
        PLANT_SCOPED = {"WERKS", "LGORT", "LGPRO", "DISPO", "FEVOR", "BWKEY"}
        is_plant_scoped = decision.sap_field in PLANT_SCOPED

        # Pre-build error index once: row_idx → set of bad plants for THIS
        # rule. Without this we'd O(errors × materials) on every preview —
        # 100k+ errors × 10k materials = a billion compares = 2s on the
        # Peenya FG file. With the index, it's a dict lookup per row.
        bad_plants_by_row: dict[int, set[str]] = {}
        if is_plant_scoped:
            # Reverse-engineer the WERKS that emitted each error by looking
            # at which plant_row(s) carry the error's bad value. Multiple
            # plant_rows may share the same value (e.g. all 11 plants have
            # LGPRO='FEU1', or all have BWKEY blank) — collect ALL matching
            # plants, not just the first one. Without this, the preview
            # for multi-plant materials only ever shows PE01 (the first
            # plant_row, alphabetically) even though SP01-WH01 all have
            # the same problem.
            for e in state["errors"]:
                if e.rule_id != decision.rule_id:
                    continue
                if e.sap_field != decision.sap_field:
                    continue
                if not (0 <= e.row_idx < len(materials)):
                    continue
                m = materials[e.row_idx]
                bucket = bad_plants_by_row.setdefault(e.row_idx, set())
                # If we've already found all plants for this row from a
                # previous error pass, skip the inner loop.
                already_have_all = len(bucket) == len(m.plant_rows)
                if already_have_all:
                    continue
                for pr in m.plant_rows:
                    pr_val = str(pr.get(decision.sap_field, "") or "")
                    if pr_val == e.value:
                        werks = str(pr.get("WERKS", "") or "")
                        if werks:
                            bucket.add(werks)
                        # NOTE: no break — multiple plant_rows can share
                        # the same value (especially for blank fields).

        # Build the (row_idx, plant_row) display tuples. Sample evenly
        # across plants so a multi-plant material doesn't get its first
        # plant overrepresented in the truncated preview. Strategy:
        # group eligible (row_idx, plant_row) tuples by WERKS, then take
        # equal shares from each plant up to PREVIEW_MAX.
        display_tuples: list[tuple[int, object | None]] = []
        if is_plant_scoped:
            # Build (row_idx, plant_row) tuples grouped by plant
            by_plant: dict[str, list[tuple[int, object]]] = {}
            for row_idx in unique_row_ids:
                if not (0 <= row_idx < len(materials)):
                    continue
                m = materials[row_idx]
                bad = bad_plants_by_row.get(row_idx)
                for pr in m.plant_rows:
                    werks = str(pr.get("WERKS", "") or "")
                    if not werks:
                        continue
                    if bad and werks not in bad:
                        continue
                    by_plant.setdefault(werks, []).append((row_idx, pr))
            # Sort plants for deterministic output
            sorted_plants = sorted(by_plant.keys())
            n_plants = len(sorted_plants)
            if n_plants > 0:
                # Equal share per plant (rounded up so we don't lose budget)
                per_plant_quota = max(1, PREVIEW_MAX // n_plants)
                # Round-robin fill: take quota from each plant in turn
                cursor: dict[str, int] = {p: 0 for p in sorted_plants}
                while len(display_tuples) < PREVIEW_MAX:
                    progress = False
                    for plant in sorted_plants:
                        if cursor[plant] < min(len(by_plant[plant]), per_plant_quota):
                            display_tuples.append(by_plant[plant][cursor[plant]])
                            cursor[plant] += 1
                            progress = True
                            if len(display_tuples) >= PREVIEW_MAX:
                                break
                    if not progress:
                        # All plants exhausted within their quota — top up
                        # from any plant still has rows.
                        any_left = False
                        for plant in sorted_plants:
                            if cursor[plant] < len(by_plant[plant]):
                                display_tuples.append(by_plant[plant][cursor[plant]])
                                cursor[plant] += 1
                                any_left = True
                                if len(display_tuples) >= PREVIEW_MAX:
                                    break
                        if not any_left:
                            break
        else:
            for row_idx in unique_row_ids:
                if not (0 <= row_idx < len(materials)):
                    continue
                display_tuples.append((row_idx, None))
                if len(display_tuples) >= PREVIEW_MAX:
                    break

        truncated = len(display_tuples) >= PREVIEW_MAX and total_count > PREVIEW_MAX

        rows = []
        for row_idx, plant_row in display_tuples:
            m = materials[row_idx]
            # When plant-scoped, read all fields from plant_row; main is the
            # fallback for any missing keys (basic-data fields like MAKTX
            # that aren't on plant_rows individually).
            source = plant_row if plant_row is not None else m.main
            row_cells = {
                "MATNR": m.matnr,
                "MTART": str(m.main.get("MTART", "") or ""),
                "WERKS": str(source.get("WERKS", "") or m.main.get("WERKS", "") or ""),
                "MAKTX": str(m.main.get("MAKTX", "") or ""),
            }
            if decision.sap_field and decision.sap_field not in row_cells:
                # For plant-scoped, this might be empty (BWKEY missing) —
                # which is fine: the column shows blank with a red hint
                # indicating the missing-mandatory error.
                val = source.get(decision.sap_field)
                if val is None and plant_row is not None:
                    val = m.main.get(decision.sap_field)
                row_cells[decision.sap_field] = "" if val is None else str(val)
            rows.append({
                "row_idx": row_idx,
                "xml_row": m.source_excel_row,
                "cells": row_cells,
            })

        return {
            "decision_id": decision_id,
            "rule_id": decision.rule_id,
            "rule_name": decision.rule_name,
            "column_label": decision.column_label,
            "sheet": decision.sheet,
            "key_columns": key_cols,
            "rows": rows,
            "total_count": total_count,
            "truncated": truncated,
            "show_sales_areas": False,
        }

    # ── SD path (unchanged) ──
    wb = state["workbook"]
    sheet = wb.sheets.get(decision.sheet)

    # Core identifying columns — always shown. SO/DC/DIV included so they render
    # whenever the sheet is Sales Data (which has them natively).
    core_key_labels = [
        "Customer Number", "Name", "Name 2", "City",
        "Sales Org", "Distribution Channel", "Division",
    ]
    extra_key_labels_for_duplicate = [
        "Street", "Postal Code", "Country", "PAN Number", "GSTIN",
        "Mobile Phone", "Email Address", "Telephone No 1",
    ]
    # Contextual columns — rule-specific fields shown alongside the core
    # set so SMEs can judge the right fix at a glance.
    # Example: for Inco Location errors, showing the Incoterms code itself
    # (e.g. "CIF", "FOB") next to the Location helps decide what's truly junk
    # description vs a legitimate port name.
    contextual_columns_by_rule: dict[str, list[str]] = {
        "inco_location_description": ["Incoterms", "Inco. Location1"],
        "dl_expired":                ["Drug Licence Number", "Drug Licence Expiry Date"],
        "length_exceeded":           [],  # column under scrutiny is the decision's own column
        "invalid_state_in":          ["State", "Country/Region", "Postal Code"],
        "invalid_pan":               ["Permanent Account Number", "GSTIN"],
        "gstin_length":              ["GSTIN", "State"],
        "junk_value":                [],
        "duplicate_record":          [],  # duplicates already get the extra_key_labels set
    }
    rule_context = contextual_columns_by_rule.get(decision.rule_id, [])

    is_duplicate = "duplicate" in decision.rule_id.lower()
    labels_to_include = core_key_labels + rule_context + (extra_key_labels_for_duplicate if is_duplicate else [])
    # Always show the column under scrutiny itself (so user sees what's being judged)
    if decision.column_label and decision.column_label not in labels_to_include:
        labels_to_include.append(decision.column_label)

    # Match spec labels either exactly OR as a variant-with-suffix
    # (e.g. "Drug Licence Number 2" for wanted "Drug Licence Number").
    # We explicitly do NOT match when the wanted word appears mid-label
    # (e.g. "City" would otherwise match "City Code", which is a separate
    # SAP field). The rule is: spec.label must START with wanted plus a
    # space/digit/paren, OR equal wanted exactly.
    import re as _re
    key_cols = []
    seen_col_idx = set()
    for wanted in labels_to_include:
        # Anchor at start of the spec label. Trailing character, if any,
        # must be whitespace, digit, or an opening bracket — which is how
        # SAP's variant naming convention works ("Foo 2", "Foo Bar" is NOT
        # a variant of "Foo").
        pattern = _re.compile(
            r"^" + _re.escape(wanted) + r"(\s*\d|\s*\(|\s*\[|$)",
            _re.IGNORECASE,
        )
        for spec in sheet.specs:
            if spec.col_idx in seen_col_idx:
                continue
            if spec.label == wanted or pattern.search(spec.label):
                key_cols.append({"col_idx": spec.col_idx, "label": spec.label})
                seen_col_idx.add(spec.col_idx)
                break

    name_col = next((k["col_idx"] for k in key_cols if k["label"] == "Name"), None)
    num_col  = next((k["col_idx"] for k in key_cols if k["label"] == "Customer Number"), None)

    # Sales area cross-lookup — only when decision sheet is NOT Sales Data itself
    # AND sheet has Customer Number to join on.
    sales_area_map: dict[str, list[dict]] = {}
    show_sales_areas_column = (
        is_duplicate
        and decision.sheet != "Sales Data"
        and num_col is not None
    )
    if show_sales_areas_column:
        sd = wb.sheets.get("Sales Data")
        if sd:
            sd_num = next((s.col_idx for s in sd.specs if s.label == "Customer Number"), None)
            sd_so  = next((s.col_idx for s in sd.specs if s.label == "Sales Org"), None)
            sd_dc  = next((s.col_idx for s in sd.specs if s.label == "Distribution Channel"), None)
            sd_div = next((s.col_idx for s in sd.specs if s.label == "Division"), None)
            if sd_num is not None and (sd_so or sd_dc or sd_div):
                for row in sd.data_rows:
                    cn = str(row.get(sd_num, "")).strip()
                    if not cn:
                        continue
                    triple = {
                        "sales_org": str(row.get(sd_so, "")) if sd_so else "",
                        "distribution_channel": str(row.get(sd_dc, "")) if sd_dc else "",
                        "division": str(row.get(sd_div, "")) if sd_div else "",
                    }
                    sales_area_map.setdefault(cn, []).append(triple)

    has_any_sa = any(sa for sa in sales_area_map.values()) if sales_area_map else False
    show_sales_areas_column = show_sales_areas_column and has_any_sa

    # Collect all row indexes to potentially show
    all_row_ids = list(decision.error_row_indexes)
    if is_duplicate and num_col is not None:
        flagged_nums = set()
        for row_idx in decision.error_row_indexes:
            if 0 <= row_idx < len(sheet.data_rows):
                cn = str(sheet.data_rows[row_idx].get(num_col, "")).strip()
                if cn:
                    flagged_nums.add(cn)
        first_seen = {}
        for idx, row in enumerate(sheet.data_rows):
            cn = str(row.get(num_col, "")).strip()
            if cn in flagged_nums and cn not in first_seen:
                first_seen[cn] = idx
        for first_idx in first_seen.values():
            if first_idx not in all_row_ids:
                all_row_ids.append(first_idx)

        def cluster_key(idx):
            row = sheet.data_rows[idx] if 0 <= idx < len(sheet.data_rows) else {}
            return (str(row.get(num_col, "")).strip(), idx)
        all_row_ids.sort(key=cluster_key)

    total_count = len(all_row_ids)
    row_ids_to_show = all_row_ids[:PREVIEW_MAX]
    truncated = total_count > PREVIEW_MAX

    def completeness(row) -> int:
        return sum(1 for v in row.values() if v and str(v).strip())

    rows = []
    for row_idx in row_ids_to_show:
        if 0 <= row_idx < len(sheet.data_rows):
            row = sheet.data_rows[row_idx]
            key_values = {}
            for k in key_cols:
                v = row.get(k["col_idx"], "")
                key_values[k["label"]] = str(v) if v is not None else ""
            cust_num = str(row.get(num_col, "")) if num_col is not None else ""
            rows.append({
                "row_idx": row_idx,
                "xml_row": row_idx + 9,
                "customer_num": cust_num,
                "customer_name": str(row.get(name_col, "")) if name_col is not None else "",
                "current_value": str(row.get(decision.col_idx, "")),
                "key_values": key_values,
                "completeness": completeness(row),
                "is_flagged_duplicate": row_idx in decision.error_row_indexes,
                "sales_areas": sales_area_map.get(cust_num.strip(), []) if show_sales_areas_column else [],
            })

    # For duplicates: tag the "best" (most complete) row in each customer
    # cluster. If tied on completeness, prefer the kept (non-flagged) one,
    # then lowest row_idx. This helps SMEs pick which to keep at a glance.
    if is_duplicate:
        by_customer: dict[str, list[dict]] = {}
        for r in rows:
            by_customer.setdefault(r["customer_num"], []).append(r)
        for cluster in by_customer.values():
            if len(cluster) < 2:
                continue
            cluster.sort(key=lambda r: (
                -r["completeness"],               # most complete first
                r["is_flagged_duplicate"],         # non-dup beats dup
                r["row_idx"],                      # earlier row wins
            ))
            cluster[0]["is_best"] = True
            for r in cluster[1:]:
                r["is_best"] = False
    # For non-duplicates, no "best" concept — leave is_best unset

    return {
        "decision": decision.as_dict(),
        "rows": rows,
        "is_duplicate": is_duplicate,
        "key_columns": [k["label"] for k in key_cols],
        "has_sales_areas": show_sales_areas_column,
        "column_label": decision.column_label,
        "rule_name": decision.rule_name,
        "rule_id": decision.rule_id,
        "severity": decision.severity,
        "total_count": total_count,
        "shown_count": len(rows),
        "truncated": truncated,
    }


class ApplyActionRequest(BaseModel):
    decision_id: str
    action_id: str
    value: str | None = None
    reason: str | None = None
    strategy: str | None = None  # for delete_duplicates
    row_indexes: list[int] | None = None  # for delete_rows


@app.post("/api/session/decisions/apply")
def apply_action(req: ApplyActionRequest, session: str | None = Cookie(None), user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    # PP doesn't support pattern-based decisions in this build. PP errors
    # are surfaced as raw rows in the Error Grid; SMEs resolve them by
    # fixing the source file. Adding decision grouping for PP is M2.
    if state.get("pp_bundle"):
        raise HTTPException(
            501,
            "PP decision-application is not yet implemented. "
            "Fix issues in the source BOM/Routing file and re-upload.",
        )

    decision = next((d for d in state["decisions"] if d.decision_id == req.decision_id), None)
    if not decision:
        raise HTTPException(404, "Decision not found")

    # Enforce business reason on all mutating actions
    mutating_actions = {"replace_with", "fill_with", "clear_all", "set_urp",
                        "truncate_all", "delete_duplicates", "delete_rows", "accept_all",
                        "set_ltmc_default"}
    if req.action_id in mutating_actions:
        if not req.reason or not req.reason.strip():
            raise HTTPException(400, "A business reason is required for this action")

    # ── v62: set_ltmc_default — the SME has entered a value for an
    # LTMC-mandatory field that's missing from the customer source.
    # Store it in the session overrides map and mark the decision
    # resolved. No per-row data mutation, no SD-applier path (which
    # would crash on MM), no re-validation needed (the rule won't
    # re-fire because the LTMC generator has the value now).
    if req.action_id == "set_ltmc_default":
        if not state.get("mm_bundle"):
            raise HTTPException(400,
                "set_ltmc_default is only available for MM sessions")
        from services.mm_ltmc_overrides import set_override
        set_override(state, decision.sap_field, req.value or "")

        # Suppress this rule's errors so they don't keep showing up.
        # The LTMC export will pick up the override; revalidation should
        # NOT re-emit these errors.
        for row_idx in decision.error_row_indexes:
            state.setdefault("accepted_errors", set()).add(
                (decision.rule_id, decision.sheet, decision.col_idx, row_idx)
            )

        state["resolved_decisions"].add(decision.decision_id)

        # Audit log entry — same shape as group_replace so changes_summary
        # can render it. _snapshots is the per-row "before" state (all
        # blank for missing-from-source fields); new_value is the SME
        # value. _ltmc_override flag distinguishes this from a real
        # per-row mutation.
        snapshots = {str(rx): "" for rx in decision.error_row_indexes}
        audit_log.log(
            user=user, action="decision_set_ltmc_default",
            file_id=state["file_id"], filename=state["filename"],
            module=state["module"], sheet=decision.sheet, rule_id=decision.rule_id,
            affected_count=decision.affected_count,
            reason=req.reason or "",
            details={
                "column": decision.column_label,
                "sap_field": decision.sap_field,
                "value": req.value or "",
                "_col_idx": decision.col_idx,
                "_snapshots": snapshots,
                "_ltmc_override": True,   # marker for changes_summary
            },
        )

        # Re-filter errors for the suppressed rows so the count drops
        new_errors = [e for e in state["errors"]
                       if (e.rule_id, e.sheet, e.col_idx, e.row_idx)
                       not in state["accepted_errors"]]
        state["errors"] = new_errors

        if state["file_id"]:
            try:
                repo.update_stats(
                    state["file_id"],
                    state.get("mm_bundle", {}).get("merged").materials.__len__()
                        if state.get("mm_bundle") and state["mm_bundle"].get("merged")
                        else 0,
                    len(new_errors),
                    len([d for d in state["decisions"]
                         if d.decision_id not in state["resolved_decisions"]]),
                )
            except Exception:
                pass

        return {
            "success": True,
            "audit_entry": {
                "action": "decision_set_ltmc_default",
                "decision_id": decision.decision_id,
                "value": req.value or "",
                "affected_count": decision.affected_count,
            },
            "new_error_count": len(new_errors),
            "ltmc_overrides": dict(state.get("ltmc_overrides") or {}),
        }


    payload = {
        "value": req.value or "",
        "reason": req.reason or "",
        "strategy": req.strategy or "keep_first",
        "row_indexes": req.row_indexes or [],
    }

    try:
        entry = apply_decision(state["workbook"], decision, req.action_id, payload,
                                user=user["display_name"])
        state["audit_log"].append(entry)

        # accept_all: suppress these specific errors so they stop reappearing on re-validation
        if req.action_id in ("accept_all", "ignore"):
            for row_idx in decision.error_row_indexes:
                state.setdefault("accepted_errors", set()).add(
                    (decision.rule_id, decision.sheet, decision.col_idx, row_idx)
                )

        # Rebuild from current workbook, but keep the just-applied decision in the list
        # so it shows up in the "Resolved" filter. Keep its resolved status.
        applied_id = decision.decision_id
        # Snapshot the decision metadata before rebuilding (in case the rule no longer fires)
        applied_snapshot = decision
        errors = _revalidate_and_rebuild(state)
        # Re-insert the resolved snapshot if it disappeared (it almost certainly did)
        if not any(d.decision_id == applied_id for d in state["decisions"]):
            state["decisions"].append(applied_snapshot)
        state["resolved_decisions"].add(applied_id)

        # Persist working copy back to repo
        if state["file_id"]:
            xml_bytes = write_xml(state["workbook"])
            repo.save_working_copy(state["file_id"], xml_bytes)
            repo.update_stats(state["file_id"], state["workbook"].customer_count, len(errors),
                              len([d for d in state["decisions"] if d.decision_id not in state["resolved_decisions"]]))

        # Global audit — include snapshots + col_idx so Changes Summary can
        # reconstruct per-row old→new pairs later.
        # Both `column` (friendly) and `sap_field` (SAP code) are stored
        # so the changes_summary endpoint can resolve to a friendly label
        # even when the friendly one wasn't recorded (legacy entries) or
        # when the friendly side ages out of sync with the source upload.
        audit_details = {
            "column": decision.column_label,
            "sap_field": decision.sap_field,
            "value": req.value or "",
            "strategy": req.strategy or "",
            "_col_idx": decision.col_idx,
            "_snapshots": entry.details.get("_snapshots", {}),
        }
        audit_log.log(
            user=user, action=f"decision_{req.action_id}",
            file_id=state["file_id"], filename=state["filename"],
            module=state["module"], sheet=decision.sheet, rule_id=decision.rule_id,
            affected_count=decision.affected_count,
            reason=req.reason or "",
            details=audit_details,
        )

        return {"success": True, "audit_entry": entry.as_dict(), "new_error_count": len(errors)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/session/undo/{audit_index}")
def undo(audit_index: int, session: str | None = Cookie(None), user: dict = Depends(current_user)):
    """Undo the nth (0-indexed) most recent entry in the audit log."""
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    # audit_log is in insertion order; latest is last
    # We accept negative indexes or index from end
    log = state["audit_log"]
    if audit_index < 0 or audit_index >= len(log):
        raise HTTPException(404, "Audit entry not found")

    entry = log[audit_index]
    ok = undo_entry(state["workbook"], entry)
    if not ok:
        raise HTTPException(400, "Cannot undo this action")

    # Remove the entry from audit log
    log.pop(audit_index)

    # If we undid a decision, unmark it resolved
    # Reconstruct resolved set from remaining entries that resolved decisions
    remaining_resolved = set()
    for e in log:
        # Not tracked perfectly — we use rule_id + sheet matching
        pass  # simpler: just fully revalidate

    # Revalidate + rebuild decisions
    errors = _revalidate_and_rebuild(state)
    # Recompute resolved set against the rebuilt decisions (IDs may have changed)
    active_rule_sheet_col = {(e.rule_id, e.sheet, e.col_idx) for e in errors}
    state["resolved_decisions"] = {
        d.decision_id for d in state["decisions"]
        if (d.rule_id, d.sheet, d.col_idx) not in active_rule_sheet_col
    }

    if state["file_id"]:
        xml_bytes = write_xml(state["workbook"])
        repo.save_working_copy(state["file_id"], xml_bytes)

    return {"success": True}


@app.get("/api/session/records/{sheet_name}/{row_idx}")
def get_record(sheet_name: str, row_idx: int,
               decision_id: str | None = None,
               session: str | None = Cookie(None),
               user: dict = Depends(current_user)):
    """Load a single record for the editor view.

    `decision_id` (optional query param) scopes j/k navigation — when present,
    `next_error_row`/`prev_error_row`/`error_rows_in_sheet` only contain row
    indexes that this specific decision flagged. Without it, navigation
    walks every error on the sheet (legacy behaviour for direct links).
    """
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    # PP record-editor view is not yet implemented (M2 follow-up). For now
    # the Error Grid is the primary surface for PP; clicking a row in the
    # grid won't open this endpoint for PP sessions because the UI doesn't
    # link there. This guard returns a friendly 501 if the URL is hit
    # directly (e.g. via the Records page sidebar), so PP users see a
    # clear message rather than a 500 about state["workbook"] being None.
    if state.get("pp_bundle"):
        raise HTTPException(
            501,
            "PP record editor is not yet implemented. "
            "Use the Error Grid to review issues; fix them in the source "
            "BOM/Routing file and re-upload.",
        )

    # ── MM branch: the "sheet" concept is "Materials" and rows come from
    #    the merged bundle, not a Workbook. Build an SD-shaped response
    #    so the existing record-editor UI renders without changes.
    if state.get("mm_bundle"):
        bundle = state["mm_bundle"]
        materials = bundle["merged"].materials
        main_fields = bundle["main_loaded"].sap_fields  # ordered SAP codes

        if sheet_name != "Materials":
            raise HTTPException(404, f"Sheet not found (MM has only 'Materials', got '{sheet_name}')")
        if not (0 <= row_idx < len(materials)):
            raise HTTPException(404, "Material not found")

        material = materials[row_idx]

        # Decision scope for j/k navigation
        decision = None
        if decision_id:
            decision = next((d for d in state["decisions"]
                             if d.decision_id == decision_id), None)

        # Row-level errors for this material (same shape as SD)
        row_errors = []
        for e in state["errors"]:
            if e.sheet != "Materials" or e.row_idx != row_idx:
                continue
            d = e.as_dict()
            # No ETE type/length for MM (xlsx doesn't carry SAP DDIC types).
            # The record editor already treats these as optional.
            d["ete_type"] = ""
            d["ete_length"] = 0
            d["related_fields"] = []
            row_errors.append(d)

        # Build "grouped fields" for the editor. Group all main fields
        # under "Main" and alt/longtext under their own groups, mirroring
        # how SD groups fields by section.
        error_cols = {e["col_idx"] for e in row_errors}

        # v63: per-column error messages — drives the tooltip on the
        # red "!" indicator in the Records editor. Before v63 the pill
        # was just a visual marker with no explanation on hover. Now
        # each field carries the list of error messages flagging it so
        # the UI can show "Material Number (MATNR) does not conform to
        # the number range for its Material Type ZATD" as a hover tip
        # instead of forcing the SME to leave the editor and scan the
        # decisions list to figure out what's wrong.
        errors_by_col: dict[int, list[str]] = {}
        for e in row_errors:
            ci = e.get("col_idx")
            if ci is None:
                continue
            msg = e.get("message") or e.get("rule_description") or ""
            if msg:
                errors_by_col.setdefault(ci, []).append(msg)

        # Friendly labels: row 1 of the source upload. header_labels is a
        # parallel list to sap_fields — same length, same order — so we
        # zip them. When the source has no friendly label for a column
        # (sap_field-only, e.g. older templates), fall back to the SAP
        # code so the UI never shows a blank label.
        labels = (bundle["main_loaded"].header_labels
                  if hasattr(bundle["main_loaded"], "header_labels")
                  else [])
        # The mandatory-field set drives the asterisk indicator in the editor.
        # Pulled from the v58-narrowed list so SMEs see exactly which 16
        # fields the validator will hard-fail on if blank.
        try:
            from services.mm_mandatory_by_team import MANDATORY_FIELDS_BY_TEAM
            mandatory_set = {f for fields in MANDATORY_FIELDS_BY_TEAM.values()
                              for f in fields}
        except Exception:
            mandatory_set = set()

        # Main fields in original column order; 1-based col_idx to match SD
        main_group = []
        for i, sap_code in enumerate(main_fields, start=1):
            raw = material.main.get(sap_code)
            # i-1 because labels list is 0-indexed; sap_code as fallback.
            friendly = (labels[i - 1] if 0 <= i - 1 < len(labels) and labels[i - 1]
                        else sap_code)
            main_group.append({
                "col_idx": i,
                "label": friendly,           # business-friendly name, e.g. "Sales org."
                "sap_field": sap_code,       # SAP code, e.g. "VKORG" — shown small below
                "value": "" if raw is None else (
                    str(int(raw)) if isinstance(raw, float) and raw.is_integer() else str(raw)
                ),
                "ete_length": 0,
                "ete_type": "",
                "mandatory": sap_code in mandatory_set,
                "has_error": i in error_cols,
                # v63: messages for this specific field's errors on this row.
                # Frontend renders these inside the red "!" pill's title attr
                # so the SME can hover and see WHY the field is flagged
                # without leaving the Records editor.
                "error_messages": errors_by_col.get(i, []),
            })

        grouped: dict[str, list[dict]] = {"Main": main_group}

        # Alt UoMs as a repeated-group (rendered read-only in the editor)
        if material.alt_uoms:
            alt_rows = []
            for a in material.alt_uoms:
                alt_rows.append({
                    "col_idx": 0, "label": "Alt UoM",
                    "sap_field": "MEINH",
                    "value": f"{a.get('MEINH','')}  ({a.get('UMREZ','')} / {a.get('UMREN','')})",
                    "ete_length": 0, "ete_type": "", "mandatory": False,
                    "has_error": False,
                    "error_messages": [],  # v63: tooltip data, empty for read-only rows
                })
            grouped["Alt UoMs"] = alt_rows

        if material.longtexts:
            lt_rows = []
            for t in material.longtexts:
                lt_rows.append({
                    "col_idx": 0, "label": "BASE_TEXT",
                    "sap_field": "BASE_TEXT",
                    "value": t.get("BASE_TEXT", "") or "",
                    "ete_length": 0, "ete_type": "", "mandatory": False,
                    "has_error": False,
                    "error_messages": [],  # v63: tooltip data, empty for read-only rows
                })
            grouped["Long Texts"] = lt_rows

        # Navigation pointers (same shape as SD response)
        if decision is not None:
            error_row_ids = sorted(decision.error_row_indexes)
        else:
            error_row_ids = sorted({e.row_idx for e in state["errors"]
                                    if e.sheet == "Materials"})
        next_error = next((i for i in error_row_ids if i > row_idx), None)
        prev_error = next((i for i in reversed(error_row_ids) if i < row_idx), None)

        return {
            "sheet": "Materials",
            "row_idx": row_idx,
            "xml_row": material.source_excel_row,
            "total_rows_in_sheet": len(materials),
            "groups": grouped,
            "errors": row_errors,
            "context": {},
            "next_error_row": next_error,
            "prev_error_row": prev_error,
            "error_rows_in_sheet": error_row_ids,
        }

    # ── SD path (unchanged) ──
    sheet = state["workbook"].sheets.get(sheet_name)
    if not sheet:
        raise HTTPException(404, f"Sheet not found")
    if not (0 <= row_idx < len(sheet.data_rows)):
        raise HTTPException(404, f"Row not found")

    # Resolve the optional decision scope
    decision = None
    if decision_id:
        decision = next((d for d in state["decisions"]
                         if d.decision_id == decision_id), None)
        # If the requested decision doesn't exist (e.g. was resolved + cleaned),
        # silently fall back to sheet-wide navigation. Don't 404 — the record
        # is still valid, we just can't scope navigation.

    row = sheet.data_rows[row_idx]

    # Build a quick col_idx -> spec lookup for this sheet
    spec_by_col = {s.col_idx: s for s in sheet.specs}

    # Related fields — shown read-only next to the editor for context so
    # users don't have to expand "All fields" to see related info.
    # Example: for a dl_expired error, showing the DL Number alongside the
    # expiry date helps decide renewal. For state errors, showing the
    # Postal Code + Country helps verify the right state code.
    related_fields_by_rule: dict[str, list[str]] = {
        "dl_expired":                 ["Drug Licence Number"],
        "invalid_state_in":           ["Postal Code", "Country/Region", "City"],
        "invalid_pan":                ["GSTIN", "Name"],
        "gstin_length":               ["State", "Name"],
        "gstin_format":               ["State", "Name"],
        "pan_format":                 ["GSTIN", "Name"],
        "inco_location_description":  ["Incoterms"],
        "mandatory_missing":          [],
    }

    # Errors on this row, augmented with spec info so the client can pick
    # the right editor (date picker, truncate button, etc.)
    row_errors = []
    for e in state["errors"]:
        if e.sheet != sheet_name or e.row_idx != row_idx:
            continue
        d = e.as_dict()
        spec = spec_by_col.get(e.col_idx)
        if spec:
            d["ete_type"] = spec.ete_type       # 'C', 'N', 'D', ...
            d["ete_length"] = spec.ete_length   # max allowed length
        # Attach related-field snapshots (label + value) based on rule.
        related_labels = related_fields_by_rule.get(e.rule_id, [])
        related = []
        for rl in related_labels:
            rspec = next((s for s in sheet.specs if s.label == rl), None)
            if rspec:
                raw = row.get(rspec.col_idx)
                related.append({
                    "label": rl,
                    "value": "" if raw is None else str(raw),
                })
        d["related_fields"] = related
        row_errors.append(d)

    # Find next/prev error rows in same sheet (for error navigation).
    # If we're inside a decision's flow, scope to rows that THIS decision
    # flagged — otherwise j/k walks every error, which is disorienting
    # when the user entered via "Fix individually" on one specific rule.
    if decision is not None:
        error_row_ids = sorted(decision.error_row_indexes)
    else:
        error_row_ids = sorted({e.row_idx for e in state["errors"] if e.sheet == sheet_name})
    next_error = next((i for i in error_row_ids if i > row_idx), None)
    prev_error = next((i for i in reversed(error_row_ids) if i < row_idx), None)

    # Grouped fields
    grouped: dict[str, list[dict]] = {}
    error_cols = {e["col_idx"] for e in row_errors}
    # v63: per-column error messages for the tooltip on the red "!"
    # indicator. Mirrors the MM path above. Same shape, same purpose.
    sd_errors_by_col: dict[int, list[str]] = {}
    for e in row_errors:
        ci = e.get("col_idx")
        if ci is None:
            continue
        msg = e.get("message") or e.get("rule_description") or ""
        if msg:
            sd_errors_by_col.setdefault(ci, []).append(msg)
    for spec in sheet.specs:
        group = spec.group or "General"
        grouped.setdefault(group, []).append({
            "col_idx": spec.col_idx,
            "label": spec.label,
            "sap_field": spec.sap_field,
            "value": "" if (v := row.get(spec.col_idx, "")) is None else str(v),
            "ete_length": spec.ete_length,
            "ete_type": spec.ete_type,
            "mandatory": spec.mandatory,
            "has_error": spec.col_idx in error_cols,
            "error_messages": sd_errors_by_col.get(spec.col_idx, []),
        })

    # Linked context
    context: dict = {}
    if sheet_name == "General Data":
        num_spec = next((s for s in sheet.specs if s.label == "Customer Number"), None)
        if num_spec:
            cust_num = str(row.get(num_spec.col_idx, ""))
            for linked in ("Sales Data", "Tax Numbers", "Company Data", "BP Roles"):
                lsheet = state["workbook"].sheets.get(linked)
                if not lsheet:
                    continue
                lspec = next((s for s in lsheet.specs if s.label == "Customer Number"), None)
                if not lspec:
                    continue
                count = sum(1 for r in lsheet.data_rows if str(r.get(lspec.col_idx, "")) == cust_num)
                context[linked] = count

    return {
        "sheet": sheet_name,
        "row_idx": row_idx,
        "xml_row": row_idx + 9,
        "total_rows_in_sheet": len(sheet.data_rows),
        "groups": grouped,
        "errors": row_errors,
        "context": context,
        "next_error_row": next_error,
        "prev_error_row": prev_error,
        "error_rows_in_sheet": error_row_ids,
    }


class EditCellRequest(BaseModel):
    sheet: str
    row_idx: int
    col_idx: int
    value: Any
    # Optional SAP field code — required for MM when the field is not
    # in the original main file column set (SPRAS, ALAND, WAERS, CURTP,
    # BWKEY are LTMC-defaults that don't have a source column). For SD
    # and for MM fields-in-source, col_idx alone is sufficient and this
    # is ignored.
    sap_field: str | None = None


@app.post("/api/session/records/edit")
def edit_cell(req: EditCellRequest, session: str | None = Cookie(None), user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    # PP doesn't support inline cell editing in this build (M2 follow-up).
    # SMEs fix issues in the source BOM/Routing file and re-upload.
    if state.get("pp_bundle"):
        raise HTTPException(
            501,
            "PP cell editing is not yet implemented. "
            "Fix issues in the source file and re-upload.",
        )

    # ── MM branch: edit a field on a merged material in-memory, re-run the
    #    MM validator, and return the same response shape SD uses. LTMC
    #    export later reads from state["mm_bundle"]["merged"] so edits
    #    flow through to the generated XML.
    if state.get("mm_bundle"):
        if req.sheet != "Materials":
            raise HTTPException(400, "MM edits must target sheet='Materials'")
        bundle = state["mm_bundle"]
        materials = bundle["merged"].materials
        main_fields = bundle["main_loaded"].sap_fields
        if not (0 <= req.row_idx < len(materials)):
            raise HTTPException(404, "Material not found")
        # Map col_idx → SAP code. For original main-file columns col_idx
        # is 1-based within main_fields. For LTMC-default fields (SPRAS,
        # ALAND, WAERS, CURTP, BWKEY, etc.) the field doesn't exist in the
        # source — col_idx is a synthetic value and the frontend MUST
        # supply sap_field explicitly. Without that we can't know which
        # field the user meant to edit.
        if 1 <= req.col_idx <= len(main_fields):
            sap_code = main_fields[req.col_idx - 1]
            # If frontend sent sap_field too, prefer it (defensive)
            if req.sap_field and req.sap_field != sap_code:
                sap_code = req.sap_field
        elif req.sap_field:
            sap_code = req.sap_field
        else:
            raise HTTPException(
                400,
                f"col_idx {req.col_idx} is outside the source column range "
                f"and no sap_field provided. The frontend must include "
                f"sap_field for LTMC-default fields not present in the "
                f"source main file."
            )
        material = materials[req.row_idx]
        old_value = material.main.get(sap_code)
        # Apply (None for blanks so mandatory checks fire)
        new_val = req.value

        # ── v67: BWKEY edit ALWAYS auto-derives from each plant's WERKS ──
        # The bug: for plant-level valuation (HML's setup), the rule is
        # BWKEY = WERKS *per plant_row*. When a multi-plant material has
        # plant_rows [PE01, PE02], the only correct state is plant_rows[0]
        # with BWKEY=PE01 and plant_rows[1] with BWKEY=PE02. There is no
        # single "BWKEY value" for the material as a whole.
        #
        # Before this fix: user edits BWKEY to PE02 → v65 propagates PE02
        # to all plant_rows → plant_row PE01 now has BWKEY=PE02 mismatched
        # with WERKS=PE01 → error persists, decision never dismisses. SME
        # screenshot showed exactly this: "Valuation Area is PE01 but
        # Plant is PE02" — they typed PE02, expected dismissal, didn't get
        # it because the OTHER plant_row's mismatch flared up.
        #
        # The fix: when sap_code is BWKEY, ignore the typed value and set
        # each plant_row's BWKEY equal to its own WERKS. The user's typed
        # value is treated as a "fix this" gesture rather than as a literal
        # value to write. Audit log records what we actually did so the
        # paper trail is honest. This also short-circuits the v65 generic
        # propagation below (which would otherwise overwrite our per-plant
        # values with the user's typed one).
        if sap_code == "BWKEY":
            # Inline string-coercer (avoids importing from validator module
            # since main.py shouldn't grow that dependency for one call).
            def _as_str_local(v):
                if v is None:
                    return ""
                if isinstance(v, float) and v.is_integer():
                    return str(int(v))
                return str(v).strip()
            for pr in material.plant_rows:
                pr_werks = _as_str_local(pr.get("WERKS"))
                if pr_werks:
                    pr.values["BWKEY"] = pr_werks
                else:
                    pr.values.pop("BWKEY", None)
            # The user's typed value is irrelevant — overwrite for the
            # audit-log "new_value" with what we actually wrote on main.
            new_val = _as_str_local(material.main.get("WERKS")) or ""

        # ── v67: VPRSV (price control) normalization ─────────────────
        # Rule `mm_price_control_missing_or_invalid` requires VPRSV to be
        # exactly "S" or "V" (uppercase). Users typing "s" or "v" had
        # their edit silently rejected by the validator on next run, so
        # the decision didn't dismiss. Uppercase the value to be forgiving
        # of typing style. Only applies to VPRSV (other letter-code fields
        # may be case-sensitive at SAP — only normalize where we know).
        elif sap_code == "VPRSV":
            if isinstance(new_val, str):
                new_val = new_val.strip().upper()

        if isinstance(new_val, str) and new_val.strip() == "":
            material.main.values.pop(sap_code, None)
        else:
            material.main.values[sap_code] = new_val

        # ── v65: propagate edits of plant-scoped fields to ALL plant rows ──
        # The bug we're fixing: when a multi-plant material has e.g. an
        # invalid LGORT on plant PE02 (but a different LGORT or blank on
        # PE01), the Records editor only shows `material.main` (which is
        # plant_rows[0] = PE01's row). Editing LGORT updates main only.
        # The validator's `_handle_kds_nested_lookup` iterates ALL plant_rows
        # — so the bad LGORT on PE02 still fires the error after the user
        # "fixes" the value via the editor. The decision never dismisses.
        #
        # HML's typical pattern is one storage-location code per material
        # used uniformly across plants, so propagating the new value to
        # every plant_row matches their data model and lets the SME
        # actually clear the error with one edit.
        #
        # v67: skip propagation for BWKEY (handled above per-plant) since
        # its correct value differs per plant_row.
        #
        # We skip identity fields (WERKS uniquely identifies a plant_row;
        # MATNR is the material key). For those the edit only updates main.
        # All other plant-scoped fields propagate.
        _IDENTITY_FIELDS = {"WERKS", "MATNR", "REC_NO", "BWKEY"}
        if sap_code not in _IDENTITY_FIELDS:
            for pr in material.plant_rows:
                # plant_rows[0] is the same object as material.main for
                # single-plant materials (already updated above). For
                # multi-plant the other plant_row objects need explicit
                # propagation here.
                if pr is material.main:
                    continue
                if isinstance(new_val, str) and new_val.strip() == "":
                    pr.values.pop(sap_code, None)
                else:
                    pr.values[sap_code] = new_val

        # Re-validate the whole session — same pattern as SD. Cheap enough
        # for MM (thousands of materials × ~65 rules = under 1s).
        from services.mm_validator import validate_mm as _mm_validate
        bundle = state.get("mm_bundle") or {}
        alt_loaded = bundle.get("alt_loaded")
        lt_loaded = bundle.get("lt_loaded")
        errors = _mm_validate(
            materials, _get_mm_catalogs(), main_fields,
            merged_result=bundle.get("merged"),
            alt_uom_rows=(alt_loaded.rows if alt_loaded else None),
            longtext_rows=(lt_loaded.rows if lt_loaded else None),
            friendly_labels=getattr(bundle.get("main_loaded"), "header_labels", None),
        )
        accepted = state.get("accepted_errors") or set()
        if accepted:
            errors = [e for e in errors
                      if (e.rule_id, e.sheet, e.col_idx, e.row_idx) not in accepted]
        state["errors"] = errors
        state["decisions"] = group_errors(errors)
        state["dirty"] = True

        # Build an audit entry compatible with the SD shape the UI expects
        entry = AuditEntry(
            timestamp=dt_module.datetime.now().isoformat(timespec="seconds"),
            user=user.get("display_name", user["username"]),
            action="edit_cell",
            rule_id="",
            sheet="Materials",
            affected_count=1,
            reason=f"Edited {sap_code}",
            details={
                "row_idx": req.row_idx,
                "col_idx": req.col_idx,
                "column_label": sap_code,
                "sap_field": sap_code,
                "old_value": "" if old_value is None else str(old_value),
                "new_value": str(new_val),
            },
        )
        state["audit_log"].append(entry)

        audit_log.log(
            user=user, action="edit_cell",
            file_id=state["file_id"], filename=state["filename"],
            module="MM", sheet="Materials", affected_count=1,
            details={
                "row_idx": req.row_idx, "col_idx": req.col_idx,
                "sap_field": sap_code,
                "old_value": "" if old_value is None else str(old_value)[:200],
                "new_value": str(new_val)[:200],
            },
        )
        return {"success": True, "audit_entry": entry.as_dict(),
                "new_error_count": len(errors)}

    # ── SD path ──
    try:
        entry = apply_single_edit(state["workbook"], req.sheet, req.row_idx, req.col_idx, req.value,
                                   user=user["display_name"])
        state["audit_log"].append(entry)
        errors = _revalidate_and_rebuild(state)

        # Flag the workbook as having unsaved changes; the XML is regenerated
        # lazily on export/bulk-action/explicit-save. Writing 25 MB of XML on
        # every keystroke was adding ~4 seconds per edit on large files.
        state["dirty"] = True

        # Look up column label for audit clarity
        sheet = state["workbook"].sheets.get(req.sheet)
        spec = next((s for s in sheet.specs if s.col_idx == req.col_idx), None) if sheet else None
        col_label = spec.label if spec else f"col {req.col_idx}"

        audit_log.log(
            user=user, action="edit_cell",
            file_id=state["file_id"], filename=state["filename"],
            module=state["module"], sheet=req.sheet,
            affected_count=1,
            details={
                "row_idx": req.row_idx,
                "col_idx": req.col_idx,
                "column_label": col_label,
                "old_value": str(entry.details.get("old_value", ""))[:200],
                "new_value": str(req.value)[:200],
            },
        )

        return {"success": True, "audit_entry": entry.as_dict(), "new_error_count": len(errors)}
    except Exception as e:
        raise HTTPException(500, str(e))


# ────────────────────────────────────────────────────────────────────────────
# Find & Replace across error cells
# ────────────────────────────────────────────────────────────────────────────

class FindReplaceRequest(BaseModel):
    """Find/replace scoped to error cells only. Never touches clean data.

    Scope defaults to the whole workbook's flagged cells. Narrow with:
      - sheet           → restrict to one sheet
      - column_label    → restrict to one column on that sheet

    Matching is exact by default; two optional relaxations:
      - case_insensitive → 'DELHI' == 'delhi'
      - trim_whitespace  → 'Delhi ' == 'Delhi'

    confirm=False asks the server to count matches and return them for
    preview; the client may then resend with confirm=True to actually
    write the changes. The > 10 match confirmation threshold is enforced
    server-side so a misbehaving client can't skip it.
    """
    sheet: str | None = None
    column_label: str | None = None
    find: str
    replace: str
    case_insensitive: bool = False
    trim_whitespace: bool = False
    reason: str = ""
    confirm: bool = False


FIND_REPLACE_HARD_CAP = 500
FIND_REPLACE_CONFIRM_THRESHOLD = 10


def _fr_matches(value: Any, needle: str, *,
                case_insensitive: bool, trim_whitespace: bool) -> bool:
    """Return True if `value` matches `needle` under the rules.
    Exact match only — no substring/regex. Treats None as blank."""
    s = "" if value is None else str(value)
    n = needle
    if trim_whitespace:
        s = s.strip()
        n = n.strip()
    if case_insensitive:
        return s.casefold() == n.casefold()
    return s == n


@app.post("/api/session/find_replace")
def find_replace(req: FindReplaceRequest,
                  session: str | None = Cookie(None),
                  user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    if not req.find:
        raise HTTPException(400, "'Find' cannot be blank.")
    if req.find == req.replace and not (req.case_insensitive or req.trim_whitespace):
        raise HTTPException(400, "Find and Replace are identical — nothing to do.")

    wb = state["workbook"]

    # Build list of errors restricted to the chosen scope. Find-replace
    # ONLY touches cells that a validator already flagged, so clean data
    # is never accidentally modified.
    scope_errors = []
    for e in state["errors"]:
        if req.sheet and e.sheet != req.sheet:
            continue
        if req.column_label and e.column_label != req.column_label:
            continue
        scope_errors.append(e)

    # Find actual matches by reading the current cell value
    matches = []
    seen_cells = set()
    for e in scope_errors:
        key = (e.sheet, e.row_idx, e.col_idx)
        if key in seen_cells:
            # Same cell might be flagged by multiple rules; count once.
            continue
        sheet_obj = wb.sheets.get(e.sheet)
        if not sheet_obj or not (0 <= e.row_idx < len(sheet_obj.data_rows)):
            continue
        current_value = sheet_obj.data_rows[e.row_idx].get(e.col_idx)
        if _fr_matches(current_value, req.find,
                       case_insensitive=req.case_insensitive,
                       trim_whitespace=req.trim_whitespace):
            seen_cells.add(key)
            matches.append({
                "sheet": e.sheet,
                "row_idx": e.row_idx,
                "col_idx": e.col_idx,
                "column_label": e.column_label,
                "xml_row": e.xml_row,
                "current_value": str(current_value) if current_value is not None else "",
                "rule_id": e.rule_id,
            })

    total = len(matches)

    if total > FIND_REPLACE_HARD_CAP:
        raise HTTPException(
            400,
            f"{total} matches exceed the hard limit of {FIND_REPLACE_HARD_CAP}. "
            f"Narrow the scope (pick a specific column) or use Bulk Change on a decision."
        )

    # Preview (dry-run)
    if not req.confirm:
        return {
            "matches": matches[:100],
            "preview_truncated": total > 100,
            "total_matches": total,
            "needs_confirmation": total > FIND_REPLACE_CONFIRM_THRESHOLD,
        }

    # Confirm mode — write changes. Enforce reason for large batches.
    if total > FIND_REPLACE_CONFIRM_THRESHOLD and not req.reason.strip():
        raise HTTPException(400,
            f"A reason is required when replacing more than "
            f"{FIND_REPLACE_CONFIRM_THRESHOLD} cells.")

    if total == 0:
        return {"replaced_count": 0, "matches": [], "new_error_count": len(state["errors"])}

    # Apply via the SAME path as single-cell edits. Each replaced cell
    # becomes its own audit entry, which preserves per-cell undo
    # (Recent Activity can reverse any individual cell).
    replaced = []
    for m in matches:
        try:
            entry = apply_single_edit(
                wb, m["sheet"], m["row_idx"], m["col_idx"], req.replace,
                user=user["display_name"],
            )
            # Annotate for audit trail + undo context
            entry.reason = req.reason
            entry.details["find_replace_batch"] = True
            entry.details["find"] = req.find
            entry.details["replace"] = req.replace
            state["audit_log"].append(entry)
            replaced.append(m)
        except Exception as exc:
            # One cell failing shouldn't abort the whole batch.
            print(f"[find_replace] edit failed for {m}: {exc}")

    errors = _revalidate_and_rebuild(state)
    state["dirty"] = True

    # Consolidated audit entry for the batch (separate from per-cell entries).
    # Shows up in Audit Trail as "Sanjay replaced 23 cells of 'DELHI' → 'Delhi'"
    # instead of 23 individual lines.
    audit_log.log(
        user=user, action="find_replace",
        file_id=state["file_id"], filename=state["filename"],
        module=state["module"], sheet=req.sheet or "(all sheets)",
        affected_count=len(replaced), reason=req.reason,
        details={
            "column_label": req.column_label or "(all columns)",
            "find": req.find[:100],
            "replace": req.replace[:100],
            "case_insensitive": req.case_insensitive,
            "trim_whitespace": req.trim_whitespace,
            "replaced_count": len(replaced),
        },
    )

    return {
        "replaced_count": len(replaced),
        "matches": replaced[:100],
        "total_matches": total,
        "new_error_count": len(errors),
    }


# ────────────────────────────────────────────────────────────────────────────
# Group-by-value bulk replace
# ────────────────────────────────────────────────────────────────────────────
#
# Pattern: show the SME every distinct bad value flagged by ONE decision,
# with counts, and let them type the replacement next to each. For KDS-
# validated columns, an "auto-populate" toggle fills in closest-match
# suggestions from the catalog. Much faster than Fix Individually when
# errors cluster into a small number of distinct values (typical for
# categorical columns: Incoterms, Sales Group, Customer Group, State).
#
# Endpoints:
#   GET  /api/session/decisions/{id}/groups        — load distinct values
#   POST /api/session/decisions/{id}/group_replace — apply replacements

# A decision is "categorical" (worth group-by-value UI) when:
#   - it has a catalog backing it (any *_not_in_kds rule + inco_location),
#   - OR its errors cluster into <= 20 distinct values (small enough to fit
#     on screen and make individual-replacement UI useful).
GROUP_CATEGORICAL_MAX_DISTINCT = 20

# Caps for group-replace. These are deliberately different from find/replace:
# the whole point of group-replace is that ONE correct replacement can fix
# thousands of cells (e.g. one typo used in 18,000 rows). So the cap is on
# the number of distinct rules (typos/variants) being applied at once, not
# on the total cell count.
GROUP_REPLACE_MAX_DISTINCT_RULES = 50     # > 50 distinct rules = SME should sanity-check
GROUP_REPLACE_HARD_CAP_CELLS = 50_000     # absolute cell ceiling (a whole file's worth)
GROUP_REPLACE_CONFIRM_THRESHOLD = 10      # cells above which the batch needs a reason


@app.get("/api/session/decisions/{decision_id:path}/history")
def decision_history(decision_id: str,
                     session: str | None = Cookie(None),
                     user: dict = Depends(current_user)):
    """v67: return the audit-log entries that pertain to a specific decision.

    The SME wants per-decision history: "for this BWKEY decision, what was
    edited, by whom, when, with what reason." Without this they'd have to
    scan the whole session-wide audit log and filter mentally.

    We match audit entries to the decision via two channels:
      1. `rule_id` — every action carries the rule_id of the decision it
         was acting on (edit_cell, group_replace, set_ltmc_default,
         decision_accept_all all stamp this).
      2. `details.col_idx` / `details.sap_field` — narrower match for
         per-cell edits that affect the decision's specific column.

    Returned newest-first so the SME sees the most recent attempt first.
    Each entry includes a `summary` line that renders cleanly inline
    (e.g. "Set LGORT='FEU1' for row 8 — by Admin · 2026-05-11 14:23")
    so the UI can show the history as a simple list without parsing
    `details` structures itself.
    """
    state = _get_session_state(session)
    decisions = state.get("decisions") or []
    decision = next((d for d in decisions if d.decision_id == decision_id), None)
    if not decision:
        raise HTTPException(404, "Decision not found")

    audit_log_entries = state.get("audit_log") or []
    matching: list[dict] = []
    for entry in reversed(audit_log_entries):   # newest first
        # Primary match: rule_id matches the decision's rule
        if getattr(entry, "rule_id", None) != decision.rule_id:
            continue
        details = getattr(entry, "details", {}) or {}
        # Build a readable one-line summary
        action = entry.action
        if action == "set_ltmc_default":
            summary = (f"Set LTMC default {details.get('sap_field','')}="
                       f"'{details.get('value','')}' "
                       f"(applies to all {details.get('affected_count','?')} rows)")
        elif action == "group_replace":
            summary = (f"Group replace '{details.get('find','')}' "
                       f"→ '{details.get('replace','')}' "
                       f"on row {details.get('row_idx','?')}")
        elif action == "edit_cell":
            sap_field = details.get('sap_field', '') or details.get('column_label', '')
            summary = (f"Edited {sap_field}: "
                       f"'{details.get('old_value','')}' → "
                       f"'{details.get('new_value','')}' "
                       f"on row {details.get('row_idx','?')}")
        elif action == "decision_accept_all":
            summary = (f"Marked all {details.get('affected_count','?')} "
                       f"errors as accepted (won't block export)")
        else:
            summary = f"{action} on row {details.get('row_idx','?')}"

        matching.append({
            "timestamp": entry.timestamp,
            "user": entry.user,
            "action": action,
            "summary": summary,
            "reason": entry.reason or "",
            "affected_count": getattr(entry, "affected_count", 1),
            "details": details,
        })

    return {
        "decision_id": decision_id,
        "rule_id": decision.rule_id,
        "rule_name": decision.rule_name,
        "column_label": decision.column_label,
        "sap_field": decision.sap_field,
        "history": matching,
        "history_count": len(matching),
    }


@app.get("/api/session/decisions/{decision_id:path}/groups")
def decision_groups(decision_id: str,
                     session: str | None = Cookie(None),
                     user: dict = Depends(current_user)):
    """Return the distinct bad values for one decision, with counts and
    (if the rule is catalog-backed) closest-match suggestions from KDS.

    Response shape:
        {
          "column_label": "Incoterms Location1",
          "sheet": "General Data",
          "total_errors": 1234,
          "distinct_values": 12,
          "is_categorical": true,     # UI should offer group-replace
          "has_catalog": true,        # KDS suggestions available
          "catalog_sample": [...],    # first N entries (for dropdown preview)
          "groups": [
            {
              "value": "DELHI",
              "count": 23,
              "sample_rows": [12, 45, 67],
              "suggestions": [
                {"code": "07", "description": "Delhi", "ratio": 1.0},
                ...
              ]
            },
            ...
          ]
        }
    """
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    decision = next((d for d in state["decisions"]
                     if d.decision_id == decision_id), None)
    if not decision:
        raise HTTPException(404, "Decision not found")

    # ── MM branch ──────────────────────────────────────────────────────────
    # MM sessions have no "workbook" — data lives in mm_bundle.merged.materials.
    # Each error for this decision represents ONE cell: for plant-scoped fields
    # (LGPRO/LGORT/WERKS), one error per (material × plant); for plant-
    # independent fields (MATKL/BKLAS), one error per material. We just
    # iterate the error list and bucket by value — the errors already have
    # the correct per-cell granularity, so the count matches affected_count.
    if state.get("mm_bundle"):
        bundle = state["mm_bundle"]
        materials = bundle["merged"].materials

        buckets: dict[str, dict] = {}
        first_err_with_options = None
        for e in state["errors"]:
            if e.rule_id != decision.rule_id:
                continue
            if e.sap_field != decision.sap_field:
                continue
            val = e.value
            bucket = buckets.setdefault(val, {
                "value": val, "count": 0, "sample_rows": [],
                "validator_suggestion": None,
            })
            bucket["count"] += 1
            if len(bucket["sample_rows"]) < 5 and e.row_idx not in bucket["sample_rows"]:
                bucket["sample_rows"].append(e.row_idx)
            if bucket["validator_suggestion"] is None and e.suggested_value:
                bucket["validator_suggestion"] = e.suggested_value
            if first_err_with_options is None and e.suggested_options:
                first_err_with_options = e

        groups = sorted(buckets.values(), key=lambda b: -b["count"])

        # MM suggestions: pull from the error's suggested_options. Same
        # (rule_id, value) share the same options (for flat catalogs).
        for g in groups:
            suggestions: list[dict] = []
            vs = g.pop("validator_suggestion", None)
            if vs:
                suggestions.append({
                    "code": vs, "description": "", "ratio": 1.0,
                    "source": "validator",
                })
            # Use the first-with-options as the sample for all groups of
            # this rule. Flat-catalog rules (Plant, MATKL) have the same
            # options on every error; scoped rules (LGPRO) differ per
            # plant — in which case the per-error dropdown in Fix
            # Individually has the scoped options, and this group-level
            # sample is the catalog-wide typical set.
            if first_err_with_options:
                for opt in first_err_with_options.suggested_options[:3]:
                    if opt["value"] != vs:
                        suggestions.append({
                            "code": opt["value"],
                            "description": opt["label"],
                            "ratio": 0.0,
                            "source": "kds",
                        })
            g["suggestions"] = suggestions

        # Catalog sample for the UI's typeahead dropdown.
        # v66.1: bumped from 50 → 300 so the ISO UoM catalog (230 codes)
        # fits in full. The original 50 cap was chosen when only flat
        # KDS catalogs (Material Group: 451, Sales Group: 1653) drove this
        # — those are still capped at 300 (still scrollable in a <select>
        # but no longer truncates ISO UoM at "F"). For very large catalogs
        # SMEs should still use the free-text input next to the dropdown.
        catalog_sample = []
        if first_err_with_options:
            catalog_sample = [
                {"code": o["value"], "description": o["label"]}
                for o in first_err_with_options.suggested_options[:300]
            ]

        distinct = len(groups)
        is_categorical = bool(catalog_sample) or distinct <= GROUP_CATEGORICAL_MAX_DISTINCT

        return {
            "decision_id": decision_id,
            "rule_id": decision.rule_id,
            "rule_name": decision.rule_name,
            "sheet": decision.sheet,
            "column_label": decision.column_label,
            "total_errors": decision.affected_count,
            "distinct_values": distinct,
            "is_categorical": is_categorical,
            "has_catalog": bool(catalog_sample),
            "catalog_sample": catalog_sample,
            "groups": groups,
        }

    # ── SD branch (original logic unchanged) ───────────────────────────────
    wb = state["workbook"]
    sheet = wb.sheets.get(decision.sheet)
    if not sheet:
        raise HTTPException(404, f"Sheet not found: {decision.sheet}")

    # Collect the current cell value for each row flagged by this decision.
    # Dedup by (row, col) — a cell flagged by multiple rules should only be
    # counted once in this decision's groups.
    # Build bucket-by-value index AND collect per-error suggested_value from
    # the validator. Some rules (invalid_state_in, gstin_checksum) don't have
    # a KDS catalog but the validator already knows the fix — we carry that
    # through to the UI so the "Replace with" gets a concrete suggestion.
    buckets: dict[str, dict] = {}   # value → {count, sample_rows, validator_suggestions}
    seen_cells: set[tuple[int, int]] = set()
    # Pre-index this decision's errors by (row, col) so we can look up
    # suggested_value without re-scanning the whole error list.
    err_by_cell: dict[tuple[int, int], "Error"] = {}
    for e in state["errors"]:
        if e.sheet == decision.sheet and e.rule_id == decision.rule_id:
            err_by_cell[(e.row_idx, e.col_idx)] = e
    for row_idx in decision.error_row_indexes:
        key = (row_idx, decision.col_idx)
        if key in seen_cells:
            continue
        seen_cells.add(key)
        if not (0 <= row_idx < len(sheet.data_rows)):
            continue
        raw = sheet.data_rows[row_idx].get(decision.col_idx)
        value = "" if raw is None else str(raw)
        bucket = buckets.setdefault(value, {
            "value": value, "count": 0, "sample_rows": [],
            "validator_suggestion": None,
        })
        bucket["count"] += 1
        if len(bucket["sample_rows"]) < 5:
            bucket["sample_rows"].append(row_idx)
        # Carry the validator's suggested_value through if present. For
        # a bucket, we take the first non-empty suggestion we see — all
        # errors in the same bucket share the same bad value, so they
        # should yield the same suggestion.
        if bucket["validator_suggestion"] is None:
            err = err_by_cell.get(key)
            if err is not None and getattr(err, "suggested_value", None):
                bucket["validator_suggestion"] = err.suggested_value

    # Sort by count descending so the most-impactful cleanup items are at top
    groups = sorted(buckets.values(), key=lambda b: -b["count"])

    # Build final suggestion list per group:
    #   1. Validator's suggested_value (if any) — highest confidence (1.0)
    #   2. KDS closest matches — if rule is catalog-backed
    catalog = kds.CATALOG_BY_RULE.get(decision.rule_id)
    has_catalog = catalog is not None

    for g in groups:
        suggestions: list[dict] = []
        # Validator-provided suggestion first (e.g. state code 8 → 08 Rajasthan)
        vs = g.pop("validator_suggestion", None)
        if vs:
            # Try to enrich with a description if we can resolve it. For
            # state codes this maps via INDIA_STATE_CODE_NAMES.
            from services.validator import INDIA_STATE_CODE_NAMES
            desc = INDIA_STATE_CODE_NAMES.get(vs, "")
            suggestions.append({
                "code": vs,
                "description": desc,
                "ratio": 1.0,
                "source": "validator",   # lets the UI badge it differently
            })
        # KDS closest-match as a secondary list
        if has_catalog:
            for s in kds.closest_matches(g["value"], catalog, top_n=3):
                s["source"] = "kds"
                suggestions.append(s)
        g["suggestions"] = suggestions

    if has_catalog:
        catalog_sample = [{"code": c, "description": d}
                          for c, d in list(catalog.items())[:50]]
    else:
        catalog_sample = []

    distinct = len(groups)
    is_categorical = has_catalog or distinct <= GROUP_CATEGORICAL_MAX_DISTINCT

    return {
        "decision_id": decision_id,
        "rule_id": decision.rule_id,
        "rule_name": decision.rule_name,
        "sheet": decision.sheet,
        "column_label": decision.column_label,
        "total_errors": decision.affected_count,
        "distinct_values": distinct,
        "is_categorical": is_categorical,
        "has_catalog": has_catalog,
        "catalog_sample": catalog_sample,
        "groups": groups,
    }


class GroupReplaceRequest(BaseModel):
    """Apply replacements built up in the group-by-value modal.

    `replacements` is a list of (find_value → replace_value) pairs.
    Only cells in the decision's error set that currently equal the
    `find_value` are changed. Unmatched cells are untouched — safe by
    construction.
    """
    replacements: list[dict]   # each: {"find": str, "replace": str}
    reason: str = ""
    confirm: bool = False


@app.post("/api/session/decisions/{decision_id:path}/group_replace")
def group_replace(decision_id: str, req: GroupReplaceRequest,
                   session: str | None = Cookie(None),
                   user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    decision = next((d for d in state["decisions"]
                     if d.decision_id == decision_id), None)
    if not decision:
        raise HTTPException(404, "Decision not found")

    # Normalise replacement rules, dropping blanks & noops.
    # Empty `find` is intentionally allowed: when a field is blank (e.g.
    # SPRAS not populated, ALAND missing), the "bad value" that Group &
    # Replace matches against IS the empty string. Rejecting it would
    # make "apply SPRAS=EN to all 10,565 rows that have SPRAS blank"
    # impossible. Empty `replace` is also allowed — SMEs sometimes want
    # to null out a cell (e.g. clear LGPRO=FEU1 and let LTMC default it).
    rules: dict[str, str] = {}
    for r in req.replacements:
        find = r.get("find")
        replace = r.get("replace")
        if find is None or replace is None:
            continue
        if find == replace:
            # noop — skip silently
            continue
        # Both "" → "EN" (fill blanks) and "FEU1" → "PKS1" (replace a value) valid
        rules[find] = replace

    if not rules:
        raise HTTPException(400,
            "No replacements to apply. Tick rows and provide replacement values first.")

    # ── MM branch ──────────────────────────────────────────────────────────
    # For MM, "cells" are (material × plant_row × sap_field) tuples, not
    # (row × col) in a sheet. We iterate every plant_row of every material
    # in the decision's error set. For plant-scoped fields (LGPRO/LGORT/
    # WERKS), the same bad value usually appears on every plant_row for a
    # material — so one material contributes N cells (N = # plant rows).
    # For plant-independent fields (MATKL/MTART/BKLAS), each material
    # contributes 1 cell (main only).
    if state.get("mm_bundle"):
        bundle = state["mm_bundle"]
        materials = bundle["merged"].materials
        main_fields = bundle["main_loaded"].sap_fields
        sap_field = decision.sap_field

        # ── v66.2: cross-file Group & Replace ──────────────────────────
        # When a decision's sheet is "AlternateUnits" or "LongText",
        # row_idx points into the alt_loaded.rows or lt_loaded.rows source
        # data — NOT into `materials`. The values to find/replace live in
        # those source rows (e.g. MEINH='PC' lives in alt_loaded.rows[i].
        # values['MEINH'], not in any material.main or plant_row). The
        # bug we're fixing: previously the MM branch always looked up
        # `materials[row_idx]` for cross-file decisions too, found no
        # matching values (because MEINH isn't a main-file field), and
        # silently produced 0 edits — SME saw "10 errors" become "10 errors"
        # after their replacement, with no error and no fix applied.
        if decision.sheet in ("AlternateUnits", "LongText"):
            alt_loaded = bundle.get("alt_loaded")
            lt_loaded = bundle.get("lt_loaded")
            source_loaded = alt_loaded if decision.sheet == "AlternateUnits" else lt_loaded
            if source_loaded is None or not getattr(source_loaded, "rows", None):
                raise HTTPException(400,
                    f"Cross-file decision points to '{decision.sheet}' sheet "
                    f"but that source file is not loaded. Re-upload all 3 "
                    f"files (main + alt UoM + long text) and try again.")

            source_rows = source_loaded.rows
            unique_row_indexes = list(dict.fromkeys(decision.error_row_indexes))

            # Plan edits against the alt-UoM or long-text source rows.
            xfile_edits: list[dict] = []
            for row_idx in unique_row_indexes:
                if not (0 <= row_idx < len(source_rows)):
                    continue
                src_row = source_rows[row_idx]
                current = src_row.values.get(sap_field)
                current_s = "" if current is None else str(current)
                if current_s in rules:
                    xfile_edits.append({
                        "row_idx": row_idx,
                        "sap_field": sap_field,
                        "old_value": current_s,
                        "new_value": rules[current_s],
                        "matnr": src_row.values.get("MATNR") or "",
                    })

            total = len(xfile_edits)

            if not req.confirm:
                per_rule: dict[str, int] = {}
                for e in xfile_edits:
                    per_rule[e["old_value"]] = per_rule.get(e["old_value"], 0) + 1
                return {
                    "total_cells": total,
                    "per_rule": [{"find": f, "count": c} for f, c in per_rule.items()],
                    "needs_confirmation": total > GROUP_REPLACE_CONFIRM_THRESHOLD,
                }

            if total > GROUP_REPLACE_CONFIRM_THRESHOLD and not req.reason.strip():
                raise HTTPException(400,
                    f"A reason is required when replacing more than "
                    f"{GROUP_REPLACE_CONFIRM_THRESHOLD} cells.")

            if total == 0:
                return {"replaced_count": 0, "new_error_count": len(state["errors"])}

            # Apply edits to the source rows in-place. The merger's
            # MergedMaterial.alt_uoms / .longtexts hold REFERENCES to
            # these same LoadedRow objects, so editing here also flows
            # through to the LTMC export which reads from material.alt_uoms.
            for e in xfile_edits:
                src_row = source_rows[e["row_idx"]]
                new_val = e["new_value"]
                if isinstance(new_val, str) and new_val.strip() == "":
                    src_row.values.pop(e["sap_field"], None)
                else:
                    src_row.values[e["sap_field"]] = new_val

                entry = AuditEntry(
                    timestamp=dt_module.datetime.now().isoformat(timespec="seconds"),
                    user=user.get("display_name", user["username"]),
                    action="group_replace",
                    rule_id=decision.rule_id,
                    sheet=decision.sheet,
                    affected_count=1,
                    reason=req.reason,
                    details={
                        "row_idx": e["row_idx"],
                        "sap_field": e["sap_field"],
                        "column_label": decision.column_label,
                        "old_value": e["old_value"],
                        "new_value": str(new_val),
                        "matnr": e["matnr"],
                        "group_replace_batch": True,
                        "cross_file_source": decision.sheet,  # tag for changes_summary
                        "find": e["old_value"],
                        "replace": str(new_val),
                    },
                )
                state["audit_log"].append(entry)

            # Re-validate to refresh decisions (the cross-file rule
            # should now find fewer mismatches, possibly zero).
            from services.mm_validator import validate_mm as _mm_validate
            errors = _mm_validate(
                materials, _get_mm_catalogs(), main_fields,
                merged_result=bundle.get("merged"),
                alt_uom_rows=(alt_loaded.rows if alt_loaded else None),
                longtext_rows=(lt_loaded.rows if lt_loaded else None),
                friendly_labels=getattr(bundle.get("main_loaded"), "header_labels", None),
            )
            accepted = state.get("accepted_errors") or set()
            if accepted:
                errors = [err for err in errors
                          if (err.rule_id, err.sheet, err.col_idx, err.row_idx) not in accepted]
            state["errors"] = errors
            state["decisions"] = group_errors(errors)
            state["dirty"] = True

            return {
                "replaced_count": total,
                "new_error_count": len(errors),
                "summary": f"Updated {total} cell(s) in {decision.sheet} source data.",
            }

        PLANT_SCOPED = {"WERKS", "LGORT", "LGPRO", "DISPO", "FEVOR", "BWKEY"}

        # Edit plan: (row_idx, plant_row_index, old, new). plant_row_index
        # is the index into material.plant_rows; -1 means "main" (non-
        # plant-scoped). When sap_field is plant-scoped we edit every
        # plant_row whose value matches a rule; when plant-independent
        # we edit material.main only.
        #
        # Dedupe row_indexes first: for plant-scoped rules the same
        # row_idx appears in error_row_indexes once per plant_row (11×
        # for a Peenya material at 11 plants). We want to visit each
        # material once and enumerate its plant_rows ourselves, otherwise
        # we'd visit plant_rows 11×11 = 121 times.
        unique_row_indexes = list(dict.fromkeys(decision.error_row_indexes))
        mm_edits: list[dict] = []
        for row_idx in unique_row_indexes:
            if not (0 <= row_idx < len(materials)):
                continue
            mat = materials[row_idx]
            if sap_field in PLANT_SCOPED:
                for p_idx, plant_row in enumerate(mat.plant_rows):
                    current = plant_row.get(sap_field)
                    current_s = "" if current is None else str(current)
                    if current_s in rules:
                        mm_edits.append({
                            "row_idx": row_idx,
                            "plant_row_idx": p_idx,
                            "sap_field": sap_field,
                            "old_value": current_s,
                            "new_value": rules[current_s],
                            "werks": plant_row.get("WERKS") or "",
                        })
            else:
                current = mat.main.get(sap_field)
                current_s = "" if current is None else str(current)
                if current_s in rules:
                    mm_edits.append({
                        "row_idx": row_idx,
                        "plant_row_idx": -1,   # edit main
                        "sap_field": sap_field,
                        "old_value": current_s,
                        "new_value": rules[current_s],
                        "werks": "",
                    })

        total = len(mm_edits)

        if len(rules) > GROUP_REPLACE_MAX_DISTINCT_RULES:
            raise HTTPException(400,
                f"{len(rules)} distinct replacements exceed the safety cap of "
                f"{GROUP_REPLACE_MAX_DISTINCT_RULES}. Apply in smaller batches.")
        if total > GROUP_REPLACE_HARD_CAP_CELLS:
            raise HTTPException(400,
                f"{total:,} cells exceed the absolute ceiling of "
                f"{GROUP_REPLACE_HARD_CAP_CELLS:,}. Use Fix Individually.")

        if not req.confirm:
            per_rule = {}
            for e in mm_edits:
                per_rule[e["old_value"]] = per_rule.get(e["old_value"], 0) + 1
            return {
                "total_cells": total,
                "per_rule": [{"find": f, "count": c} for f, c in per_rule.items()],
                "needs_confirmation": total > GROUP_REPLACE_CONFIRM_THRESHOLD,
            }

        if total > GROUP_REPLACE_CONFIRM_THRESHOLD and not req.reason.strip():
            raise HTTPException(400,
                f"A reason is required when replacing more than "
                f"{GROUP_REPLACE_CONFIRM_THRESHOLD} cells.")

        if total == 0:
            return {"replaced_count": 0, "new_error_count": len(state["errors"])}

        # Apply: mutate plant_rows / main in place
        replaced = []
        for e in mm_edits:
            mat = materials[e["row_idx"]]
            target = mat.main if e["plant_row_idx"] == -1 else mat.plant_rows[e["plant_row_idx"]]
            new_val = e["new_value"]
            if isinstance(new_val, str) and new_val.strip() == "":
                target.values.pop(e["sap_field"], None)
            else:
                target.values[e["sap_field"]] = new_val

            # Per-edit audit entry
            entry = AuditEntry(
                timestamp=dt_module.datetime.now().isoformat(timespec="seconds"),
                user=user.get("display_name", user["username"]),
                action="group_replace",
                rule_id=decision.rule_id,
                sheet="Materials",
                affected_count=1,
                reason=req.reason,
                details={
                    "row_idx": e["row_idx"],
                    "sap_field": e["sap_field"],
                    "column_label": decision.column_label,
                    "old_value": e["old_value"],
                    "new_value": str(new_val),
                    "werks": e["werks"],
                    "group_replace_batch": True,
                    "find": e["old_value"],
                    "replace": str(new_val),
                },
            )
            state["audit_log"].append(entry)
            replaced.append(e)

        # Revalidate
        from services.mm_validator import validate_mm as _mm_validate
        bundle = state.get("mm_bundle") or {}
        alt_loaded = bundle.get("alt_loaded")
        lt_loaded = bundle.get("lt_loaded")
        errors = _mm_validate(
            materials, _get_mm_catalogs(), main_fields,
            merged_result=bundle.get("merged"),
            alt_uom_rows=(alt_loaded.rows if alt_loaded else None),
            longtext_rows=(lt_loaded.rows if lt_loaded else None),
            friendly_labels=getattr(bundle.get("main_loaded"), "header_labels", None),
        )
        accepted = state.get("accepted_errors") or set()
        if accepted:
            errors = [err for err in errors
                      if (err.rule_id, err.sheet, err.col_idx, err.row_idx) not in accepted]
        state["errors"] = errors
        state["decisions"] = group_errors(errors)
        state["dirty"] = True

        # Consolidated audit log. We embed `_snapshots` (the same shape SD's
        # decision_* actions use) so the changes_summary endpoint can expand
        # this single DB row into N per-cell entries — matches the SD flow,
        # so SMEs see one row per replaced cell in the changes summary table
        # instead of nothing (which is what was happening before v55: the
        # consolidated row was logged but no per-cell snapshot data, so the
        # changes_summary's `decision_*` branch would skip it).
        #
        # Why store snapshots in the DB log instead of pulling from
        # state["audit_log"]: the in-memory log is volatile (lost on
        # service restart) and not joined into changes_summary's data
        # source. The DB log is durable and is the one truthful list of
        # what happened. SD's decision flow has done this since day one;
        # MM was the outlier.
        snapshots = {str(e["row_idx"]): e["old_value"] for e in replaced}
        audit_log.log(
            user=user, action="decision_group_replace",
            file_id=state["file_id"], filename=state["filename"],
            module="MM", sheet="Materials", rule_id=decision.rule_id,
            affected_count=len(replaced),
            reason=req.reason[:200] if req.reason else "",
            details={
                "rule_id": decision.rule_id,
                "sap_field": sap_field,
                "column": decision.column_label,
                "_col_idx": decision.col_idx,
                "_snapshots": snapshots,
                # The find→replace map. When multiple distinct bad values
                # were replaced in one go (e.g. "FEU1" and "" both replaced
                # with "FGEP"), changes_summary needs the per-row old→new
                # mapping which lives in _snapshots (per row_idx) plus
                # _new_values_by_row below.
                "_new_values_by_row": {
                    str(e["row_idx"]): str(e["new_value"]) for e in replaced
                },
                "_rules": rules,
                "total_replacements": len(rules),
                "reason": req.reason[:200] if req.reason else "",
            },
        )
        return {
            "replaced_count": len(replaced),
            "new_error_count": len(errors),
            "pending_decisions": len(state["decisions"]),
            "distinct_values_replaced": len(rules),
        }

    # ── SD branch (original logic) ─────────────────────────────────────────
    wb = state["workbook"]
    sheet = wb.sheets.get(decision.sheet)

    # Collect (row_idx, col_idx, old, new) for every cell that matches one
    # of the replacement rules. Dedup per cell — multi-rule flags would
    # otherwise apply twice.
    edits: list[dict] = []
    seen_cells: set[tuple[int, int]] = set()
    for row_idx in decision.error_row_indexes:
        key = (row_idx, decision.col_idx)
        if key in seen_cells:
            continue
        seen_cells.add(key)
        if not (0 <= row_idx < len(sheet.data_rows)):
            continue
        current = sheet.data_rows[row_idx].get(decision.col_idx)
        current_s = "" if current is None else str(current)
        if current_s in rules:
            edits.append({
                "row_idx": row_idx,
                "col_idx": decision.col_idx,
                "old_value": current_s,
                "new_value": rules[current_s],
            })

    total = len(edits)

    if len(rules) > GROUP_REPLACE_MAX_DISTINCT_RULES:
        raise HTTPException(400,
            f"{len(rules)} distinct replacements exceed the safety cap of "
            f"{GROUP_REPLACE_MAX_DISTINCT_RULES}. Apply in smaller batches.")

    if total > GROUP_REPLACE_HARD_CAP_CELLS:
        raise HTTPException(400,
            f"{total:,} cells exceed the absolute ceiling of "
            f"{GROUP_REPLACE_HARD_CAP_CELLS:,}. Use Fix Individually.")

    if not req.confirm:
        # Dry-run: tell the client how many cells would change per rule
        per_rule = {}
        for e in edits:
            per_rule[e["old_value"]] = per_rule.get(e["old_value"], 0) + 1
        return {
            "total_cells": total,
            "per_rule": [{"find": f, "count": c} for f, c in per_rule.items()],
            "needs_confirmation": total > GROUP_REPLACE_CONFIRM_THRESHOLD,
        }

    if total > GROUP_REPLACE_CONFIRM_THRESHOLD and not req.reason.strip():
        raise HTTPException(400,
            f"A reason is required when replacing more than "
            f"{GROUP_REPLACE_CONFIRM_THRESHOLD} cells.")

    if total == 0:
        return {"replaced_count": 0, "new_error_count": len(state["errors"])}

    # Apply each edit via the same per-cell path single-cell edits use.
    # This preserves per-cell undo via Recent Activity.
    replaced = []
    for e in edits:
        try:
            entry = apply_single_edit(
                wb, decision.sheet, e["row_idx"], e["col_idx"], e["new_value"],
                user=user["display_name"],
            )
            entry.reason = req.reason
            entry.rule_id = decision.rule_id
            entry.details["group_replace_batch"] = True
            entry.details["find"] = e["old_value"]
            entry.details["replace"] = e["new_value"]
            entry.details["column_label"] = decision.column_label
            state["audit_log"].append(entry)
            replaced.append(e)
        except Exception as exc:
            print(f"[group_replace] edit failed for {e}: {exc}")

    errors = _revalidate_and_rebuild(state)
    state["dirty"] = True

    # One consolidated audit trail entry for the operation.
    audit_log.log(
        user=user, action="group_replace",
        file_id=state["file_id"], filename=state["filename"],
        module=state["module"], sheet=decision.sheet,
        affected_count=len(replaced), reason=req.reason,
        details={
            "decision_id": decision_id,
            "rule_id": decision.rule_id,
            "column_label": decision.column_label,
            "distinct_values_replaced": len(rules),
            "replaced_count": len(replaced),
        },
    )

    return {
        "replaced_count": len(replaced),
        "distinct_values_replaced": len(rules),
        "total_cells": total,
        "new_error_count": len(errors),
    }


@app.delete("/api/session/records/{sheet_name}/{row_idx}")
def delete_row(sheet_name: str, row_idx: int, reason: str = "",
               session: str | None = Cookie(None), user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")
    try:
        entry = delete_single_row(state["workbook"], sheet_name, row_idx, user=user["display_name"])
        state["audit_log"].append(entry)
        errors = _revalidate_and_rebuild(state)

        if state["file_id"]:
            xml_bytes = write_xml(state["workbook"])
            repo.save_working_copy(state["file_id"], xml_bytes)

        audit_log.log(
            user=user, action="delete_row",
            file_id=state["file_id"], filename=state["filename"],
            module=state["module"], sheet=sheet_name,
            affected_count=1, reason=reason,
            details={"row_idx": row_idx, "snapshot": entry.details.get("snapshot", {})},
        )

        return {"success": True, "audit_entry": entry.as_dict(), "new_error_count": len(errors)}
    except Exception as e:
        raise HTTPException(500, str(e))


class BulkDeleteRequest(BaseModel):
    sheet: str
    row_indexes: list[int]
    reason: str = ""


@app.post("/api/session/records/bulk_delete")
def bulk_delete_rows(req: BulkDeleteRequest, session: str | None = Cookie(None),
                     user: dict = Depends(current_user)):
    """Delete multiple rows at once. Always requires a business reason."""
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")
    if not req.reason or not req.reason.strip():
        raise HTTPException(400, "A business reason is required for bulk deletion")
    if not req.row_indexes:
        raise HTTPException(400, "No rows selected")

    try:
        # Delete in descending order so earlier indexes don't shift
        deleted = 0
        for row_idx in sorted(set(req.row_indexes), reverse=True):
            try:
                entry = delete_single_row(state["workbook"], req.sheet, row_idx, user=user["display_name"])
                state["audit_log"].append(entry)
                deleted += 1
            except Exception:
                continue

        errors = _revalidate_and_rebuild(state)

        if state["file_id"]:
            xml_bytes = write_xml(state["workbook"])
            repo.save_working_copy(state["file_id"], xml_bytes)

        audit_log.log(
            user=user, action="bulk_delete_rows",
            file_id=state["file_id"], filename=state["filename"],
            module=state["module"], sheet=req.sheet,
            affected_count=deleted, reason=req.reason,
            details={"row_indexes": req.row_indexes},
        )

        return {"success": True, "deleted_count": deleted, "new_error_count": len(errors)}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/session/audit")
def get_audit(session: str | None = Cookie(None), user: dict = Depends(current_user)):
    """Recent activity for the currently-open file.
    Now reads from Postgres audit table by file_id, so it survives file switches & restarts."""
    state = _get_session_state(session)
    file_id = state.get("file_id") if state else None
    if not file_id:
        # No file open — return in-memory session log as fallback
        log = state["audit_log"] if state else []
        return {
            "log": [{**e.as_dict(), "audit_index": idx} for idx, e in enumerate(log)][::-1]
        }
    # File is open — pull persistent audit for this file
    entries = audit_log.list_for_file(file_id)
    # Also include in-session entries with audit_index for undo
    session_log = state.get("audit_log") or []
    session_indexed = [{**e.as_dict(), "audit_index": idx} for idx, e in enumerate(session_log)][::-1]
    # Return both: persistent audit is display-only; session log has audit_index for undo
    return {
        "log": session_indexed,
        "persistent": entries,
    }


def _release_mm_export_cache(state: dict) -> None:
    """Clean up any prior MM export's temp dir."""
    import shutil
    cache = state.get("mm_export_cache")
    if not cache:
        return
    try:
        shutil.rmtree(cache["dir"], ignore_errors=True)
    except Exception:
        pass
    state["mm_export_cache"] = None


@app.post("/api/session/export_ltmc")
def export_ltmc_manifest(session: str | None = Cookie(None),
                          user: dict = Depends(current_user)):
    """Generate LTMC-upload-ready SpreadsheetML XML for an MM session.

    Two-phase flow matching PP's:
      1. POST /api/session/export_ltmc      → manifest with chunk list
      2. GET  /api/session/export_ltmc/chunk/{n} → stream one xml file

    Why chunked: SAP LTMC has a ~100 MB practical per-file limit and
    Excel struggles with SpreadsheetML 2003 files much over that. For
    a Healthium-scale 10k-material file the single-file output came
    in at 113.9 MB. We split at MATNR boundaries (a material's rows
    on Basic Data + Plant Data + Storage Locations + Inspection Setup
    Data + Valuation Data all stay in the same chunk) so each chunk
    is independently importable into SAP.

    Single-chunk-friendly: the splitter runs first; if all materials
    fit in one chunk under the cap, we return a 1-element manifest
    and the frontend offers a single download. Multi-chunk files are
    presented as a numbered series.

    Only emits sheets that are populated for the data on hand. For a
    typical FG material at Healthium, that's 11 sheets:
      Basic Data, Additional Descriptions, Alternative Units of Measure,
      Class Data, Distribution Chains, Point of Sale Data,
      Tax Classification, Plant Data, Storage Locations,
      Inspection Setup Data, Valuation Data
    """
    import tempfile, time
    from pathlib import Path
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")
    if not state.get("mm_bundle"):
        raise HTTPException(400, "LTMC export is only available for MM sessions")

    from services.ltmc_generator import generate_ltmc_xml
    from services.mm_splitter import split_into_chunks, SAFE_MAX_BYTES
    bundle = state["mm_bundle"]
    merged = bundle["merged"]
    materials = merged.materials

    # Pull BWKEY map from the bundled MM_KDS catalogs.
    catalogs = _get_mm_catalogs()
    bwkey_map = catalogs.get("bwkey_by_werks", {}) if isinstance(catalogs, dict) else {}

    # v62: session-level LTMC overrides from "Set LTMC default value"
    # Decisions. Pass-through to the generator so user-entered values
    # for missing-from-source LTMC-mandatory fields land in the export.
    from services.mm_ltmc_overrides import get_overrides
    ltmc_overrides = dict(get_overrides(state))

    base = (state["filename"] or "materials").rsplit(".", 1)[0]

    # Clean prior export if any, then prep new temp dir.
    _release_mm_export_cache(state)
    tmp_dir = Path(tempfile.mkdtemp(prefix="mm_export_"))
    chunks_meta: list[dict] = []

    # Try as a single file first — if it fits under the safe cap we
    # ship one file and skip the per-chunk template scaffolding overhead
    # (~1.2 MB per chunk). Same pattern PP uses since v55.6.
    single_bytes = generate_ltmc_xml(materials, bwkey_map=bwkey_map,
                                     ltmc_overrides=ltmc_overrides)
    if len(single_bytes) <= SAFE_MAX_BYTES:
        filename = f"{base}_LTMC.xml"
        (tmp_dir / filename).write_bytes(single_bytes)
        chunks_meta.append({
            "index": 0,
            "filename": filename,
            "size_bytes": len(single_bytes),
            "material_count": len(materials),
        })
    else:
        # Bin-pack into chunks. The splitter sorts by source order
        # within each chunk so the output preserves the SME's input
        # ordering — easier to find a known MATNR for review.
        chunks = split_into_chunks(merged)
        n = len(chunks)
        width = len(str(n))
        for ch in chunks:
            ch_bytes = generate_ltmc_xml(ch.materials, bwkey_map=bwkey_map,
                                         ltmc_overrides=ltmc_overrides)
            fname = f"{base}_LTMC_part{ch.chunk_index+1:0{width}d}of{n}.xml"
            (tmp_dir / fname).write_bytes(ch_bytes)
            chunks_meta.append({
                "index": ch.chunk_index,
                "filename": fname,
                "size_bytes": len(ch_bytes),
                "material_count": len(ch.materials),
            })

    state["mm_export_cache"] = {
        "dir": tmp_dir,
        "chunks": chunks_meta,
        "kind": "mm",
        "created_at": time.time(),
    }

    return {
        "chunk_count": len(chunks_meta),
        "single_file": len(chunks_meta) == 1,
        "material_count": len(materials),
        "chunks": chunks_meta,
    }


@app.get("/api/session/export_ltmc/chunk/{chunk_index}")
def export_ltmc_chunk(chunk_index: int,
                       session: str | None = Cookie(None),
                       user: dict = Depends(current_user)):
    """Stream one cached MM export chunk back to the client.

    The chunk must have been prepared by a prior POST to
    /api/session/export_ltmc — this endpoint only reads from the
    in-session temp dir. If the cache is missing or the requested
    chunk is out of range, returns 410 Gone (re-run the POST).
    """
    from pathlib import Path
    from fastapi.responses import StreamingResponse

    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    cache = state.get("mm_export_cache")
    if not cache:
        raise HTTPException(410, "Export cache missing — re-run the export.")

    if chunk_index < 0 or chunk_index >= len(cache["chunks"]):
        raise HTTPException(404, f"Chunk index {chunk_index} out of range")

    meta = cache["chunks"][chunk_index]
    file_path = Path(cache["dir"]) / meta["filename"]
    if not file_path.exists():
        raise HTTPException(410, "Cached chunk missing on disk — re-run the export.")

    def _stream():
        with open(file_path, "rb") as f:
            while True:
                blk = f.read(1024 * 1024)
                if not blk:
                    break
                yield blk

    return StreamingResponse(
        _stream(),
        media_type="application/xml",
        headers={
            "Content-Disposition": f'attachment; filename="{meta["filename"]}"',
            "Content-Length": str(meta["size_bytes"]),
            "X-Chunk-Index": str(chunk_index),
            "X-Chunk-Count": str(len(cache["chunks"])),
        },
    )


@app.get("/api/session/export_review_xlsx")
def export_review_xlsx(session: str | None = Cookie(None),
                        user: dict = Depends(current_user)):
    """Generate a colored xlsx review file for the current MM session.

    Distinct from /export_ltmc: that one produces SAP-import-ready
    SpreadsheetML 2003 XML (no colors). This one produces a modern
    .xlsx with cells fill-colored to indicate validation errors so
    SMEs can review their data visually in Excel and figure out
    which cells need fixing before they re-export the clean LTMC file.

    The output file is intentionally named "*_Review.xlsx" with a
    prominent banner row reading "DO NOT UPLOAD TO SAP" so it can't
    be confused with the LTMC import file. SAP's importer doesn't
    accept .xlsx anyway (LTMC needs SpreadsheetML 2003 XML), so even
    if an SME tried to upload it directly, SAP would reject it with
    a format error rather than silently importing colored cells.

    MM-only for now. PP/Routing have a similar use case but their
    multi-sheet structure makes a single review xlsx more complex
    to lay out — separate effort.
    """
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")
    if not state.get("mm_bundle"):
        raise HTTPException(
            400,
            "Review xlsx export is only available for MM sessions today. "
            "PP/Routing review export is on the roadmap."
        )

    from services.mm_review_export import build_mm_review_xlsx
    bundle = state["mm_bundle"]
    base = (state["filename"] or "materials").rsplit(".", 1)[0]

    # The bundle holds a MergeResult on .merged. Errors live on state.
    # main_loaded provides the SAP field order + friendly labels.
    # v57: also pass alt/lt loaded files so the review xlsx can render
    # cross-file errors on dedicated sheets ("Alternate Units" + "Long Text")
    # for SME inspection in the same file.
    fname, xlsx_bytes = build_mm_review_xlsx(
        merged=bundle["merged"],
        errors=state["errors"],
        main_loaded=bundle["main_loaded"],
        base_filename=base,
        alt_loaded=bundle.get("alt_loaded"),
        lt_loaded=bundle.get("lt_loaded"),
    )

    return FastResponse(
        content=xlsx_bytes,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/session/export")
def export_xml(session: str | None = Cookie(None), user: dict = Depends(current_user)):
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    # ── MM branch: emit a cleaned XLSX bundle ─────────────────────────────
    # Phase 1 export (NOT full LTMC MATMAS XML — Phase 3 pending).
    # This produces a single xlsx with one sheet per input file containing
    # the corrected values. LTMC field renaming (e.g. MATNR → PRODUCT for
    # S/4HANA) is NOT applied — we emit source-format so SMEs can visually
    # diff against the original and, if needed, feed the cleaned xlsx into
    # the client's existing LTMC process.
    #
    # If the session has zero errors (all decisions resolved), the output
    # is "ready to hand off". If errors remain, the output is "work in
    # progress" — SMEs can still download and review before committing.
    if state.get("mm_bundle"):
        import io
        from openpyxl import Workbook
        bundle = state["mm_bundle"]
        merged = bundle["merged"]
        main_loaded = bundle["main_loaded"]
        main_fields = main_loaded.sap_fields

        wb = Workbook()

        # === Sheet 1: Main material master — one row per (MATNR × plant) ===
        # Column set = original main-file fields + any new fields added during
        # the session (SPRAS, ALAND, WAERS, CURTP, BWKEY get added by the
        # mandatory_with_default rules' fixes). Without unioning, the export
        # would silently drop the very fields the SME just filled in.
        ws_main = wb.active
        ws_main.title = "Main"
        added_fields: list[str] = []
        original_set = set(main_fields)
        seen_added = set(original_set)
        for mat in merged.materials:
            for plant_row in mat.plant_rows:
                for k in plant_row.values.keys():
                    if k not in seen_added:
                        seen_added.add(k)
                        added_fields.append(k)
        export_fields = list(main_fields) + added_fields  # original order preserved
        # Row 1: human-friendly column labels (copy from loader if we have them).
        # For added fields we don't have a label so use the SAP code itself.
        original_labels = main_loaded.header_labels if hasattr(main_loaded, "header_labels") else main_fields
        export_labels = list(original_labels) + list(added_fields)
        ws_main.append(export_labels)
        # Row 2: SAP codes — what the loader used as keys (and what the export reader expects)
        ws_main.append(export_fields)
        # Row 3+: cleaned data, one row per plant_row per material
        for mat in merged.materials:
            for plant_row in mat.plant_rows:
                row_out = []
                for sap_field in export_fields:
                    v = plant_row.get(sap_field)
                    row_out.append("" if v is None else v)
                ws_main.append(row_out)

        # === Sheet 2: Alternate Units of Measure ===
        ws_alt = wb.create_sheet("AlternateUnits")
        alt_loaded = bundle.get("alt_loaded")
        if alt_loaded:
            alt_fields = alt_loaded.sap_fields
            ws_alt.append(list(alt_loaded.header_labels if hasattr(alt_loaded, "header_labels") else alt_fields))
            ws_alt.append(list(alt_fields))
            for mat in merged.materials:
                for au in mat.alt_uoms:
                    row_out = []
                    for sap_field in alt_fields:
                        v = au.get(sap_field)
                        row_out.append("" if v is None else v)
                    ws_alt.append(row_out)
        else:
            ws_alt.append(["(no alternate units of measure file uploaded)"])

        # === Sheet 3: Long Text ===
        ws_lt = wb.create_sheet("LongText")
        lt_loaded = bundle.get("lt_loaded")
        if lt_loaded:
            lt_fields = lt_loaded.sap_fields
            ws_lt.append(list(lt_loaded.header_labels if hasattr(lt_loaded, "header_labels") else lt_fields))
            ws_lt.append(list(lt_fields))
            for mat in merged.materials:
                for lt in mat.longtexts:
                    row_out = []
                    for sap_field in lt_fields:
                        v = lt.get(sap_field)
                        row_out.append("" if v is None else v)
                    ws_lt.append(row_out)
        else:
            ws_lt.append(["(no long-text file uploaded)"])

        # === Sheet 4: Change Log ===
        # Every edit captured in this session, so SMEs can audit what
        # the tool did vs the original input.
        ws_log = wb.create_sheet("ChangeLog")
        ws_log.append(["Timestamp", "User", "Action", "Rule", "MATNR", "WERKS",
                       "SAP Field", "Old Value", "New Value", "Reason"])
        audit_entries = state.get("audit_log") or []
        for entry in audit_entries:
            # AuditEntry is a dataclass; convert to dict
            if hasattr(entry, "timestamp"):
                d = {
                    "timestamp": entry.timestamp,
                    "user": entry.user,
                    "action": entry.action,
                    "rule_id": entry.rule_id,
                    "details": entry.details or {},
                    "reason": entry.reason or "",
                }
            else:
                d = entry
            details = d.get("details") or {}
            # Find the MATNR for this row_idx
            row_idx = details.get("row_idx")
            matnr = ""
            if row_idx is not None and 0 <= row_idx < len(merged.materials):
                matnr = merged.materials[row_idx].matnr
            ws_log.append([
                d.get("timestamp", ""),
                d.get("user", ""),
                d.get("action", ""),
                d.get("rule_id", ""),
                matnr,
                details.get("werks", ""),
                details.get("sap_field", details.get("column_label", "")),
                details.get("old_value", ""),
                details.get("new_value", ""),
                d.get("reason", ""),
            ])

        # Summary sheet as the first visible tab
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.append(["Master Data Validator — MM Export Summary"])
        ws_summary.append([])
        ws_summary.append(["File ID", state.get("file_id", "")])
        ws_summary.append(["Original filename", state.get("filename", "")])
        ws_summary.append(["Exported by", user.get("display_name", user["username"])])
        ws_summary.append(["Exported at", dt_module.datetime.now().isoformat(timespec="seconds")])
        ws_summary.append([])
        ws_summary.append(["Materials", len(merged.materials)])
        ws_summary.append(["Plant rows", merged.summary.get("plant_row_count", 0)])
        ws_summary.append(["Remaining errors", len(state.get("errors", []))])
        ws_summary.append(["Pending decisions", len(state.get("decisions", []))])
        ws_summary.append(["Edits applied this session", len(audit_entries)])
        if len(state.get("errors", [])) > 0:
            ws_summary.append([])
            ws_summary.append(["⚠ WARNING", "This file still has unresolved errors. Review before LTMC upload."])

        # Serialize
        out = io.BytesIO()
        wb.save(out)
        xlsx_bytes = out.getvalue()

        if state.get("dirty"):
            state["dirty"] = False  # best-effort clear

        fname = (state["filename"] or "materials").rsplit(".", 1)[0] + "_cleaned.xlsx"
        return FastResponse(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    # ── SD branch: original XML export ────────────────────────────────────
    xml_bytes = write_xml(state["workbook"])
    # Opportunistically flush the working copy since we're serializing anyway.
    if state.get("dirty") and state.get("file_id"):
        repo.save_working_copy(state["file_id"], xml_bytes)
        state["dirty"] = False
    fname = state["filename"].replace(".xml", "_clean.xml")
    if not fname.endswith(".xml"):
        fname += "_clean.xml"

    # Warn (don't fail) if SD output exceeds the 95 MB safe threshold.
    # MM and PP both chunk at this size; SD doesn't yet because no SD
    # customer master we've seen approaches it. If the file IS this big,
    # SAP and Excel may both have trouble — surface that via a custom
    # response header so the frontend can show a warning toast.
    SD_SAFE_MAX = 95 * 1024 * 1024
    headers = {"Content-Disposition": f'attachment; filename="{fname}"'}
    if len(xml_bytes) > SD_SAFE_MAX:
        headers["X-SD-Size-Warning"] = (
            f"output is {len(xml_bytes)//(1024*1024)} MB — over the 95 MB "
            f"comfortable limit for SAP LTMC import and Excel viewing. "
            f"If SAP rejects the import, contact support — chunked SD "
            f"export isn't yet implemented."
        )
    return FastResponse(
        content=xml_bytes,
        media_type="application/xml",
        headers=headers,
    )


@app.get("/api/session/errors")
def all_errors(
    limit: int = 10000,
    offset: int = 0,
    session: str | None = Cookie(None),
    user: dict = Depends(current_user),
):
    """Full error list for grid view.

    Paginated to handle PP/Routing files which can produce hundreds of
    thousands of errors. Default limit 10000 covers SD/MM cases (which
    rarely exceed a few thousand) without OOMing the response on a
    large Routing file. Frontend can request additional pages via
    ?offset=N&limit=M.
    """
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")
    all_errs = state["errors"]
    total = len(all_errs)
    # Clamp limit to a sane ceiling
    limit = max(1, min(limit, 50000))
    offset = max(0, offset)
    page = all_errs[offset:offset + limit]
    return {
        "errors": [e.as_dict() for e in page],
        "total": total,
        "offset": offset,
        "limit": limit,
        "truncated": total > offset + limit,
    }


@app.get("/api/session/changes_summary")
def changes_summary(session: str | None = Cookie(None), user: dict = Depends(current_user)):
    """Flat list of every cell-level change made during validation.
    Each entry: sheet, row, column, old value, new value, rule, reason, user, time.
    SMEs can review before download; IT can attach as audit evidence."""
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    entries = audit_log.list_for_file(state["file_id"]) if state["file_id"] else []
    changes = []

    # Build a friendly-label resolver from the current session's loaded
    # files. The audit log might contain raw SAP codes for `column_label`
    # if the entry was created before v59 (when Decision objects didn't
    # yet carry friendly labels) — for those entries we look up the
    # friendly name from header_labels and a small canonical-name fallback
    # for LTMC standard fields not in customer source columns (BWKEY,
    # ALAND, WAERS, CURTP, etc).
    #
    # Resolver: SAP code → friendly label, with three tiers:
    #   1) Source upload row 1 (customer's chosen friendly labels)
    #   2) Canonical SAP DDIC short text for LTMC standard fields
    #   3) Fallback to the SAP code itself (never blank)
    sap_to_friendly: dict[str, str] = {}
    if state.get("mm_bundle"):
        bundle = state["mm_bundle"]
        main_loaded = bundle.get("main_loaded")
        if main_loaded:
            sap_fields = list(getattr(main_loaded, "sap_fields", []) or [])
            labels = list(getattr(main_loaded, "header_labels", []) or [])
            for code, lbl in zip(sap_fields, labels):
                if lbl and code:
                    sap_to_friendly[code] = str(lbl)
    # Canonical names for LTMC-standard fields the customer's template
    # often doesn't include (we add them during export). These are short
    # SAP DDIC labels — what an SAP user reading the LTMC sheet would
    # see in their MM module.
    _LTMC_CANONICAL = {
        "BWKEY":  "Valuation Area",
        "ALAND":  "Country",
        "WAERS":  "Currency",
        "CURTP":  "Currency Type",
        "SPRAS":  "Language",
        "PEINH":  "Price Unit",
        "BWTAR":  "Valuation Type",
    }
    for code, name in _LTMC_CANONICAL.items():
        sap_to_friendly.setdefault(code, name)

    def _friendly(label_in_audit: str | None, sap_field: str | None = None) -> str:
        """Resolve to the friendliest name we have, in priority order:
        the audit log's column_label if it's already friendly (not equal
        to the SAP code) → header_labels lookup → canonical → SAP code."""
        # If the audit log already has a friendly label (different from
        # the SAP code), trust it — don't second-guess what the SME saw
        # at the time the change was made.
        if label_in_audit and sap_field and label_in_audit != sap_field:
            return label_in_audit
        # If the audit log column_label IS the SAP code (legacy v58 entry),
        # try to upgrade via the resolver.
        candidate = sap_field or label_in_audit or ""
        return sap_to_friendly.get(candidate, label_in_audit or candidate)

    for e in entries:
        action = e.get("action", "")
        details = e.get("details") or {}
        if isinstance(details, str):
            import json as _json
            try:
                details = _json.loads(details)
            except Exception:
                details = {}

        # Single cell edit
        if action == "edit_cell":
            raw_label = details.get("column_label", "")
            sap_field = details.get("sap_field", "") or raw_label
            changes.append({
                "type": "edit",
                "sheet": e.get("sheet") or "",
                "row_idx": details.get("row_idx"),
                "xml_row": (details.get("row_idx") or 0) + 9 if details.get("row_idx") is not None else None,
                "col_idx": details.get("col_idx"),
                "column_label": _friendly(raw_label, sap_field),
                "old_value": details.get("old_value", ""),
                "new_value": details.get("new_value", ""),
                "rule": "",
                "reason": e.get("reason") or "",
                "user": e.get("display_name") or e.get("username") or "",
                "timestamp": e.get("timestamp"),
            })
        # Bulk decision actions
        elif action.startswith("decision_"):
            act = action.replace("decision_", "")
            snapshots = details.get("_snapshots") or {}
            col_idx = details.get("_col_idx")
            # `column` (legacy "column_label") and `sap_field` are both stored
            # in audit log details. v59-onwards both should be friendly + raw,
            # but for legacy entries (created under v58) `column` may equal
            # the SAP code. The resolver upgrades it to friendly when possible.
            raw_label = details.get("column", "")
            sap_field = details.get("sap_field") or details.get("column_label_sap") or raw_label
            column_label = _friendly(raw_label, sap_field)
            new_val = details.get("new_value", details.get("value", ""))
            # For accept_all / review, snapshots captured but no new value — note it as "kept"
            if act in ("accept_all", "ignore", "review"):
                for row_str, old in snapshots.items():
                    row_idx = int(row_str)
                    changes.append({
                        "type": "accept",
                        "sheet": e.get("sheet") or "",
                        "row_idx": row_idx,
                        "xml_row": row_idx + 9,
                        "col_idx": col_idx,
                        "column_label": column_label,
                        "old_value": old,
                        "new_value": old,  # unchanged
                        "rule": e.get("rule_id") or "",
                        "reason": e.get("reason") or "",
                        "user": e.get("display_name") or e.get("username") or "",
                        "timestamp": e.get("timestamp"),
                    })
            elif act in ("replace_with", "fill_with", "clear_all", "set_urp", "truncate_all"):
                for row_str, old in snapshots.items():
                    row_idx = int(row_str)
                    changes.append({
                        "type": act,
                        "sheet": e.get("sheet") or "",
                        "row_idx": row_idx,
                        "xml_row": row_idx + 9,
                        "col_idx": col_idx,
                        "column_label": column_label,
                        "old_value": old,
                        "new_value": "" if act == "clear_all" else (
                            "URP" if act == "set_urp" else new_val
                        ),
                        "rule": e.get("rule_id") or "",
                        "reason": e.get("reason") or "",
                        "user": e.get("display_name") or e.get("username") or "",
                        "timestamp": e.get("timestamp"),
                    })
            elif act == "group_replace":
                # MM bulk Group & Replace: one DB row covers N per-cell edits.
                # `_snapshots` carries the old value per row; `_new_values_by_row`
                # carries the new value per row (different bad values can map
                # to different replacements within one operation, so a single
                # `new_value` field at the parent level isn't enough).
                # The fall-back order is: per-row map → rule lookup → details.new_value.
                new_values_by_row = details.get("_new_values_by_row") or {}
                rules_map = details.get("_rules") or {}
                for row_str, old in snapshots.items():
                    row_idx = int(row_str)
                    new_for_row = (
                        new_values_by_row.get(row_str)
                        or rules_map.get(old)
                        or new_val
                        or ""
                    )
                    changes.append({
                        "type": "replace_with",   # render as replace_with for UI
                        "sheet": e.get("sheet") or "",
                        "row_idx": row_idx,
                        "xml_row": row_idx + 9,
                        "col_idx": col_idx,
                        "column_label": column_label,
                        "old_value": old,
                        "new_value": new_for_row,
                        "rule": e.get("rule_id") or "",
                        "reason": e.get("reason") or "",
                        "user": e.get("display_name") or e.get("username") or "",
                        "timestamp": e.get("timestamp"),
                    })
            elif act in ("delete_duplicates", "delete_rows"):
                changes.append({
                    "type": "delete",
                    "sheet": e.get("sheet") or "",
                    "row_idx": None,
                    "xml_row": None,
                    "col_idx": None,
                    "column_label": "(full row)",
                    "old_value": f"{e.get('affected_count') or 0} rows deleted",
                    "new_value": "(deleted)",
                    "rule": e.get("rule_id") or "",
                    "reason": e.get("reason") or "",
                    "user": e.get("display_name") or e.get("username") or "",
                    "timestamp": e.get("timestamp"),
                })
        elif action == "delete_row":
            changes.append({
                "type": "delete",
                "sheet": e.get("sheet") or "",
                "row_idx": details.get("row_idx"),
                "xml_row": (details.get("row_idx") or 0) + 9 if details.get("row_idx") is not None else None,
                "col_idx": None,
                "column_label": "(full row)",
                "old_value": "(row deleted)",
                "new_value": "(deleted)",
                "rule": "",
                "reason": e.get("reason") or "",
                "user": e.get("display_name") or e.get("username") or "",
                "timestamp": e.get("timestamp"),
            })
        elif action == "bulk_delete_rows":
            affected = e.get("affected_count") or 0
            changes.append({
                "type": "delete",
                "sheet": e.get("sheet") or "",
                "row_idx": None,
                "xml_row": None,
                "col_idx": None,
                "column_label": "(full rows)",
                "old_value": f"{affected} rows",
                "new_value": "(deleted)",
                "rule": "",
                "reason": e.get("reason") or "",
                "user": e.get("display_name") or e.get("username") or "",
                "timestamp": e.get("timestamp"),
            })

    # Newest last — so the "story" reads forwards
    changes.sort(key=lambda c: str(c.get("timestamp") or ""))

    # ── v63: enrich each change with sap_field so the Changes Summary
    # table can show both the friendly label AND the SAP code in the
    # Column cell (e.g. "Price ctrl (VPRSV)" instead of just "Price ctrl").
    # Two sources, in order:
    #   1. Audit entry's details already carry `sap_field` (set_ltmc_default
    #      since v62, group_replace since v55). Use it directly.
    #   2. Reverse-map column_label → sap_field by scanning the loaded
    #      main-file's header_labels/sap_fields parallel lists.
    # If neither resolves, leave sap_field empty — frontend just shows
    # the friendly label alone (graceful degradation).
    sap_by_label: dict[str, str] = {}
    if state.get("mm_bundle"):
        ml = state["mm_bundle"].get("main_loaded")
        if ml is not None:
            labels = list(getattr(ml, "header_labels", []) or [])
            sap_fields_list = list(getattr(ml, "sap_fields", []) or [])
            for code, lbl in zip(sap_fields_list, labels):
                if lbl and lbl != code:
                    sap_by_label[lbl] = code

    for c in changes:
        if c.get("sap_field"):
            continue   # already populated by the caller (group_replace etc.)
        col_label = c.get("column_label") or ""
        if col_label in sap_by_label:
            c["sap_field"] = sap_by_label[col_label]
        elif col_label and col_label.isupper() and col_label.replace("_", "").isalnum():
            # Looks like a SAP code already (all-uppercase alphanumeric, e.g.
            # 'BWKEY', 'STEUC'). Mirror to sap_field so the UI doesn't show
            # a redundant "(BWKEY)" suffix next to a "BWKEY" label.
            c["sap_field"] = col_label
        else:
            c["sap_field"] = ""

    # Summary stats
    total = len(changes)
    by_type: dict[str, int] = {}
    by_sheet: dict[str, int] = {}
    for c in changes:
        by_type[c["type"]] = by_type.get(c["type"], 0) + 1
        by_sheet[c["sheet"]] = by_sheet.get(c["sheet"], 0) + 1

    return {
        "changes": changes,
        "total": total,
        "by_type": by_type,
        "by_sheet": by_sheet,
    }


@app.get("/api/session/changes_summary.csv")
def changes_summary_csv(session: str | None = Cookie(None), user: dict = Depends(current_user)):
    """Downloadable CSV of changes summary — for audit / compliance hand-off."""
    state = _get_session_state(session)
    if not _session_loaded(state):
        raise HTTPException(400, "No file loaded")

    data = changes_summary(session=session, user=user)
    changes = data["changes"]

    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "Type", "Sheet", "Row", "Column", "Old Value", "New Value",
        "Rule", "Reason", "User", "Timestamp"
    ])
    for c in changes:
        w.writerow([
            c["type"], c["sheet"], c.get("xml_row") or "", c["column_label"],
            c["old_value"], c["new_value"], c["rule"], c["reason"],
            c["user"], c["timestamp"] or "",
        ])

    csv_bytes = buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens cleanly
    fname = (state["filename"] or "changes").replace(".xml", "") + "_changes.csv"
    return FastResponse(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ──────────────────────────────────────────────────────────────────────────
# Reference data
# ──────────────────────────────────────────────────────────────────────────

# ISO 3166-1 alpha-2 for the ~50 countries most commonly on SAP customer
# masters. Keeping the list short so the dropdown is scannable; users can
# still type unlisted codes into the freeform text fallback.
COUNTRY_CODES = [
    ("IN", "India"),            ("US", "United States"),    ("GB", "United Kingdom"),
    ("DE", "Germany"),          ("FR", "France"),           ("IT", "Italy"),
    ("ES", "Spain"),            ("NL", "Netherlands"),      ("BE", "Belgium"),
    ("CH", "Switzerland"),      ("AT", "Austria"),          ("SE", "Sweden"),
    ("NO", "Norway"),           ("DK", "Denmark"),          ("FI", "Finland"),
    ("IE", "Ireland"),          ("PT", "Portugal"),         ("GR", "Greece"),
    ("PL", "Poland"),           ("CZ", "Czech Republic"),   ("HU", "Hungary"),
    ("RO", "Romania"),          ("TR", "Turkey"),           ("RU", "Russia"),
    ("CN", "China"),            ("JP", "Japan"),            ("KR", "South Korea"),
    ("TW", "Taiwan"),           ("HK", "Hong Kong"),        ("SG", "Singapore"),
    ("MY", "Malaysia"),         ("TH", "Thailand"),         ("VN", "Vietnam"),
    ("PH", "Philippines"),      ("ID", "Indonesia"),        ("AU", "Australia"),
    ("NZ", "New Zealand"),      ("CA", "Canada"),           ("MX", "Mexico"),
    ("BR", "Brazil"),           ("AR", "Argentina"),        ("CL", "Chile"),
    ("CO", "Colombia"),         ("ZA", "South Africa"),     ("EG", "Egypt"),
    ("SA", "Saudi Arabia"),     ("AE", "UAE"),              ("IL", "Israel"),
    ("QA", "Qatar"),            ("KW", "Kuwait"),           ("BH", "Bahrain"),
    ("OM", "Oman"),             ("LK", "Sri Lanka"),        ("BD", "Bangladesh"),
    ("NP", "Nepal"),            ("PK", "Pakistan"),
]


@app.get("/api/reference/states/IN")
def states_india():
    """SAP GST state codes for India — for the State dropdown editor."""
    from services.validator import INDIA_STATE_CODE_NAMES
    # Return as list for stable ordering. Skip historical "Andhra Pradesh (old)"
    # entry 28; keep 37 "Andhra Pradesh (new)" as primary. Users can still type it.
    items = []
    for code, name in INDIA_STATE_CODE_NAMES.items():
        if code == "28":
            continue  # deprecated
        label = name.replace(" (new)", "")
        items.append({"code": code, "name": label})
    items.sort(key=lambda x: int(x["code"]))
    return {"items": items}


@app.get("/api/reference/countries")
def countries():
    """ISO 3166-1 country codes for the Country dropdown editor."""
    return {
        "items": [{"code": c, "name": n} for c, n in COUNTRY_CODES],
    }


# ──────────────────────────────────────────────────────────────────────────
# Frontend
# ──────────────────────────────────────────────────────────────────────────

@app.get("/")
def index():
    return HTMLResponse((FRONTEND_DIR / "index.html").read_text(encoding="utf-8"))


@app.get("/api/health")
def health():
    """Health + operational metrics. Safe to poll every 10-30s for monitoring."""
    ok, info = db_svc.ping()
    loaded_workbooks = sum(1 for s in WORKING.values() if s.get("workbook"))
    total_sessions = len(WORKING)

    try:
        import psutil, os as _os
        rss_mb = psutil.Process(_os.getpid()).memory_info().rss / (1024 * 1024)
    except ImportError:
        rss_mb = None

    # MDV_ENV is set by NSSM (via setup_production.py) to "production" or
    # "staging". Frontend uses this to show a visual banner. Default to
    # "production" so deployments without the env var (e.g. local dev)
    # don't show a scary banner.
    env_name = os.environ.get("MDV_ENV", "production")

    return {
        "status": "ok" if ok else "degraded",
        "env": env_name,
        "database": "up" if ok else "down",
        "db_info": info[:80] if ok else info,
        "loaded_workbooks": loaded_workbooks,
        "total_sessions": total_sessions,
        "max_workbooks": MAX_LOADED_WORKBOOKS,
        "session_ttl_seconds": SESSION_TTL_SECONDS,
        "process_rss_mb": round(rss_mb, 1) if rss_mb else None,
        "storage_path": str(repo.STORAGE_ROOT),
    }


@app.post("/api/session/close")
def close_file(session: str | None = Cookie(None), user: dict = Depends(current_user)):
    """Voluntarily release the loaded workbook so RAM is freed immediately.
    If there are unflushed edits, we persist to disk BEFORE freeing memory
    so no edits are lost."""
    state = _get_session_state(session)
    if not state:
        return {"closed": False}
    had_file = state.get("file_id")
    # Flush dirty working copy before releasing the workbook
    if state.get("dirty") and state.get("workbook") and state.get("file_id"):
        try:
            xml_bytes = write_xml(state["workbook"])
            repo.save_working_copy(state["file_id"], xml_bytes)
        except Exception as e:
            print(f"[close_file] flush failed for {state['file_id']}: {e}")
    _free_workbook_in_session(state)
    return {"closed": True, "had_file": bool(had_file)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
